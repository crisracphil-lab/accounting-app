from decimal import Decimal
import importlib
from pathlib import Path
import sys

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


@pytest.fixture
def fresh_db(tmp_path, monkeypatch):
    monkeypatch.setenv("ACCOUNTING_DB", str(tmp_path / "test.db"))
    from app import db as db_module
    importlib.reload(db_module)
    db_module.init_db()
    yield db_module


def _make_fs(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BS"
    ws.cell(row=5, column=1).value = "A/C No."
    ws.cell(row=5, column=2).value = "Acct Title"
    ws.cell(row=5, column=3).value = "Current"
    ws.cell(row=5, column=4).value = "Prior"
    ws.cell(row=5, column=5).value = "Inc./Dec."
    ws.cell(row=5, column=6).value = "Reasons/Remarks"
    ws.cell(row=6, column=1).value = 5000
    ws.cell(row=6, column=2).value = "Operating Expense"
    ws.cell(row=6, column=3).value = 250000
    ws.cell(row=6, column=4).value = 100000
    ws.cell(row=6, column=5).value = 150000
    wb.save(path)


def _make_sl(path: Path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PER SYSTEM"
    ws.cell(row=5, column=1).value = "A/C No."
    ws.cell(row=5, column=2).value = "Acct Title"
    ws.cell(row=6, column=1).value = 5000
    ws.cell(row=6, column=2).value = "Operating Expense"
    ws.cell(row=6, column=12).value = 100000
    ws.cell(row=7, column=10).value = 150000
    ws.cell(row=7, column=11).value = 0
    ws.cell(row=7, column=12).value = 250000
    wb.save(path)


def test_closing_from_financial_statement_and_subsidiary_ledger_generates_remarks(tmp_path, fresh_db):
    fs_path = tmp_path / "financial_statements.xlsx"
    sl_path = tmp_path / "subsidiary_ledger.xlsx"
    _make_fs(fs_path)
    _make_sl(sl_path)

    from app.services.closing_books import save_closing_run_from_files
    run_id = save_closing_run_from_files(fs_path, sl_path, "Test close", Decimal("100000"))

    with fresh_db.db() as conn:
        run = conn.execute("SELECT * FROM closing_runs WHERE id=?", (run_id,)).fetchone()
        row = conn.execute("SELECT * FROM closing_account_changes WHERE run_id=? AND account_code='5000'", (run_id,)).fetchone()

    assert run["financial_statement_filename"].endswith("financial_statements.xlsx")
    assert run["subsidiary_ledger_filename"].endswith("subsidiary_ledger.xlsx")
    assert row["flagged"] == 1
    assert row["sheet"] == "BS"
    assert row["fs_inc_dec"] == "150000"
    assert row["explanation"].startswith("Increase because")
    assert "subsidiary ledger movement" in row["explanation"]
    assert "Subsidiary ledger 5000" in row["basis"]
    assert row["ledger_rows"] == 1


def test_closing_export_writes_remarks_into_uploaded_fs_last_column(tmp_path, fresh_db):
    fs_path = tmp_path / "financial_statements.xlsx"
    sl_path = tmp_path / "subsidiary_ledger.xlsx"
    out_path = tmp_path / "fs_with_remarks.xlsx"
    _make_fs(fs_path)
    _make_sl(sl_path)

    from app.services.closing_books import (
        save_closing_run_from_files,
        export_closing_financial_statement_with_remarks,
    )
    run_id = save_closing_run_from_files(fs_path, sl_path, "Test close", Decimal("100000"))
    export_closing_financial_statement_with_remarks(run_id, out_path)

    import openpyxl
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb["BS"]
    assert ws.cell(row=5, column=6).value == "Reasons/Remarks"
    remark = ws.cell(row=6, column=6).value
    assert remark is not None
    assert remark.startswith("Increase because")
    assert "subsidiary ledger movement" in remark.lower()
