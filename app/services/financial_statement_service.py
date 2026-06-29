"""FS ingestion, editing, export."""
from __future__ import annotations
import json
from pathlib import Path
from typing import List, Optional
from app.db import db, log_action
from app.parsers.financial_statement import parse_fs, FSWorkbook, FSRow


def ingest_fs(file_path, period_label: str) -> int:
    fs = parse_fs(file_path)
    with db() as conn:
        conn.execute("BEGIN")
        try:
            cur = conn.execute(
                "INSERT INTO fs_uploads (period_label, company_name, source_filename, "
                "is_columns_json, bs_columns_json) VALUES (?, ?, ?, ?, ?)",
                (period_label, fs.company_name, str(file_path),
                 json.dumps(fs.is_columns), json.dumps(fs.bs_columns)))
            up_id = cur.lastrowid
            for r in fs.is_rows + fs.bs_rows:
                conn.execute(
                    "INSERT INTO fs_rows (fs_upload_id, sheet, row_number, account_code, "
                    "account_title, columns_json, inc_dec, remarks) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (up_id, r.sheet, r.row_number, r.account_code, r.account_title,
                     json.dumps(r.as_dict()["columns"]),
                     str(r.inc_dec) if r.inc_dec is not None else None,
                     r.remarks))
            log_action(conn, "ingest", "fs_upload", up_id,
                       {"period": period_label, "is_rows": len(fs.is_rows),
                        "bs_rows": len(fs.bs_rows)})
            conn.execute("COMMIT")
            return up_id
        except Exception:
            conn.execute("ROLLBACK")
            raise


def update_remarks(row_id: int, remarks: str) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE fs_rows SET remarks = ?, edited_at = datetime('now') WHERE id = ?",
            (remarks or None, row_id))
        log_action(conn, "edit_remarks", "fs_row", row_id,
                   {"remarks_len": len(remarks)})


def export_fs_xlsx(fs_upload_id: int, output_path: Path) -> Path:
    """Export FS as a single .xlsx with IS + BS sheets and the edited remarks."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    with db() as conn:
        up = conn.execute("SELECT * FROM fs_uploads WHERE id = ?", (fs_upload_id,)).fetchone()
        if up is None:
            raise ValueError(f"FS upload {fs_upload_id} not found")
        rows = conn.execute(
            "SELECT * FROM fs_rows WHERE fs_upload_id = ? ORDER BY sheet, row_number",
            (fs_upload_id,)).fetchall()

    is_cols = json.loads(up["is_columns_json"] or "[]")
    bs_cols = json.loads(up["bs_columns_json"] or "[]")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_label, headers in (("IS", is_cols), ("BS", bs_cols)):
        if not headers:
            continue
        ws = wb.create_sheet(sheet_label if sheet_label != "IS" else "IS3")
        ws.cell(1, 1, up["company_name"] or "Financial Statement").font = Font(bold=True, size=14)
        ws.cell(2, 1, ("Income Statement" if sheet_label == "IS" else "Balance Sheet")).font = Font(bold=True)
        ws.cell(3, 1, f"Period: {up['period_label']}")
        ws.cell(4, 1, f"Exported with edited Reasons/Remarks").font = Font(italic=True, color="777777")
        # Header row
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(5, c, h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        sheet_rows = [r for r in rows if r["sheet"] == sheet_label]
        for ri, r in enumerate(sheet_rows, start=6):
            cols = json.loads(r["columns_json"] or "{}")
            ws.cell(ri, 1, r["account_code"] or "")
            ws.cell(ri, 2, r["account_title"])
            for c, label in enumerate(headers[2:-2 if sheet_label in ("IS", "BS") else None],
                                      start=3):
                v = cols.get(label)
                if v is None:
                    continue
                try:
                    ws.cell(ri, c, float(v))
                except (ValueError, TypeError):
                    ws.cell(ri, c, str(v))
            # Inc/Dec column (second-to-last)
            if r["inc_dec"]:
                try:
                    ws.cell(ri, len(headers) - 1, float(r["inc_dec"]))
                except (TypeError, ValueError) as exc:
                    raise ValueError(f"Invalid FS increase/decrease value for row {r['id']}: {r['inc_dec']}") from exc
            # Remarks column (last)
            if r["remarks"]:
                cell = ws.cell(ri, len(headers), r["remarks"])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Auto-size columns roughly
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 18
        if headers:
            ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 35
            ws.column_dimensions[openpyxl.utils.get_column_letter(len(headers))].width = 60

    wb.save(output_path)
    return output_path
