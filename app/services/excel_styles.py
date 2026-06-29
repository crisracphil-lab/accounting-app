"""
Corporate Excel styling for BookPoint exports.

Usage:
    from app.services.excel_styles import build_workbook_header, style_data_rows, corp_sheet

Colours follow RAC Phil Corp's navy/gold identity.
"""
from __future__ import annotations
import io
from typing import Iterable, Sequence, Optional
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── Brand colours ────────────────────────────────────────────────────────────
NAVY        = "1E3A5F"   # dark navy — header fill
NAVY_LIGHT  = "2E5090"   # mid navy — sub-header fill
GOLD        = "C9A84C"   # accent line / highlight
WHITE       = "FFFFFF"
ROW_ALT     = "EBF3FB"   # alternate row light blue
TOTAL_FILL  = "D9E8F5"   # summary / total row

# ── Common styles ─────────────────────────────────────────────────────────────
_THIN_SIDE   = Side(style="thin",   color="B0B8C5")
_MEDIUM_SIDE = Side(style="medium", color="1E3A5F")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE,
                      top=_THIN_SIDE,  bottom=_THIN_SIDE)
_BOTTOM_MEDIUM = Border(bottom=Side(style="medium", color=NAVY))

HDR_FONT  = Font(bold=True, color=WHITE, size=10, name="Calibri")
CORP_FONT = Font(bold=True, color=NAVY,  size=12, name="Calibri")
SUB_FONT  = Font(bold=True, color=WHITE, size=9,  name="Calibri")
DATA_FONT = Font(size=9, name="Calibri")
BOLD_FONT = Font(bold=True, size=9, name="Calibri")

HDR_FILL    = PatternFill("solid", fgColor=NAVY)
SUB_FILL    = PatternFill("solid", fgColor=NAVY_LIGHT)
TOTAL_FILL_ = PatternFill("solid", fgColor=TOTAL_FILL)
ALT_FILL    = PatternFill("solid", fgColor=ROW_ALT)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=False)
WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

MONEY_FMT = '#,##0.00'
INT_FMT   = '#,##0'


# ── Helpers ───────────────────────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    return get_column_letter(n)


def _merge_and_write(ws, row: int, col_start: int, col_end: int,
                     value, font=None, fill=None, alignment=None):
    """Merge cells and write a value."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end,
    )
    cell = ws.cell(row, col_start, value)
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    return cell


def add_corp_header(ws, report_title: str, period: str = "", num_cols: int = 10):
    """
    Write a 3-row corporate header block at the top of the worksheet.
    Row 1 — Company name (navy bg, white bold)
    Row 2 — Report title  (navy bg, white bold)
    Row 3 — Period / As of (light navy bg, white)
    Returns the next available data row (4 if period supplied, else 3).
    """
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    _merge_and_write(ws, 1, 1, num_cols,
                     "RAC PHIL CORP",
                     font=Font(bold=True, color=WHITE, size=13, name="Calibri"),
                     fill=HDR_FILL, alignment=CENTER)

    _merge_and_write(ws, 2, 1, num_cols,
                     report_title.upper(),
                     font=Font(bold=True, color=WHITE, size=10, name="Calibri"),
                     fill=HDR_FILL, alignment=CENTER)

    if period:
        ws.row_dimensions[3].height = 15
        _merge_and_write(ws, 3, 1, num_cols,
                         period,
                         font=Font(bold=False, color=WHITE, size=9, name="Calibri"),
                         fill=SUB_FILL, alignment=CENTER)
        return 4
    return 3


def write_column_headers(ws, row: int, headers: Sequence[str]):
    """Write bold navy column headers at the given row."""
    ws.row_dimensions[row].height = 16
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = _THIN_BORDER
    ws.freeze_panes = ws.cell(row + 1, 1)


def style_data_rows(ws, data_row_start: int, data_row_end: int,
                    num_cols: int, money_cols: Iterable[int] = (),
                    int_cols: Iterable[int] = ()):
    """
    Apply alternating fill, font, borders and number formats to data rows.
    money_cols / int_cols are 1-based column indices.
    """
    money_cols = set(money_cols)
    int_cols   = set(int_cols)

    for r in range(data_row_start, data_row_end + 1):
        fill = ALT_FILL if r % 2 == 0 else None
        for c in range(1, num_cols + 1):
            cell = ws.cell(r, c)
            cell.font   = DATA_FONT
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            if c in money_cols:
                cell.number_format = MONEY_FMT
                cell.alignment     = RIGHT
            elif c in int_cols:
                cell.number_format = INT_FMT
                cell.alignment     = RIGHT
            else:
                cell.alignment = LEFT


def add_total_row(ws, row: int, num_cols: int,
                  totals: dict[int, float]):
    """Write a shaded totals row. totals = {col_index: value}."""
    ws.row_dimensions[row].height = 15
    for c in range(1, num_cols + 1):
        cell = ws.cell(row, c)
        cell.fill   = TOTAL_FILL_
        cell.font   = BOLD_FONT
        cell.border = Border(
            top=Side(style="medium", color=NAVY),
            bottom=Side(style="medium", color=NAVY),
            left=_THIN_SIDE, right=_THIN_SIDE,
        )
        if c in totals:
            cell.value         = totals[c]
            cell.number_format = MONEY_FMT
            cell.alignment     = RIGHT
        elif c == 1:
            cell.value     = "TOTAL"
            cell.alignment = LEFT


def auto_col_width(ws, min_w: int = 8, max_w: int = 45):
    """Set column widths based on content. Skips merged/broken cells safely."""
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = min_w
        for cell in col:
            try:
                # Skip header rows (merged banner rows) and cells without a value attr
                if cell.row < 3 or not hasattr(cell, "value"):
                    continue
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_w)
        except Exception:
            pass


def finalize_workbook(wb: Workbook) -> bytes:
    """Save workbook to bytes and return."""
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
