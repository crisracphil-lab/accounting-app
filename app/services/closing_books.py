"""
Closing-of-books worksheet generator.
Supports two closing workflows:
1. Legacy subsidiary-ledger only analysis.
2. Financial statement + subsidiary ledger analysis. The financial statement
   supplies the report account rows and inc./dec. values; the subsidiary ledger
   supplies the basis used to generate remarks/reasoning for material movements.
Parser failures and invalid uploads raise exceptions; no generated sample rows are used.
"""
from __future__ import annotations
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
import re
import shutil
import uuid
from collections import defaultdict
from typing import List, Optional
from app.db import db, log_action
from app.parsers.subsidiary_ledger import (
    parse_subsidiary_ledger, SubsidiaryLedger, AccountActivity,
)
from app.parsers.financial_statement import parse_fs, FSRow
DEFAULT_THRESHOLD = Decimal("100000.00")


def _account_threshold(account_code: str) -> Decimal:
    """Return materiality threshold per account code range.

    Expenses (6xxx / 7xxx)  10,000  highest scrutiny
    Revenue (4xxx)          50,000
    Cost of Sales (5xxx)   100,000
    Cash (110x)            200,000  large movements expected
    Current Assets (11xx)  100,000
    Other CA (12xx)         50,000
    Prepayments (125x)      30,000
    Fixed Assets (15x)      20,000
    Other Assets (18x)     100,000
    Liabilities (2xx)       50,000
    Default                100,000
    """
    code = str(account_code).strip().rstrip(".0").lstrip("0") or "0"
    if code.startswith("7"):
        return Decimal("10000")
    if code.startswith("6"):
        return Decimal("10000")
    if code.startswith("56") or code.startswith("5"):
        return Decimal("100000")
    if code.startswith("4"):
        return Decimal("50000")
    if code.startswith("1102") or code.startswith("110"):
        return Decimal("200000")
    if code.startswith("11"):
        return Decimal("100000")
    if code.startswith("125"):
        return Decimal("30000")
    if code.startswith("12"):
        return Decimal("50000")
    if code.startswith("15"):
        return Decimal("20000")
    if code.startswith("18"):
        return Decimal("100000")
    if code.startswith("2"):
        return Decimal("50000")
    return DEFAULT_THRESHOLD
@dataclass
class FlaggedAccount:
    account_code: str
    account_title: str
    opening_balance: Decimal
    closing_balance: Decimal
    net_change: Decimal
    flagged: bool
def _money(v: Decimal) -> str:
    return f"{v:,.2f}"
def _movement_word(amount: Decimal) -> str:
    return "increase" if amount >= 0 else "decrease"
def _build_basis(activity: AccountActivity, fs_inc_dec: Optional[Decimal] = None) -> str:
    basis = (
        f"Subsidiary ledger {activity.code} - {activity.title}: "
        f"opening balance {_money(activity.opening_balance)}, "
        f"period debit {_money(activity.period_debit)}, "
        f"period credit {_money(activity.period_credit)}, "
        f"closing balance {_money(activity.closing_balance)}, "
        f"net movement {_money(activity.net_change)}, "
        f"posting rows {activity.rows}."
    )
    if fs_inc_dec is not None:
        variance = fs_inc_dec - activity.net_change
        if abs(variance) > Decimal("0.05"):
            basis += f" FS inc./dec. differs from SL movement by {_money(variance)}; review mapping or report formula."
    return basis
def _normalize_driver_text(text: str) -> str:
    text = re.sub(r"\s+", " ", (text or "").strip())
    text = re.sub(r"-BC$", "", text, flags=re.I).strip()
    upper = text.upper()
    if not upper:
        return ""
    if "MBC" in upper or "MANILA BROADCAST" in upper:
        return "payment to Manila Broadcasting for radio ads"
    if "MCCANN" in upper:
        return "payment to McCann"
    if "PAYMENT TO PGI" in upper or re.search(r"\bTO PGI\b", upper):
        return "payment to PGI"
    if "TO FUND PKO" in upper or "FUND PKO" in upper:
        return "fund transfer to PKO"
    if "PAYMENT SOLUTIONS" in upper:
        return "cash-in from payment solutions"
    if "PAGCOR" in upper and "AUDIT" in upper:
        return "PAGCOR audit fee payments"
    if "PAGCOR" in upper and "SHARE" in upper:
        return "PAGCOR share payments"
    if "PAGCOR" in upper:
        return "PAGCOR payments"
    if "BIR" in upper or "WITHHOLDING TAX" in upper:
        return "BIR tax payment"
    if "GCASH" in upper:
        return "GCash transactions"
    if "MAYA" in upper:
        return "Maya transactions"
    cleaned = upper
    for prefix in (
        "TO RECORD ", "TO CLOSE ", "FULL PAYMENT ", "PAYMENT FOR ",
        "PAYMENT TO ", "LKW ", "DF ", "RAC PHIL CORP ",
    ):
        cleaned = cleaned.replace(prefix, "")
    cleaned = re.sub(r"\bFEB(RUARY)?\b\.?\s*\d{0,4}", "", cleaned)
    cleaned = re.sub(r"\bJAN(UARY)?\b\.?\s*\d{0,4}", "", cleaned)
    cleaned = re.sub(r"\b20\d{2}\b", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -")
    if not cleaned:
        cleaned = upper
    return cleaned[:1].lower() + cleaned[1:].lower()
def _join_drivers(drivers: list) -> str:
    """Accept List[str] or List[(str, Decimal)] and produce readable remark text.

    Amounts shown only when >= 50,000 (independently material).
    """
    pairs = []
    seen: set = set()
    for item in drivers:
        if isinstance(item, tuple):
            text, amt = item[0], item[1]
        else:
            text, amt = str(item), None
        text = text.strip().rstrip(".")
        if not text or text in seen:
            continue
        seen.add(text)
        pairs.append((text, amt))
    if not pairs:
        return "subsidiary ledger movement"

    def _fmt(text: str, amt) -> str:
        if amt is not None and amt >= Decimal("50000"):
            return f"{text} amounting to ₱{amt:,.2f}"
        return text

    parts = [_fmt(t, a) for t, a in pairs]
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]} and {parts[1]}"


def _activity_drivers(activity: AccountActivity, movement: Decimal) -> list:
    """Return at most 2 truly material (text, Decimal) driver pairs from the SL.

    Rules:
    1. Only postings on the dominant side (debit if movement>0, credit if <0).
    2. Aggregate amounts by normalised description.
    3. A driver is included only if it meets EITHER:
       >= 30% of abs(movement)  -- dominant single cause
       >= 50,000 absolute        -- independently material
    4. Return at most 2 drivers, largest first.
    5. If nothing passes the filter, return the single largest driver only.
    """
    postings = getattr(activity, "postings", []) or []
    if not postings:
        return []

    abs_movement = abs(movement)

    if movement < 0:
        side = "credit" if getattr(activity, "period_credit", 0) != 0 else "debit"
    elif movement > 0:
        side = "debit" if getattr(activity, "period_debit", 0) != 0 else "credit"
    else:
        return []

    totals: dict = {}
    for p in postings:
        p_side = getattr(p, "side", "")
        if p_side != side:
            continue
        raw_text = getattr(p, "text", "") or ""
        text = _normalize_driver_text(raw_text)
        if not text:
            continue
        if any(w in text.lower() for w in ("service charge", "bank charge", "monthly fee", "annual fee")):
            continue
        amt = abs(getattr(p, "amount", Decimal("0")) or Decimal("0"))
        totals[text] = totals.get(text, Decimal("0")) + amt

    if not totals:
        return []

    ranked = sorted(totals.items(), key=lambda x: x[1], reverse=True)

    MATERIAL_PCT = Decimal("0.30")
    MATERIAL_ABS = Decimal("50000")
    material = [
        (text, amt) for text, amt in ranked
        if (abs_movement > 0 and amt / abs_movement >= MATERIAL_PCT) or amt >= MATERIAL_ABS
    ]

    if material:
        return material[:2]
    return [ranked[0]]


def generate_closing_remark(account_title: str,
                            movement: Decimal,
                            activity: Optional[AccountActivity],
                            fs_inc_dec: Optional[Decimal] = None) -> str:
    """Generate direct closing remarks from real subsidiary-ledger drivers."""
    direction = "Increase" if movement >= 0 else "Decrease"
    if activity is None:
        return (
            f"{direction} because no matching subsidiary ledger account was found for {account_title}; "
            "review account mapping before finalizing."
        )
    drivers = _activity_drivers(activity, movement)
    reason = _join_drivers(drivers)
    remark = f"{direction} because of {reason}."
    if fs_inc_dec is not None:
        variance = fs_inc_dec - activity.net_change
        if abs(variance) > Decimal("0.05"):
            remark += f" Review mapping because FS inc./dec. differs from SL movement by {_money(variance)}."
    return remark
def evaluate_closing(ledger: SubsidiaryLedger,
                     threshold: Decimal = DEFAULT_THRESHOLD) -> List[FlaggedAccount]:
    out = []
    for a in ledger.accounts:
        acct_threshold = _account_threshold(a.code)
        out.append(FlaggedAccount(
            account_code=a.code,
            account_title=a.title,
            opening_balance=a.opening_balance,
            closing_balance=a.closing_balance,
            net_change=a.net_change,
            flagged=abs(a.net_change) >= acct_threshold,
        ))
    return out
def save_closing_run(file_path,
                     period_label: str,
                     threshold: Decimal = DEFAULT_THRESHOLD) -> int:
    """Legacy path: parse one subsidiary ledger and persist material changes."""
    ledger = parse_subsidiary_ledger(file_path)
    flagged = evaluate_closing(ledger, threshold)
    by_code = {a.code: a for a in ledger.accounts}
    with db() as conn:
        conn.execute("BEGIN")
        try:
            cur = conn.execute(
                "INSERT INTO closing_runs (period_label, threshold, source_filename, subsidiary_ledger_filename, status) "
                "VALUES (?, ?, ?, ?, 'in_progress')",
                (period_label, str(threshold), str(file_path), str(file_path)),
            )
            run_id = cur.lastrowid
            for f in flagged:
                activity = by_code.get(f.account_code)
                basis = _build_basis(activity) if activity else None
                explanation = generate_closing_remark(f.account_title, f.net_change, activity) if f.flagged else None
                conn.execute(
                    "INSERT INTO closing_account_changes "
                    "(run_id, account_code, account_title, opening_balance, "
                    " closing_balance, net_change, flagged, explanation, basis, "
                    " ledger_debit, ledger_credit, ledger_rows) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (run_id, f.account_code, f.account_title,
                     str(f.opening_balance), str(f.closing_balance),
                     str(f.net_change), 1 if f.flagged else 0, explanation, basis,
                     str(activity.period_debit) if activity else "0",
                     str(activity.period_credit) if activity else "0",
                     activity.rows if activity else 0),
                )
            log_action(conn, "closing_run", "closing_run", run_id, {
                "period": period_label,
                "threshold": str(threshold),
                "accounts": len(flagged),
                "flagged": sum(1 for f in flagged if f.flagged),
                "mode": "subsidiary_ledger_only",
            })
            conn.execute("COMMIT")
            return run_id
        except Exception:
            conn.execute("ROLLBACK")
            raise
def _closing_storage_dir() -> Path:
    data_dir = Path(__file__).resolve().parents[2] / "data" / "closing_files"
    data_dir.mkdir(parents=True, exist_ok=True)
    return data_dir
def _persist_financial_statement(source_path) -> Path:
    src = Path(source_path)
    if not src.exists():
        raise ValueError(f"Financial statement file is missing: {src}")
    suffix = src.suffix.lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError(f"Financial statement must be Excel .xls, .xlsx, or .xlsm, got {suffix}")
    dest = _closing_storage_dir() / f"fs_{uuid.uuid4().hex}{suffix}"
    shutil.copy2(src, dest)
    return dest
def _find_sheet_name(wb, sheet_label: str) -> str:
    names = wb.sheetnames
    if sheet_label == "IS":
        exact = next((n for n in names if n.upper() == "IS3"), None)
        if exact:
            return exact
        match = next((n for n in names if n.upper().startswith("IS")), None)
        if match:
            return match
    if sheet_label == "BS":
        match = next((n for n in names if n.upper() == "BS"), None)
        if match:
            return match
    raise ValueError(f"Unable to find worksheet for {sheet_label} in financial statement workbook")
def _remarks_column(ws) -> int:
    header_row = None
    for r in range(1, min(16, ws.max_row + 1)):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "A/C No" in v:
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"Unable to locate financial statement header row in worksheet {ws.title}")
    last_used = ws.max_column
    remarks_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        h_text = str(h or "").strip().lower()
        if "reason" in h_text or "remark" in h_text:
            remarks_col = c
    if remarks_col is not None:
        return remarks_col
    new_col = last_used + 1
    ws.cell(row=header_row, column=new_col, value="Remarks")
    return new_col
def _load_xls_as_openpyxl_workbook(source: Path):
    """Convert an uploaded .xls workbook to an openpyxl workbook for xlsx export."""
    import xlrd
    import openpyxl
    wb_xls = xlrd.open_workbook(source, formatting_info=False)
    out = openpyxl.Workbook()
    # Remove default sheet after creating real sheets.
    default = out.active
    for idx, sheet_name in enumerate(wb_xls.sheet_names()):
        ws_xls = wb_xls.sheet_by_name(sheet_name)
        ws = out.create_sheet(title=sheet_name[:31])
        for r in range(ws_xls.nrows):
            for c in range(ws_xls.ncols):
                value = ws_xls.cell_value(r, c)
                if value == "":
                    continue
                ws.cell(row=r + 1, column=c + 1, value=value)
        for c in range(1, max(2, ws_xls.ncols + 1)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 16
    out.remove(default)
    return out
def export_closing_financial_statement_with_remarks(run_id: int, output_path: Path) -> Path:
    """Copy the uploaded FS workbook and write generated closing remarks into its remarks/last column."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    with db() as conn:
        run = conn.execute("SELECT * FROM closing_runs WHERE id = ?", (run_id,)).fetchone()
        if run is None:
            raise ValueError(f"Closing run {run_id} not found")
        stored_path = run["financial_statement_stored_path"] if "financial_statement_stored_path" in run.keys() else None
        if not stored_path:
            raise ValueError("This closing run has no stored financial statement workbook to export")
        rows = conn.execute(
            "SELECT * FROM closing_account_changes WHERE run_id = ? AND sheet IN ('IS','BS') AND fs_row_number IS NOT NULL "
            "ORDER BY sheet, fs_row_number",
            (run_id,),
        ).fetchall()
    source = Path(stored_path)
    if not source.exists():
        raise ValueError(f"Stored financial statement workbook no longer exists: {source}")
    if source.suffix.lower() == ".xls":
        wb = _load_xls_as_openpyxl_workbook(source)
    elif source.suffix.lower() in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(source)
    else:
        raise ValueError(f"Expected .xls, .xlsx or .xlsm source workbook, got {source.suffix}")
    sheet_cache = {}
    remarks_cache = {}
    for r in rows:
        explanation = (r["explanation"] or "").strip()
        if not explanation:
            continue
        sheet_label = r["sheet"]
        if sheet_label not in sheet_cache:
            ws = wb[_find_sheet_name(wb, sheet_label)]
            sheet_cache[sheet_label] = ws
            remarks_cache[sheet_label] = _remarks_column(ws)
            header_cell = ws.cell(row=max(1, r["fs_row_number"] - 1), column=remarks_cache[sheet_label])
        ws = sheet_cache[sheet_label]
        col = remarks_cache[sheet_label]
        cell = ws.cell(row=int(r["fs_row_number"]), column=col, value=explanation)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r["flagged"]:
            cell.fill = PatternFill("solid", fgColor="ECFDF5")
        cell.font = Font(color="064E3B")
        if ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width < 55:
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 55
    wb.save(output_path)
    return output_path
def _combined_activity_for_fs_code(code: str, account_title: str, ledger_by_code: dict[str, AccountActivity]) -> Optional[AccountActivity]:
    """Return exact SL activity, or a parent-rollup of child SL accounts for FS group rows."""
    exact = ledger_by_code.get(code)
    if exact is not None:
        return exact
    children = [a for child_code, a in ledger_by_code.items()
                if child_code != code and child_code.startswith(code)]
    if not children:
        return None
    children.sort(key=lambda a: a.code)
    return AccountActivity(
        code=code,
        title=account_title,
        opening_balance=sum((a.opening_balance for a in children), Decimal("0")),
        period_debit=sum((a.period_debit for a in children), Decimal("0")),
        period_credit=sum((a.period_credit for a in children), Decimal("0")),
        closing_balance=sum((a.closing_balance for a in children), Decimal("0")),
        rows=sum(a.rows for a in children),
        postings=[p for a in children for p in getattr(a, "postings", [])],
    )
def save_closing_run_from_files(financial_statement_path,
                                subsidiary_ledger_path,
                                period_label: str,
                                threshold: Decimal = DEFAULT_THRESHOLD) -> int:
    """Parse FS + SL and persist material changes with generated remarks."""
    fs = parse_fs(financial_statement_path)
    stored_fs_path = _persist_financial_statement(financial_statement_path)
    ledger = parse_subsidiary_ledger(subsidiary_ledger_path)
    ledger_by_code = {a.code: a for a in ledger.accounts}
    fs_rows: List[FSRow] = [r for r in (fs.bs_rows + fs.is_rows) if r.account_code]
    if not fs_rows:
        raise ValueError("Financial statement contained no account rows with account codes")
    with db() as conn:
        conn.execute("BEGIN")
        try:
            cur = conn.execute(
                "INSERT INTO closing_runs "
                "(period_label, threshold, source_filename, financial_statement_filename, subsidiary_ledger_filename, financial_statement_stored_path, status) "
                "VALUES (?, ?, ?, ?, ?, ?, 'in_progress')",
                (period_label, str(threshold), str(financial_statement_path),
                 str(financial_statement_path), str(subsidiary_ledger_path), str(stored_fs_path)),
            )
            run_id = cur.lastrowid
            seen_codes = set()
            flagged_count = 0
            for row in fs_rows:
                code = str(row.account_code).strip()
                seen_codes.add(code)
                activity = _combined_activity_for_fs_code(code, row.account_title, ledger_by_code)
                fs_movement = row.inc_dec
                movement = fs_movement if fs_movement is not None else (activity.net_change if activity else Decimal("0"))
                opening = activity.opening_balance if activity else Decimal("0")
                closing = activity.closing_balance if activity else Decimal("0")
                flagged = abs(movement) >= threshold
                if flagged:
                    flagged_count += 1
                basis = _build_basis(activity, fs_movement) if activity else None
                explanation = generate_closing_remark(row.account_title, movement, activity, fs_movement) if flagged else (row.remarks or None)
                conn.execute(
                    "INSERT INTO closing_account_changes "
                    "(run_id, sheet, fs_row_number, account_code, account_title, opening_balance, "
                    " closing_balance, net_change, fs_inc_dec, flagged, explanation, basis, "
                    " ledger_debit, ledger_credit, ledger_rows) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (run_id, row.sheet, row.row_number, code, row.account_title,
                     str(opening), str(closing), str(movement),
                     str(fs_movement) if fs_movement is not None else None,
                     1 if flagged else 0, explanation, basis,
                     str(activity.period_debit) if activity else "0",
                     str(activity.period_credit) if activity else "0",
                     activity.rows if activity else 0),
                )
            # Include material SL accounts that are absent from the FS so the close does not hide real movements.
            for code, activity in ledger_by_code.items():
                if code in seen_codes or abs(activity.net_change) < threshold:
                    continue
                flagged_count += 1
                basis = _build_basis(activity)
                explanation = generate_closing_remark(activity.title, activity.net_change, activity)
                conn.execute(
                    "INSERT INTO closing_account_changes "
                    "(run_id, sheet, account_code, account_title, opening_balance, closing_balance, "
                    " net_change, flagged, explanation, basis, ledger_debit, ledger_credit, ledger_rows) "
                    "VALUES (?, 'SL', ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)",
                    (run_id, code, activity.title, str(activity.opening_balance),
                     str(activity.closing_balance), str(activity.net_change), explanation,
                     basis, str(activity.period_debit), str(activity.period_credit), activity.rows),
                )
            log_action(conn, "closing_run", "closing_run", run_id, {
                "period": period_label,
                "threshold": str(threshold),
                "fs_rows": len(fs_rows),
                "sl_accounts": len(ledger.accounts),
                "flagged": flagged_count,
                "mode": "financial_statement_plus_subsidiary_ledger",
            })
            conn.execute("COMMIT")
            return run_id
        except Exception:
            conn.execute("ROLLBACK")
            raise
def update_explanation(run_id: int, account_code: str, explanation: str) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE closing_account_changes SET explanation = ?, reviewed_at = datetime('now') "
            "WHERE run_id = ? AND account_code = ?",
            (explanation, run_id, account_code),
        )
        log_action(conn, "explain", "closing_account_change", run_id,
                   {"account_code": account_code, "explanation_len": len(explanation)})
def finalize_run(run_id: int) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE closing_runs SET status = 'completed', completed_at = datetime('now') "
            "WHERE id = ?", (run_id,))
        log_action(conn, "finalize", "closing_run", run_id, None)
def update_basis(run_id: int, account_code: str, basis: str) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE closing_account_changes SET basis = ?, reviewed_at = datetime('now') "
            "WHERE run_id = ? AND account_code = ?",
            (basis or None, run_id, account_code),
        )
        log_action(conn, "set_basis", "closing_account_change", run_id,
                   {"account_code": account_code})
def update_reviewer_notes(run_id: int, account_code: str, notes: str) -> None:
    with db() as conn:
        conn.execute(
            "UPDATE closing_account_changes SET reviewer_notes = ?, reviewed_at = datetime('now') "
            "WHERE run_id = ? AND account_code = ?",
            (notes or None, run_id, account_code),
        )
        log_action(conn, "set_reviewer_notes", "closing_account_change", run_id,
                   {"account_code": account_code})