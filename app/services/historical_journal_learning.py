from __future__ import annotations

import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, List, Optional

import openpyxl

STOP = {
    'the','and','for','with','from','payment','paid','debit','credit','cash','bank','account','entry','journal',
    'invoice','receipt','ref','reference','no','number','amount','php','peso','pesos','to','of','in','on','by','via'
}
CASH_WORDS = {'cash','bank','checking','savings','petty cash','cash in bank','unionbank','bdo','bpi','metrobank'}

@dataclass
class LearnedLine:
    description: str
    account_code: str
    account_title: str
    normal_side: str
    amount: Decimal
    keywords: str


def _norm(v: Any) -> str:
    text = str(v or '').strip().lower()
    text = re.sub(r'[\n\r\t_\-]+', ' ', text)
    text = re.sub(r'[^a-z0-9/ ]+', ' ', text)
    return ' '.join(text.split())


def _dec(v: Any) -> Decimal:
    if v is None or str(v).strip() == '':
        return Decimal('0')
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    text = str(v).strip().replace(',', '').replace('₱','')
    if text.startswith('(') and text.endswith(')'):
        text = '-' + text[1:-1]
    text = re.sub(r'(?i)\b(PHP|PESO|PESOS|DR|CR)\b','',text)
    text = re.sub(r'[^0-9.\-]','',text)
    if text in {'','-','.'}:
        return Decimal('0')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def _read_rows(path: Path) -> list[list[Any]]:
    suf = path.suffix.lower()
    if suf == '.csv':
        with path.open('r', encoding='utf-8-sig', newline='') as f:
            return [r for r in csv.reader(f)]
    if suf in {'.xlsx','.xlsm'}:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheets = sorted(wb.worksheets, key=lambda w: w.max_row * max(w.max_column,1), reverse=True)
            return [[c for c in row] for row in sheets[0].iter_rows(values_only=True)] if sheets else []
        finally:
            wb.close()
    if suf == '.xls':
        import xlrd
        wb = xlrd.open_workbook(str(path))
        sh = max(wb.sheets(), key=lambda s: s.nrows * max(s.ncols, 1))
        return [[sh.cell_value(r,c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    return []


# ---------------------------------------------------------------------------
# RAC Phil Corp paired-row journal parser
# Handles the format exported by their accounting system where each journal
# line spans two rows:
#   Row A: [date] [voucher_no] [acct_code] '' [description] [dept] ... [debit] [debit]
#   Row B:  ''     ''          [acct_name] [alias] '' [dept_name] ...  [credit] [credit]
# Transactions are separated by "Subtotal:" rows.
# ---------------------------------------------------------------------------

def _is_rac_format(rows: list[list[Any]]) -> bool:
    """Detect the RAC Phil Corp paired-row journal format."""
    header_texts = []
    for row in rows[:10]:
        for cell in row:
            t = str(cell or '').strip().lower()
            if t:
                header_texts.append(t)
    header_str = ' '.join(header_texts)
    return ('voucher date' in header_str and 'voucher no' in header_str
            and 'accounting a/c' in header_str or
            ('voucher date' in header_str and 'acct title' in header_str))


def _parse_rac_single_row_journal(rows: list[list[Any]]) -> list[LearnedLine]:
    """Parse the RAC Phil Corp single-row journal format (JOURNAL.xls export).

    Each journal line is a single row:
      col 2  — Accounting A/C (account code)
      col 3  — Acct Title (account name, same row as code)
      col 6  — Description
      col 17 — Debit Amt (PHP); fallback: col 15 (Trans. Curr.)
      col 18 — Credit Amt (PHP); fallback: col 16 (Trans. Curr.)
    """
    BANK_CODES = {
        '1010', '1020', '1030', '1040', '1041',
        '1101', '1102001', '1102002', '1102004', '1102005',
        '1102006', '1102007', '1102008', '1102009', '1102300',
        '1103', '1104', '1105', '1107', '1108', '1109',
        '1110', '1111', '1112', '1113', '1114',
    }
    BANK_WORDS = {'cash on hand', 'union bank', 'eastwest bank', 'east west bank',
                  'cash in bank', 'petty cash', 'pay mongo', 'gcash', 'maya', 'pko',
                  'e-wallet', 'ewallet', 'paymaya', 'paymongo'}

    def _is_cash(code: str, name: str) -> bool:
        if code in BANK_CODES:
            return True
        return any(w in _norm(name) for w in BANK_WORDS)

    def _is_header_or_subtotal(row):
        joined = str(row).lower()
        if 'voucher date' in joined:
            return True
        r0 = str(row[0] if row else '').strip().lower()
        if r0.startswith('subtotal') or r0.startswith('created on'):
            return True
        c2 = str(row[2] if len(row) > 2 else '').strip().lower()
        return c2 in ('acct title', 'accounting a/c')

    lines: list[LearnedLine] = []
    for row in rows:
        if not row or all(str(v).strip() in ('', ' ', 'None') for v in row):
            continue
        if _is_header_or_subtotal(row):
            continue

        col2 = row[2] if len(row) > 2 else ''
        col2_str = str(col2).strip()
        if not col2_str or col2_str in ('', ' ', 'None'):
            continue

        acct_code = col2_str.replace('.0', '') if isinstance(col2, float) else col2_str
        acct_name = str(row[3] if len(row) > 3 else '').strip() or acct_code
        desc = str(row[6] if len(row) > 6 else '').strip()

        # PHP amounts: prefer col 17/18 (Debit Amt / Credit Amt);
        # fall back to col 15/16 (Trans. Curr.) for shorter exports.
        if len(row) > 17:
            debit  = _dec(row[17]) or _dec(row[15] if len(row) > 15 else 0)
            credit = _dec(row[18] if len(row) > 18 else 0) or _dec(row[16] if len(row) > 16 else 0)
        elif len(row) > 15:
            debit  = _dec(row[15])
            credit = _dec(row[16] if len(row) > 16 else 0)
        else:
            continue

        if debit == 0 and credit == 0:
            continue
        if _is_cash(acct_code, acct_name):
            continue

        side   = 'debit' if debit > 0 else 'credit'
        amount = debit if side == 'debit' else credit
        kw = _keywords(desc, acct_name) or _keywords(acct_name)
        if not kw or amount == 0:
            continue

        lines.append(LearnedLine(
            description=desc or acct_name,
            account_code=acct_code,
            account_title=acct_name,
            normal_side=side,
            amount=amount,
            keywords=kw,
        ))

    if not lines:
        raise ValueError('No non-cash journal lines found in RAC Phil Corp single-row journal.')
    return lines


def _parse_rac_paired_row_journal(rows: list[list[Any]]) -> list[LearnedLine]:
    """Parse RAC Phil Corp's legacy voucher-based paired-row journal into LearnedLine objects."""
    import datetime

    def _xldate(val):
        """Convert Excel serial date to string."""
        try:
            import xlrd
            return str(datetime.datetime(*xlrd.xldate_as_tuple(val, 0)).date())
        except Exception:
            try:
                if isinstance(val, float) and val > 40000:
                    import datetime as dt
                    epoch = dt.date(1899, 12, 30)
                    return str(epoch + dt.timedelta(days=int(val)))
            except Exception:
                pass
            return ''

    def _is_page_header(row):
        r0 = str(row[0] if row else '').strip().lower()
        r2 = str(row[2] if len(row) > 2 else '').strip().lower()
        return ('voucher date' in str(row).lower() or
                'created on:' in r0 or
                r2 in ('acct title', 'accounting a/c'))

    def _is_subtotal(row):
        return str(row[0] if row else '').strip().lower().startswith('subtotal')

    # Collect all transactions as groups of lines
    transactions: list[dict] = []
    current_voucher = None
    current_desc = ''
    current_lines: list[dict] = []

    i = 5  # skip first few rows (company name / title / headers)
    while i < len(rows):
        row = rows[i]

        if not row or all(str(v).strip() in ('', ' ', 'None') for v in row):
            i += 1
            continue

        if _is_page_header(row):
            i += 1
            continue

        if _is_subtotal(row):
            if current_voucher and current_lines:
                transactions.append({
                    'voucher': current_voucher,
                    'description': current_desc,
                    'lines': current_lines,
                })
            current_voucher = None
            current_desc = ''
            current_lines = []
            i += 1
            continue

        # Row A: has account code in col 2
        col2 = row[2] if len(row) > 2 else ''
        col2_str = str(col2).strip()
        if col2_str and col2_str not in ('', ' ', 'None'):
            col0 = row[0] if len(row) > 0 else ''
            col1 = str(row[1] if len(row) > 1 else '').strip()

            # New voucher date?
            if isinstance(col0, float) and col0 > 40000:
                if col1 and col1 not in (' ', ''):
                    current_voucher = col1
                elif not current_voucher:
                    current_voucher = f'JE-{_xldate(col0)}'
            elif col1 and col1 not in (' ', '', 'None'):
                current_voucher = col1

            # Description and dept
            desc = str(row[4] if len(row) > 4 else '').strip()
            if desc and desc not in (' ',):
                current_desc = desc

            # Account code
            acct_code = col2_str.replace('.0', '') if isinstance(col2, float) else col2_str

            # Debit from col 10 or 11
            debit = _dec(row[10] if len(row) > 10 else 0) or _dec(row[11] if len(row) > 11 else 0)

            # Row B (next row) has account name and credit
            acct_name = ''
            credit = _dec('0')
            if i + 1 < len(rows):
                row_b = rows[i + 1]
                acct_name = str(row_b[2] if len(row_b) > 2 else '').strip()
                credit = _dec(row_b[10] if len(row_b) > 10 else 0) or _dec(row_b[11] if len(row_b) > 11 else 0)
                i += 2  # consume both rows
            else:
                i += 1

            if (debit > 0 or credit > 0) and acct_code:
                current_lines.append({
                    'code': acct_code,
                    'name': acct_name or acct_code,
                    'debit': debit,
                    'credit': credit,
                    'desc': current_desc,
                })
        else:
            i += 1

    # Flush last group
    if current_voucher and current_lines:
        transactions.append({
            'voucher': current_voucher,
            'description': current_desc,
            'lines': current_lines,
        })

    if not transactions:
        raise ValueError('No transactions found in RAC Phil Corp journal format.')

    # Convert to LearnedLine objects — exclude pure cash/bank accounts
    BANK_CODES = {
        # Unionbank / standard cash codes
        '1010', '1020', '1030',
        # E-wallet codes (GCash / Maya) — fund transfers, not expenses
        '1040', '1041',
        # RAC Phil Corp extended bank codes
        '1101', '1102001', '1102002', '1102004', '1102005',
        '1102006', '1102007', '1102008', '1102009', '1102300',
        '1103', '1104', '1105', '1107', '1108', '1109',
        '1110', '1111', '1112', '1113', '1114',
    }
    BANK_WORDS = {'cash on hand', 'union bank', 'eastwest bank', 'east west bank',
                  'cash in bank', 'petty cash', 'pay mongo', 'gcash', 'maya', 'pko',
                  'e-wallet', 'ewallet', 'paymaya', 'paymongo'}

    def _is_cash(code: str, name: str) -> bool:
        if code in BANK_CODES:
            return True
        n = _norm(name)
        return any(w in n for w in BANK_WORDS)

    lines: list[LearnedLine] = []
    for txn in transactions:
        desc = txn['description']
        for line in txn['lines']:
            if _is_cash(line['code'], line['name']):
                continue
            side = 'debit' if line['debit'] > 0 else 'credit'
            amount = line['debit'] if side == 'debit' else line['credit']
            kw = _keywords(desc, line['name'])
            if not kw:
                kw = _keywords(line['name'])
            if not kw or amount == 0:
                continue
            lines.append(LearnedLine(
                description=desc or line['name'],
                account_code=line['code'],
                account_title=line['name'],
                normal_side=side,
                amount=amount,
                keywords=kw,
            ))

    if not lines:
        raise ValueError('No non-cash journal lines found in RAC Phil Corp journal.')

    return lines

def _parse_rac_journal(rows: list[list[Any]]) -> list[LearnedLine]:
    """Detect journal format and dispatch to the correct sub-parser.

    Single-row format (JOURNAL.xls): header row has 'Acct Title' in col 3.
      → calls _parse_rac_single_row_journal
    Paired-row format (legacy): account name lives on the following row.
      → calls _parse_rac_paired_row_journal
    """
    for row in rows[:10]:
        joined = ' '.join(str(c or '').strip().lower() for c in row)
        if 'voucher date' in joined or 'accounting a/c' in joined:
            col3 = str(row[3] if len(row) > 3 else '').strip().lower()
            if col3 in ('acct title', 'acct. title', 'account title'):
                return _parse_rac_single_row_journal(rows)
            break
    return _parse_rac_paired_row_journal(rows)


def _is_subsidiary_ledger(rows: list[list[Any]]) -> bool:
    """Detect the PER-ACCOUNT subsidiary ledger format (A/C No. | Acct Title | ... | Balance)."""
    for row in rows[:10]:
        joined = ' '.join(str(v).strip().lower() for v in row if v)
        if 'a/c no' in joined and 'acct title' in joined and 'balance' in joined:
            return True
    return False


def _subsidiary_ledger_to_learned_lines(path: Path) -> list[LearnedLine]:
    """Parse a subsidiary ledger file and convert it to LearnedLine objects.

    Uses the voucher map to find the expense/asset accounts (non-cash DR legs)
    for each payment voucher, which gives much cleaner account-to-description
    mappings than the paired-row journal parser.
    """
    from app.parsers.subsidiary_ledger import (
        parse_subsidiary_ledger, build_voucher_map, _CASH_PREFIXES,
    )

    _TRIVIAL = {'6603', '1218002', '1218001', '2311', '2312'}

    def _is_cash(code: str) -> bool:
        return any(code.startswith(p) for p in _CASH_PREFIXES)

    ledger = parse_subsidiary_ledger(path)
    voucher_map = build_voucher_map(ledger)

    lines: list[LearnedLine] = []
    for entry in voucher_map.values():
        # Only meaningful expense/asset accounts on the debit side
        meaningful = [
            (c, t, a) for c, t, a in entry.debit_legs
            if not _is_cash(c) and c not in _TRIVIAL and a > 0
        ]
        if not meaningful or not entry.description:
            continue
        kw = _keywords(entry.description)
        if not kw:
            continue
        for code, title, amount in meaningful:
            lines.append(LearnedLine(
                description=entry.description,
                account_code=code,
                account_title=title,
                normal_side='debit',
                amount=amount,
                keywords=kw,
            ))

    if not lines:
        raise ValueError('No non-cash expense lines found in the subsidiary ledger.')
    return lines


def extract_chart_of_accounts_from_journal(path: Path) -> list[tuple]:
    """Extract all unique (code, title, type, is_active) tuples from a journal file.

    Handles the RAC Phil Corp paired-row format automatically.
    Also accepts any file with Account Code + Account Title columns.
    Returns a list of (code, title, type_str, is_active) ready for
    chart_of_accounts upsert — includes cash/bank accounts unlike the
    learning pipeline which excludes them.
    """
    rows = _read_rows(path)
    if not rows:
        raise ValueError('Uploaded file is empty or unreadable.')

    def _infer_type(code: str, title: str) -> str:
        """Infer account type from GL code prefix and title keywords."""
        c = code.replace('.', '').replace('-', '').strip()
        prefix = c[:1] if c else ''
        t = _norm(title)
        if prefix == '1':
            return 'asset'
        if prefix == '2':
            return 'liability'
        if prefix == '3':
            return 'equity'
        if prefix == '4':
            if any(w in t for w in ('income', 'revenue', 'sales', 'commission', 'interest income',
                                     'gain', 'rental', 'service fee')):
                return 'income'
            return 'income'
        if prefix in ('5', '6', '7', '8', '9'):
            return 'expense'
        # Fallback: keyword scan
        if any(w in t for w in ('receivable', 'prepaid', 'deposit', 'asset', 'input vat',
                                  'input tax', 'advance')):
            return 'asset'
        if any(w in t for w in ('payable', 'liability', 'withholding', 'output vat',
                                  'output tax', 'loan', 'mortgage')):
            return 'liability'
        if any(w in t for w in ('capital', 'equity', 'retained', 'surplus', 'deficit')):
            return 'equity'
        if any(w in t for w in ('sales', 'income', 'revenue', 'commission earned')):
            return 'income'
        return 'expense'

    seen: dict[str, tuple] = {}   # code → (code, title, type, is_active)

    # Subsidiary ledger: extract accounts directly from the per-account headers
    if _is_subsidiary_ledger(rows):
        from app.parsers.subsidiary_ledger import parse_subsidiary_ledger
        ledger = parse_subsidiary_ledger(path)
        for acct in ledger.accounts:
            if acct.code and acct.title and acct.code not in seen:
                seen[acct.code] = (acct.code, acct.title, _infer_type(acct.code, acct.title), 1)
        if seen:
            return sorted(seen.values(), key=lambda t: t[0])

    if _is_rac_format(rows):
        # RAC format: each data row has code in col 2 and title in col 3.
        # (Older comment about "paired-row" was incorrect — both fields sit in
        #  the same row; the paired structure refers to debit/credit split across
        #  two rows of the SAME journal line, not to code/title.)
        for row in rows[1:]:
            if not row or all(str(v).strip() in ('', ' ', 'None') for v in row):
                continue
            r0 = str(row[0] if row else '').strip().lower()
            if ('voucher date' in str(row).lower() or 'created on:' in r0
                    or str(row[2] if len(row) > 2 else '').strip().lower()
                    in ('acct title', 'accounting a/c')):
                continue
            if r0.startswith('subtotal'):
                continue

            col2 = row[2] if len(row) > 2 else ''
            col2_str = str(col2).strip()
            if not col2_str or col2_str in ('', ' ', 'None'):
                continue

            raw_code = col2_str.replace('.0', '') if isinstance(col2, float) else col2_str
            # Title is in col 3 (Acct Title) of the SAME row
            title = str(row[3] if len(row) > 3 else '').strip()
            if not title:
                title = raw_code  # fallback: use code as title if title cell is empty
            if raw_code and raw_code not in seen:
                seen[raw_code] = (raw_code, title, _infer_type(raw_code, title), 1)
    else:
        # Standard flat-column format — reuse the header-detection from ALIASES
        try:
            header_idx, m = _find_header(rows)
        except ValueError:
            raise ValueError(
                'Could not detect account columns. '
                'Need "Account Code" and "Account Title" (or similar) column headers, '
                'or upload the RAC Phil Corp journal export directly.'
            )
        for row in rows[header_idx + 1:]:
            if len(row) <= max(m.values()):
                continue
            code = str(row[m['code']] or '').strip()
            title = str(row[m['title']] or '').strip()
            if not code or not title:
                continue
            if _norm(code) in {'total', 'grand total', 'subtotal'}:
                continue
            if code not in seen:
                seen[code] = (code, title, _infer_type(code, title), 1)

    if not seen:
        raise ValueError('No account codes and titles found in the uploaded file.')

    return sorted(seen.values(), key=lambda t: t[0])


ALIASES = {
    'description': {'description','particulars','memo','remarks','narration','explanation','transaction','payee','supplier','vendor','details','line description','journal description'},
    'code': {'account code','acct code','gl code','code','account no','account number'},
    'title': {'account title','account name','account','gl account','description account','coa title'},
    'debit': {'debit','dr','debit amount','dr amount'},
    'credit': {'credit','cr','credit amount','cr amount'},
}

def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str,int]]:
    best = None
    for i,row in enumerate(rows[:80]):
        n = [_norm(x) for x in row]
        mapping = {}
        for field, opts in ALIASES.items():
            opts_n = {_norm(o) for o in opts}
            for idx,h in enumerate(n):
                if h in opts_n or any(o in h for o in opts_n if len(o) >= 5):
                    mapping[field] = idx; break
        score = len(mapping)
        if {'code','title','debit','credit'}.issubset(mapping): score += 3
        if 'description' in mapping: score += 1
        if best is None or score > best[0]: best = (score,i,mapping)
    if not best or not {'code','title','debit','credit'}.issubset(best[2]):
        raise ValueError('Could not detect journal columns. Need account code, account title, debit, and credit columns.')
    return best[1], best[2]


def _keywords(text: str, account_title: str='') -> str:
    words = re.findall(r'[a-z0-9]{3,}', _norm(text))
    seen=[]
    for w in words:
        if w in STOP or w in seen: continue
        seen.append(w)
        if len(seen) >= 8: break
    return ' '.join(seen)


def parse_historical_journal(path: Path) -> list[LearnedLine]:
    rows = _read_rows(path)
    if not rows:
        raise ValueError('Uploaded journal file is empty or unreadable.')

    # Subsidiary ledger format takes priority — it has more accurate account mappings
    if _is_subsidiary_ledger(rows):
        return _subsidiary_ledger_to_learned_lines(path)

    # RAC Phil Corp paired-row journal format
    if _is_rac_format(rows):
        return _parse_rac_journal(rows)

    # Fall back to standard flat-column format
    header_idx, m = _find_header(rows)
    lines: list[LearnedLine] = []
    last_desc = ''
    for row in rows[header_idx+1:]:
        if len(row) <= max(m.values()):
            continue
        code = str(row[m['code']] or '').strip()
        title = str(row[m['title']] or '').strip()
        if not code or not title or _norm(code) in {'total','grand total'}:
            continue
        desc = str(row[m.get('description', m['title'])] or '').strip() or last_desc or title
        if desc: last_desc = desc
        debit = _dec(row[m['debit']]); credit = _dec(row[m['credit']])
        if debit == 0 and credit == 0:
            continue
        side = 'debit' if debit >= credit else 'credit'
        amount = debit if side == 'debit' else credit
        if any(w in _norm(title) for w in CASH_WORDS):
            continue
        kw = _keywords(desc, title)
        if not kw:
            continue
        lines.append(LearnedLine(desc, code, title, side, amount, kw))
    if not lines:
        raise ValueError('No usable historical journal lines were found. Check that debit/credit and account columns contain values.')
    return lines


def ensure_learning_tables(conn):
    conn.executescript('''
    CREATE TABLE IF NOT EXISTS journal_learning_basis_files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL DEFAULT 1,
        filename TEXT NOT NULL,
        stored_path TEXT NOT NULL,
        uploaded_at TEXT NOT NULL DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS journal_learning_description_patterns (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL DEFAULT 1,
        keywords TEXT NOT NULL,
        account_code TEXT NOT NULL,
        account_title TEXT NOT NULL,
        normal_side TEXT NOT NULL,
        times_seen INTEGER NOT NULL DEFAULT 1,
        sample_description TEXT,
        learned_from_filename TEXT,
        last_seen_at TEXT NOT NULL DEFAULT (datetime('now')),
        UNIQUE(company_id, keywords, account_code, normal_side)
    );
    CREATE TABLE IF NOT EXISTS journal_learning_entry_templates (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER NOT NULL DEFAULT 1,
        group_key TEXT NOT NULL,
        account_code TEXT NOT NULL,
        account_title TEXT NOT NULL,
        normal_side TEXT NOT NULL,
        amount_ratio REAL NOT NULL DEFAULT 1.0,
        times_seen INTEGER NOT NULL DEFAULT 1,
        sample_description TEXT,
        learned_from_filename TEXT,
        last_seen_at TEXT NOT NULL DEFAULT (datetime("now")),
        UNIQUE(company_id, group_key, account_code)
    );
    ''')


def _account_type(side: str, title: str) -> str:
    t = _norm(title)
    if side == 'credit' and any(w in t for w in ('sales','revenue','income','commission')):
        return 'income'
    if any(w in t for w in ('asset','receivable','advance','deposit','input vat','input tax')):
        return 'asset'
    if any(w in t for w in ('payable','liability','withholding')):
        return 'liability'
    return 'expense' if side == 'debit' else 'income'


def _upsert_account(conn, code: str, name: str, atype: str) -> None:
    """Insert or update a chart_of_accounts row.

    Update rules (applied to an existing row on conflict):
    - Always update type (derived from code prefix — reliable)
    - Update name only when the stored name looks wrong:
        • name equals code (placeholder / corrupted)
        • name is all digits and shorter than the incoming name
        • name length ≤ 8 and the incoming name is more descriptive
      Otherwise leave the existing name untouched so a correct human-entered
      name is never silently overwritten by a new upload.
    """
    conn.execute(
        """INSERT INTO chart_of_accounts (code, name, type) VALUES (?, ?, ?)
           ON CONFLICT(code) DO UPDATE SET
             type = excluded.type,
             name = CASE
               WHEN chart_of_accounts.name = chart_of_accounts.code          THEN excluded.name
               WHEN chart_of_accounts.name GLOB '[0-9]*'                     THEN excluded.name
               WHEN length(chart_of_accounts.name) <= 8
                    AND length(excluded.name) > length(chart_of_accounts.name) THEN excluded.name
               ELSE chart_of_accounts.name
             END""",
        (code, name, atype),
    )


def learn_from_historical_journal(conn, *, path: Path, company_id: int, filename: str, stored_path: str) -> list[LearnedLine]:
    ensure_learning_tables(conn)
    lines = parse_historical_journal(path)
    conn.execute('INSERT INTO journal_learning_basis_files (company_id, filename, stored_path) VALUES (?, ?, ?)', (company_id, filename, stored_path))

    # ── Step 0: Upsert ALL accounts from the file into chart_of_accounts ───────
    # This runs before the pattern loop so that every account in the journal
    # (including cash/bank accounts that the learning pipeline excludes) gets a
    # proper name in the COA.  Works for any company — RAC Phil Corp, PGI, etc.
    try:
        all_accounts = extract_chart_of_accounts_from_journal(path)
        for code, title, atype, _ in all_accounts:
            _upsert_account(conn, code, title, atype)
    except Exception:
        pass  # don't abort learning if COA extraction fails; per-line upserts below will cover it

    # ── Per-line patterns (single-account matching) ──
    for line in lines:
        _upsert_account(conn, line.account_code, line.account_title,
                        _account_type(line.normal_side, line.account_title))
        conn.execute('''INSERT INTO journal_learning_description_patterns
            (company_id, keywords, account_code, account_title, normal_side, times_seen, sample_description, learned_from_filename, last_seen_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, datetime('now'))
            ON CONFLICT(company_id, keywords, account_code, normal_side) DO UPDATE SET
              account_title=excluded.account_title,
              times_seen=times_seen+1,
              sample_description=excluded.sample_description,
              learned_from_filename=excluded.learned_from_filename,
              last_seen_at=datetime('now')''',
            (company_id, line.keywords, line.account_code, line.account_title, line.normal_side, line.description[:300], filename))

    # ── Multi-account entry templates (VAT / split-entry matching) ──
    # Group lines by raw description so co-occurring accounts (e.g. Office Supplies + Input VAT)
    # are stored with their proportional ratios.  The cash/bank side is excluded upstream, so
    # the group total represents the total debit (or credit) side of the entry.
    desc_groups: dict[str, list[LearnedLine]] = defaultdict(list)
    for line in lines:
        desc_groups[line.description].append(line)

    for desc, group in desc_groups.items():
        if len(group) < 2:
            continue  # single-account entries need no template

        # For bank-statement JE templates we only want same-side lines:
        # - outflow entries: keep the debit (Dr) lines (expenses + VAT)
        # - inflow entries:  keep the credit (Cr) lines (income + VAT payable)
        # This prevents AP / liability lines (the balancing Cr of an AP entry)
        # from appearing in templates that will be applied against bank amounts.
        debit_lines  = [l for l in group if l.normal_side == 'debit']
        credit_lines = [l for l in group if l.normal_side == 'credit']

        if len(debit_lines) >= 2:
            template_lines = debit_lines       # expense/asset split on Dr side
        elif len(credit_lines) >= 2:
            template_lines = credit_lines      # income split on Cr side
        else:
            continue  # only one account on each side – nothing to split

        total = sum(l.amount for l in template_lines)
        if total == 0:
            continue
        # Use description-only keywords (no account title noise) as the group key
        group_key = _keywords(desc)
        if not group_key:
            continue
        for line in template_lines:
            ratio = float(line.amount / total)
            conn.execute('''INSERT INTO journal_learning_entry_templates
                (company_id, group_key, account_code, account_title, normal_side, amount_ratio, times_seen, sample_description, learned_from_filename, last_seen_at)
                VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, datetime('now'))
                ON CONFLICT(company_id, group_key, account_code) DO UPDATE SET
                  account_title=excluded.account_title,
                  normal_side=excluded.normal_side,
                  amount_ratio=(amount_ratio * times_seen + excluded.amount_ratio) / (times_seen + 1),
                  times_seen=times_seen+1,
                  sample_description=excluded.sample_description,
                  learned_from_filename=excluded.learned_from_filename,
                  last_seen_at=datetime('now')''',
                (company_id, group_key, line.account_code, line.account_title, line.normal_side,
                 ratio, desc[:300], filename))

    return lines


def match_historical_pattern(conn, *, description: str, company_id: int = 1):
    """Return the single best-matching pattern row (for backward compat)."""
    ensure_learning_tables(conn)
    desc_norm = _norm(description)
    desc_words = set(re.findall(r'[a-z0-9]{3,}', desc_norm)) - STOP
    if not desc_words:
        return None
    rows = conn.execute('''SELECT * FROM journal_learning_description_patterns
                           WHERE company_id = ? ORDER BY times_seen DESC, last_seen_at DESC LIMIT 500''', (company_id,)).fetchall()
    best = None
    for r in rows:
        kws = set((r['keywords'] or '').split())
        if not kws: continue
        overlap = len(desc_words & kws)
        phrase_bonus = 2 if (r['keywords'] and r['keywords'] in desc_norm) else 0
        score = overlap + phrase_bonus + min(int(r['times_seen'] or 1), 5) * 0.1
        if overlap >= 1 and (best is None or score > best[0]):
            best = (score, r)
    if not best or best[0] < 1.1:
        return None
    return best[1]


def match_all_historical_patterns(conn, *, description: str, company_id: int = 1):
    """Return ALL matching patterns for a description.

    First tries multi-account entry templates (e.g. Office Supplies + Input VAT together).
    Falls back to individual per-line patterns if no template found.

    Returns a list of dicts with keys:
        account_code, account_title, normal_side, amount_ratio, confidence
    or None if nothing matches.
    """
    ensure_learning_tables(conn)
    desc_norm = _norm(description)
    desc_words = set(re.findall(r'[a-z0-9]{3,}', desc_norm)) - STOP
    if not desc_words:
        return None

    # ── Try entry templates first (multi-account) ──
    templates = conn.execute(
        '''SELECT * FROM journal_learning_entry_templates
           WHERE company_id = ? ORDER BY times_seen DESC, last_seen_at DESC LIMIT 500''',
        (company_id,)).fetchall()

    best_group_key = None
    best_score = 0.0
    for t in templates:
        kws = set((t['group_key'] or '').split())
        if not kws: continue
        overlap = len(desc_words & kws)
        phrase_bonus = 2 if (t['group_key'] and t['group_key'] in desc_norm) else 0
        score = overlap + phrase_bonus + min(int(t['times_seen'] or 1), 5) * 0.1
        if overlap >= 1 and score > best_score:
            best_score = score
            best_group_key = t['group_key']

    if best_group_key and best_score >= 1.1:
        group_rows = conn.execute(
            '''SELECT * FROM journal_learning_entry_templates
               WHERE company_id = ? AND group_key = ?
               ORDER BY amount_ratio DESC''',
            (company_id, best_group_key)).fetchall()
        if group_rows:
            confidence = min(0.92, 0.60 + best_score * 0.04)
            return [
                {
                    'account_code': r['account_code'],
                    'account_title': r['account_title'],
                    'normal_side': r['normal_side'],
                    'amount_ratio': r['amount_ratio'],
                    'confidence': confidence,
                    'source': f'entry_template:{best_group_key}',
                    'learned_from_filename': r['learned_from_filename'],
                }
                for r in group_rows
            ]

    # ── Fall back to single best pattern ──
    single = match_historical_pattern(conn, description=description, company_id=company_id)
    if single is None:
        return None
    return [
        {
            'account_code': single['account_code'],
            'account_title': single['account_title'],
            'normal_side': single['normal_side'],
            'amount_ratio': 1.0,
            'confidence': min(0.92, 0.60 + (single['times_seen'] or 1) * 0.03),
            'source': f'single_pattern:{single["keywords"]}',
            'learned_from_filename': single['learned_from_filename'],
        }
    ]
