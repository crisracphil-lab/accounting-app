from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from app.db import db, log_action


class OpenItemsError(ValueError):
    pass


@dataclass
class LedgerRow:
    row_number: int
    date: str
    account_code: str | None
    account_title: str | None
    reference: str | None
    description: str
    debit: Decimal
    credit: Decimal
    amount: Decimal


ALIASES = {
    "date": {
        "date", "transaction date", "posting date", "value date", "entry date",
        "journal date", "doc date", "document date",
        # RAC / local accounting system variants
        "voucher date", "trans date", "check date", "chk date", "ck date",
        "trxn date", "je date", "or date", "check voucher date",
    },
    "account_code": {
        "account code", "acct code", "gl code", "code", "account no",
        "account number", "coa code",
        # RAC / local variants — including "A/C No." which normalises to "a/c no"
        "a/c no", "a/c no.", "a/c number", "a c no",
        "accounting a c", "accounting ac", "acct no", "gl acct", "acct",
        "account id", "gl no",
    },
    "account_title": {
        "account title", "account name", "account", "gl account",
        "coa title", "title",
        # RAC variants
        "acct title", "acct name", "account description",
    },
    "reference": {
        "reference", "ref", "ref no", "reference no", "voucher no",
        "doc no", "document no", "invoice no", "receipt no",
        "je no", "journal no",
        # RAC / local variants
        "voucher number", "check no", "chk no", "ck no", "or no",
        "cv no", "si no", "jv no", "pv no", "ar no", "ap no",
        "trans no", "trxn no", "transaction no",
    },
    "description": {
        "description", "particulars", "remarks", "memo", "narration",
        "details", "supplier", "customer", "payee", "name",
        # RAC / local variants
        "explanation", "entry description", "transaction description",
        "particulars description", "payee name", "vendor name",
    },
    "debit": {
        "debit", "dr", "debit amount", "dr amount",
        # RAC variants
        "debit amt", "dr amt", "debit total",
    },
    "credit": {
        "credit", "cr", "credit amount", "cr amount",
        # RAC variants
        "credit amt", "cr amt", "credit total",
    },
    "amount": {
        "amount", "net amount", "movement", "transaction amount",
        "balance movement",
        # RAC variants
        "amt", "net amt", "trxn amount", "trans amount",
    },
    "amount_type": {"dr cr", "debit credit", "d c", "dc", "type", "direction", "sign"},
}


def _norm(v: Any) -> str:
    text = str(v or "").strip().lower()
    text = re.sub(r"[\n\r\t_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9/ ]+", " ", text)
    return " ".join(text.split())


def _col(headers: list[Any], aliases: set[str], used: set[int]) -> int | None:
    normalized = [_norm(h) for h in headers]
    norm_aliases = {_norm(a) for a in aliases}
    for i, h in enumerate(normalized):
        if i not in used and h in norm_aliases:
            return i
    for i, h in enumerate(normalized):
        if i in used:
            continue
        for a in norm_aliases:
            # len>=4 lets "date" substring-match "voucher date", "trans date", etc.
            if h == a or (len(a) >= 4 and a in h):
                return i
    return None


def _money(value: Any, row_no: int, field: str, default: Decimal | None = Decimal("0")) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).strip().upper()
    sign = Decimal("1")
    if text.endswith("DR") or " DR" in text:
        sign = Decimal("-1")
    if text.endswith("CR") or " CR" in text:
        sign = Decimal("1")
    text = re.sub(r"\b(PHP|PESO|PESOS|DR|CR)\b", "", text)
    text = text.replace(",", "").replace("â‚±", "").strip()
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return default
    try:
        return Decimal(text) * sign
    except InvalidOperation as exc:
        raise OpenItemsError(f"Invalid {field} {value!r} on row {row_no}") from exc


def _date(value: Any, row_no: int) -> str:
    if value is None or str(value).strip() == "":
        raise OpenItemsError(f"Missing date on row {row_no}")
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception as exc:
            raise OpenItemsError(f"Invalid date {value!r} on row {row_no}") from exc
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y", "%m/%d/%y", "%d/%m/%y", "%b %d %Y", "%d %b %Y", "%B %d %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise OpenItemsError(f"Invalid date {value!r} on row {row_no}")


def _read_rows(path: str | Path) -> list[list[Any]]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            return [list(r) for r in csv.reader(f)]
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise OpenItemsError("Open items upload must be .xlsx, .xlsm, .xls, or .csv.")
    best_rows: list[list[Any]] = []
    best_score = -1
    if suffix == ".xls":
        import xlrd  # type: ignore[import]
        wb_xls = xlrd.open_workbook(str(path))
        for i in range(wb_xls.nsheets):
            ws = wb_xls.sheet_by_index(i)
            rows = []
            for r in range(ws.nrows):
                row_vals = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        row_vals.append(xlrd.xldate.xldate_as_datetime(cell.value, wb_xls.datemode))
                    else:
                        row_vals.append(cell.value)
                rows.append(row_vals)
            score = 0
            for row in rows[:30]:
                header_text = " ".join(_norm(c) for c in row if c is not None)
                if any(a in header_text for a in ("account", "debit", "credit", "amount", "date")):
                    score += 1
            if score > best_score:
                best_score = score
                best_rows = rows
    else:
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            score = 0
            for row in rows[:30]:
                header_text = " ".join(_norm(c) for c in row if c is not None)
                if any(a in header_text for a in ("account", "debit", "credit", "amount", "date")):
                    score += 1
            if score > best_score:
                best_score = score
                best_rows = rows
    return best_rows


def _mapping(headers: list[Any]) -> dict[str, int]:
    used: set[int] = set()
    m: dict[str, int] = {}
    for field in ["date", "account_code", "account_title", "reference", "description", "debit", "credit", "amount", "amount_type"]:
        c = _col(headers, ALIASES[field], used)
        if c is not None:
            m[field] = c
            if field not in {"reference", "description", "account_title", "account_code"}:
                used.add(c)
    if "description" not in m and "reference" in m:
        m["description"] = m["reference"]
    if "reference" not in m and "description" in m:
        m["reference"] = m["description"]
    if "date" not in m or "description" not in m or not ({"amount", "debit", "credit"} & set(m)):
        raise OpenItemsError("Required columns not found. Need date, description/reference, and amount or debit/credit columns.")
    return m


def parse_ledger_file(path: str | Path) -> list[LedgerRow]:
    rows = _read_rows(path)
    header_idx = None
    mapping = None
    candidate_rows: list[str] = []
    for i, row in enumerate(rows[:50]):
        # Collect non-empty rows for the error message
        row_text = ", ".join(str(c) for c in row if c is not None and str(c).strip())
        if row_text:
            candidate_rows.append(f"  Row {i+1}: {row_text[:120]}")
        try:
            mapping = _mapping(row)
            header_idx = i
            break
        except OpenItemsError:
            continue
    if mapping is None or header_idx is None:
        preview = "\n".join(candidate_rows[:10]) if candidate_rows else "  (file appears empty)"
        raise OpenItemsError(
            "Could not detect a valid ledger header row. "
            "Need columns for: date, description/reference, and debit/credit/amount.\n"
            f"First rows found in file:\n{preview}"
        )
    # Pre-compute the set of known date-column header texts so we can detect
    # repeated page-break header rows inside the data section.
    _date_header_texts = {_norm(a) for a in ALIASES["date"]}

    out: list[LedgerRow] = []
    last_date: str | None = None          # carry-forward: blank date = same as previous row
    last_account_code: str | None = None  # carry-forward: blank account = same as previous row
    last_account_title: str | None = None
    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        def v(field: str):
            c = mapping.get(field)
            return row[c] if c is not None and c < len(row) else None

        # ── Skip repeated page-break header rows.
        # Multi-page exports repeat the column header row on every page.
        # Detect by checking if the date column contains a known header alias
        # (e.g. "Voucher Date", "Date") instead of an actual date value.
        raw_date_cell = v("date")
        if raw_date_cell is not None and _norm(str(raw_date_cell)) in _date_header_texts:
            continue  # this row IS a header line repeated by the pager — skip it

        # ── Carry-forward: update account code/title from EVERY non-empty row
        # (including account-group header rows that have no transaction data).
        # This means the group header row "1147001 | Accounts Receivable-Trade"
        # seeds the carry-forward even though it has no date / description.
        _raw_code_val = v("account_code")
        # Normalise float codes: xlrd returns 1147001.0 → store as "1147001"
        if isinstance(_raw_code_val, float) and _raw_code_val == int(_raw_code_val):
            _raw_code_val = str(int(_raw_code_val))
        raw_code = str(_raw_code_val or "").strip()
        raw_title = str(v("account_title") or "").strip()
        # Only update carry-forward if the code looks like a real account code
        # (numeric, or alphanumeric without long spaces). This prevents page-break
        # meta-header rows like "Created on: 2026/05/18" from polluting the carry.
        _code_looks_valid = raw_code and (
            raw_code.replace(".", "").replace("-", "").isdigit()
            or (len(raw_code) <= 20 and " " not in raw_code)
        )
        if _code_looks_valid:
            last_account_code = raw_code
        if raw_title and len(raw_title) <= 80:
            last_account_title = raw_title

        # ── Carry-forward: update date from every row that has one
        raw_date = v("date")
        raw_date_str = str(raw_date).strip() if raw_date is not None else ""
        if raw_date_str and raw_date_str not in {" ", "  "}:
            last_date = _date(raw_date, idx)

        # ── Skip rows with no transaction description/reference
        desc = str(v("description") or "").strip()
        ref = str(v("reference") or "").strip() or None
        if not desc and not ref:
            continue

        # ── Resolve date for this transaction row
        if last_date is None:
            raise OpenItemsError(
                f"Row {idx} has no date and no previous date to carry forward. "
                "Check that 'Voucher Date' column has a date somewhere above this row."
            )
        row_date = last_date

        account_code = raw_code or last_account_code
        account_title = raw_title or last_account_title
        debit = _money(v("debit"), idx, "debit") or Decimal("0")
        credit = _money(v("credit"), idx, "credit") or Decimal("0")
        amount_value = _money(v("amount"), idx, "amount", None)
        if amount_value is not None:
            direction = _norm(v("amount_type"))
            if direction in {"debit", "dr", "payment", "outflow"} and amount_value > 0:
                amount_value = -amount_value
            if amount_value < 0:
                debit = abs(amount_value)
                credit = Decimal("0")
            else:
                debit = Decimal("0")
                credit = amount_value
            amount = amount_value
        else:
            amount = credit - debit
        out.append(LedgerRow(
            row_number=idx,
            date=row_date,
            account_code=account_code or None,
            account_title=account_title or None,
            reference=ref,
            description=desc or ref or "",
            debit=debit,
            credit=credit,
            amount=amount,
        ))
    if not out:
        raise OpenItemsError("No ledger rows found after the detected header.")
    return out


def _infer_open_side(account_text: str, requested: str) -> str:
    if requested in {"debit", "credit"}:
        return requested
    text = account_text.lower()
    if any(w in text for w in ("payable", "accrued", "due to", "liability")):
        return "credit"
    return "debit"


def _score(a: LedgerRow, b: LedgerRow) -> tuple[int, int, int]:
    same_ref = 0 if (a.reference and b.reference and _norm(a.reference) == _norm(b.reference)) else 1
    text_hit = 0 if (_norm(a.description) and _norm(a.description) in _norm(b.description) or _norm(b.description) in _norm(a.description)) else 1
    days = abs((date.fromisoformat(a.date) - date.fromisoformat(b.date)).days)
    return (same_ref, text_hit, days)


def create_open_items_run(path: str | Path, account_filter: str = "", open_side: str = "auto", tolerance: Decimal = Decimal("0.05"), company_id: int = 1) -> int:
    rows = parse_ledger_file(path)
    needle = _norm(account_filter)
    if needle:
        rows = [r for r in rows if needle in _norm(r.account_code) or needle in _norm(r.account_title)]
    if not rows:
        raise OpenItemsError("No rows matched the selected account title/code.")
    account_text = account_filter or rows[0].account_title or rows[0].account_code or "ledger account"
    side = _infer_open_side(account_text, open_side)

    openings: list[dict[str, Any]] = []
    closings: list[dict[str, Any]] = []
    for r in rows:
        open_amount = r.debit if side == "debit" else r.credit
        close_amount = r.credit if side == "debit" else r.debit
        if open_amount > tolerance:
            openings.append({"row": r, "original": open_amount, "remaining": open_amount})
        if close_amount > tolerance:
            closings.append({"row": r, "original": close_amount, "remaining": close_amount})

    closures = []
    for closing in closings:
        candidates = [o for o in openings if o["remaining"] > tolerance and date.fromisoformat(o["row"].date) <= date.fromisoformat(closing["row"].date)]
        candidates.sort(key=lambda o: (_score(o["row"], closing["row"]), date.fromisoformat(o["row"].date)))
        for opening in candidates:
            if closing["remaining"] <= tolerance:
                break
            applied = min(opening["remaining"], closing["remaining"])
            opening["remaining"] -= applied
            closing["remaining"] -= applied
            closures.append((opening["row"], closing["row"], applied))

    today = date.today()
    with db() as conn:
        conn.execute("BEGIN")
        try:
            cur = conn.execute(
                "INSERT INTO open_item_runs (company_id, filename, account_filter, open_side, row_count) VALUES (?, ?, ?, ?, ?)",
                (company_id, Path(path).name, account_filter, side, len(rows)),
            )
            run_id = cur.lastrowid
            for o in openings:
                r = o["row"]
                closed_amount = o["original"] - o["remaining"]
                if o["remaining"] <= tolerance:
                    status = "closed"
                elif closed_amount > tolerance:
                    status = "partial"
                else:
                    status = "open"
                aging = (today - date.fromisoformat(r.date)).days
                conn.execute(
                    """INSERT INTO open_item_details
                       (run_id, status, open_date, account_code, account_title, reference, description,
                        original_amount, closed_amount, open_balance, aging_days)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (run_id, status, r.date, r.account_code, r.account_title, r.reference, r.description,
                     str(o["original"]), str(closed_amount), str(max(o["remaining"], Decimal("0"))), aging),
                )
            for opening, closing, applied in closures:
                conn.execute(
                    """INSERT INTO open_item_closures
                       (run_id, open_reference, close_reference, open_date, close_date, amount, close_description)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (run_id, opening.reference, closing.reference, opening.date, closing.date, str(applied), closing.description),
                )
            log_action(conn, "open_items", "open_item_run", run_id, {"rows": len(rows), "account_filter": account_filter, "open_side": side})
            conn.execute("COMMIT")
            return run_id
        except Exception:
            conn.execute("ROLLBACK")
            raise


def export_open_items_xlsx(run_id: int) -> bytes:
    from app.services.excel_styles import (
        add_corp_header, write_column_headers, style_data_rows,
        add_total_row, auto_col_width, finalize_workbook,
    )
    with db() as conn:
        run = conn.execute("SELECT * FROM open_item_runs WHERE id = ?", (run_id,)).fetchone()
        if run is None:
            raise OpenItemsError("Open item run not found.")
        details = conn.execute(
            "SELECT * FROM open_item_details WHERE run_id = ? ORDER BY status, open_date", (run_id,)
        ).fetchall()
        closures = conn.execute(
            "SELECT * FROM open_item_closures WHERE run_id = ? ORDER BY open_date, close_date", (run_id,)
        ).fetchall()

    wb = Workbook()

    # ── Sheet 1: Open Items ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Open Items"
    period = f"As of run #{run_id} | File: {run['filename']} | Created: {run['created_at']}"
    headers = ["Status", "Date", "Account Code", "Account Title", "Reference",
               "Description", "Original Amt", "Closed Amt", "Open Balance", "Aging Days"]
    num_cols = len(headers)
    data_row = add_corp_header(ws, "Open Items Report", period, num_cols)
    write_column_headers(ws, data_row, headers)
    data_row += 1

    total_orig = total_closed = total_open = 0.0
    for d in details:
        orig   = float(d["original_amount"])
        closed = float(d["closed_amount"])
        openb  = float(d["open_balance"])
        total_orig   += orig
        total_closed += closed
        total_open   += openb
        ws.append([d["status"], d["open_date"], d["account_code"], d["account_title"],
                   d["reference"], d["description"], orig, closed, openb, d["aging_days"]])

    last_data = ws.max_row
    style_data_rows(ws, data_row, last_data, num_cols,
                    money_cols={7, 8, 9}, int_cols={10})
    add_total_row(ws, last_data + 1, num_cols,
                  {7: total_orig, 8: total_closed, 9: total_open})
    auto_col_width(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["E"].width = 16

    # ── Sheet 2: Closures ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Closures")
    cl_headers = ["Open Reference", "Close Reference", "Open Date", "Close Date",
                  "Closed Amount", "Close Description"]
    cl_num = len(cl_headers)
    cl_data_row = add_corp_header(ws2, "Item Closures", period, cl_num)
    write_column_headers(ws2, cl_data_row, cl_headers)
    cl_data_row += 1

    total_cl = 0.0
    for c in closures:
        amt = float(c["amount"])
        total_cl += amt
        ws2.append([c["open_reference"], c["close_reference"],
                    c["open_date"], c["close_date"], amt, c["close_description"]])

    cl_last = ws2.max_row
    style_data_rows(ws2, cl_data_row, cl_last, cl_num, money_cols={5})
    add_total_row(ws2, cl_last + 1, cl_num, {5: total_cl})
    auto_col_width(ws2)

    return finalize_workbook(wb)
