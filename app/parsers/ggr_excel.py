"""GGR Excel parser - sums per-metric across per-outlet detail sheets."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, Optional

import openpyxl


class GGRParseError(Exception):
    pass


COL_DATE              = 1
COL_GGR_COMBINED      = 18   # Q = F + N (combined Landbased + Online GGR)
COL_PAGCOR_COMBINED   = 19   # R = G + O
COL_AUDIT_COMBINED    = 20   # S = H + P
COL_OPERATOR          = 21
COL_SERVICE_PROVIDER  = 22


@dataclass
class MetricTotals:
    positive: Decimal = Decimal("0")
    negative: Decimal = Decimal("0")
    rows_positive: int = 0
    rows_negative: int = 0
    rows_zero: int = 0

    def add(self, value: Decimal) -> None:
        if value > 0:
            self.positive += value
            self.rows_positive += 1
        elif value < 0:
            self.negative += value
            self.rows_negative += 1
        else:
            self.rows_zero += 1


@dataclass
class GGRWorkbookTotals:
    period_start: Optional[date]
    period_end: Optional[date]
    sheet_count: int
    daily_rows: int
    ggr:           MetricTotals = field(default_factory=MetricTotals)
    pagcor_share:  MetricTotals = field(default_factory=MetricTotals)
    audit_fee:     MetricTotals = field(default_factory=MetricTotals)
    operator:      MetricTotals = field(default_factory=MetricTotals)
    service_prov:  MetricTotals = field(default_factory=MetricTotals)
    per_outlet:    Dict[str, "GGRWorkbookTotals"] = field(default_factory=dict)


def _to_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    return Decimal(str(v))


def _is_outlet_sheet(name: str) -> bool:
    return len(name.strip()) >= 6 and name.strip().upper().startswith("LW")


def _coerce_date(v) -> Optional[date]:
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def parse_ggr_excel(file_path,
                    period_start: Optional[date] = None,
                    period_end: Optional[date] = None) -> GGRWorkbookTotals:
    path = Path(file_path)
    if not path.exists():
        raise GGRParseError(f"File not found: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise GGRParseError(f"Expected .xlsx, got {path.suffix}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    outlet_sheets = [s for s in wb.sheetnames if _is_outlet_sheet(s)]
    if not outlet_sheets:
        raise GGRParseError(
            f"No per-outlet sheets found (expected LWxxNNN). "
            f"Workbook has: {wb.sheetnames}")

    totals = GGRWorkbookTotals(
        period_start=period_start, period_end=period_end,
        sheet_count=len(outlet_sheets), daily_rows=0)

    for sn in outlet_sheets:
        outlet = GGRWorkbookTotals(period_start=period_start,
                                   period_end=period_end,
                                   sheet_count=1, daily_rows=0)
        ws = wb[sn]
        for r in range(3, ws.max_row + 1):
            row_date = _coerce_date(ws.cell(row=r, column=COL_DATE).value)
            if row_date is None:
                continue
            if period_start and row_date < period_start:
                continue
            if period_end and row_date > period_end:
                continue

            for attr, col in (("ggr", COL_GGR_COMBINED),
                              ("pagcor_share", COL_PAGCOR_COMBINED),
                              ("audit_fee", COL_AUDIT_COMBINED),
                              ("operator", COL_OPERATOR),
                              ("service_prov", COL_SERVICE_PROVIDER)):
                v = _to_decimal(ws.cell(row=r, column=col).value)
                getattr(totals, attr).add(v)
                getattr(outlet, attr).add(v)

            totals.daily_rows += 1
            outlet.daily_rows += 1

        totals.per_outlet[sn] = outlet

    return totals
