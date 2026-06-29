"""
UnionBank RA260 outgoing-payments register parser.

The .xlsx export has ONE 'Transactions' sheet with these columns (1-based):
   1 Company
   2 Source Account
   3 Batch ID
   4 Remarks
   5 Remittance Type   (instaPay / PESONet / UnionBank Account)
   6 Transaction Date  (mm/dd/yyyy string)
   7 Transaction Status (Successful / Released / Scheduled / Failed)
   8 Tran ID
   9 Transaction Volume (peso amount, may be a string with commas)
  10 Transaction Count
  11 Beneficiary Code
  12 Beneficiary Name
  13 Beneficiary Account Number
  14 Beneficiary Address
"""
from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

import openpyxl


class RA260ParseError(Exception):
    pass


_REQUIRED_COLUMNS = (
    "Company", "Source Account", "Batch ID", "Remarks", "Remittance Type",
    "Transaction Date", "Transaction Status", "Tran ID",
    "Transaction Volume", "Beneficiary Name",
)


@dataclass
class PaymentInstruction:
    company: Optional[str]
    source_account: Optional[str]
    batch_id: Optional[str]
    remarks: Optional[str]
    remittance_type: Optional[str]
    transaction_date: Optional[date]
    transaction_status: Optional[str]
    tran_id: str
    amount: Decimal
    transaction_count: Optional[int]
    beneficiary_code: Optional[str]
    beneficiary_name: Optional[str]
    beneficiary_account: Optional[str]
    beneficiary_address: Optional[str]

    def as_dict(self) -> dict:
        d = asdict(self)
        if self.transaction_date:
            d["transaction_date"] = self.transaction_date.isoformat()
        d["amount"] = str(self.amount)
        return d


def _coerce_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal("0")
    s = str(v).strip().replace(",", "")
    if s == "":
        return Decimal("0")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise RA260ParseError(f"Cannot parse amount {v!r}: {exc}") from exc


def _to_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_int(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return None


def parse_ra260(file_path) -> List[PaymentInstruction]:
    path = Path(file_path)
    if not path.exists():
        raise RA260ParseError(f"File not found: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise RA260ParseError(f"Expected .xlsx, got {path.suffix}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet_name = next((s for s in wb.sheetnames if s.strip().lower() == "transactions"), None)
    if sheet_name is None:
        raise RA260ParseError(
            f"Workbook must have a 'Transactions' sheet. Got: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Find header row (typically row 1) and build column map
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for idx, h in enumerate(headers, start=1):
        if isinstance(h, str):
            col_map[h.strip()] = idx
    missing = [c for c in _REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise RA260ParseError(
            f"Header row missing required columns: {missing}. "
            f"Got: {list(col_map.keys())}")

    def col(row, name):
        idx = col_map.get(name)
        if idx is None:
            return None
        return ws.cell(row=row, column=idx).value

    rows: List[PaymentInstruction] = []
    for r in range(2, ws.max_row + 1):
        tran = _to_str(col(r, "Tran ID"))
        amt_raw = col(r, "Transaction Volume")
        if not tran and (amt_raw is None or amt_raw == ""):
            continue  # blank padding row
        if not tran:
            # Tran ID can be missing on Scheduled rows - synthesize one
            batch_id = _to_str(col(r, "Batch ID")) or f"row{r}"
            tran = f"SCHED-{batch_id}-{r}"

        rows.append(PaymentInstruction(
            company             = _to_str(col(r, "Company")),
            source_account      = _to_str(col(r, "Source Account")),
            batch_id            = _to_str(col(r, "Batch ID")),
            remarks             = _to_str(col(r, "Remarks")),
            remittance_type     = _to_str(col(r, "Remittance Type")),
            transaction_date    = _coerce_date(col(r, "Transaction Date")),
            transaction_status  = _to_str(col(r, "Transaction Status")),
            tran_id             = tran,
            amount              = _to_decimal(amt_raw),
            transaction_count   = _to_int(col(r, "Transaction Count")),
            beneficiary_code    = _to_str(col(r, "Beneficiary Code")),
            beneficiary_name    = _to_str(col(r, "Beneficiary Name")),
            beneficiary_account = _to_str(col(r, "Beneficiary Account Number")),
            beneficiary_address = _to_str(col(r, "Beneficiary Address")),
        ))

    return rows
