"""
UnionBank XLSX bank statement parser.

Built against the real UB Online export format:
  - One workbook may have multiple sheets (Sheet1, Sheet1 (2), ...).
    The same Transaction ID may appear in several sheets, sometimes
    with the amount columns blank in one sheet and populated in another.
  - Each sheet has a header section, a 'TRANSACTIONS LIST:' marker on
    column A, a column-header row immediately after, then transactions.
  - Empty cells appear as None or as a literal space ' '.
  - Dates are ISO 8601 strings with time: '2026-03-02T00:00:00.000'.
  - Money columns: 'Debits' (outflows) and 'Credits' (inflows).

Output: list[ParsedTransaction] dataclass instances. Errors raised as
BankStatementParseError - no silent skipping, no fabricated values.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

try:
    import openpyxl
except ImportError as exc:
    raise ImportError(
        "openpyxl is required. Install with: pip install openpyxl"
    ) from exc


class BankStatementParseError(Exception):
    """Raised when a bank statement cannot be parsed cleanly."""


@dataclass
class StatementHeader:
    account_name: Optional[str]
    account_number: Optional[str]
    currency: Optional[str]
    period: Optional[str]


@dataclass
class ParsedTransaction:
    sheet_name: str
    transaction_id: str
    transaction_date: date
    posted_date: Optional[datetime]
    description: str
    check_number: Optional[str]
    debit_amount: Decimal       # outflow
    credit_amount: Decimal      # inflow
    net_amount: Decimal         # credit - debit (positive = inflow)
    ending_balance: Optional[Decimal]
    reference_number: Optional[str]
    remarks: Optional[str]
    branch: Optional[str]
    biller_name: Optional[str]
    counterparty_name: Optional[str]   # extracted from "Sent to <NAME>" pattern

    def as_dict(self) -> dict:
        d = asdict(self)
        d["transaction_date"] = self.transaction_date.isoformat()
        if self.posted_date:
            d["posted_date"] = self.posted_date.isoformat()
        for k in ("debit_amount", "credit_amount", "net_amount", "ending_balance"):
            v = d[k]
            d[k] = str(v) if v is not None else None
        return d


_REQUIRED_COLUMNS = (
    "Transaction Date", "Posted Date", "Transaction ID",
    "Transaction Description", "Check Number",
    "Debits", "Credits", "Ending Balance",
    "Reference Number", "Remarks", "Branch",
)


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def _parse_decimal(v, *, field: str, row_num: int, sheet: str) -> Decimal:
    if _is_blank(v):
        return Decimal("0")
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError) as exc:
        raise BankStatementParseError(
            f"[{sheet}] Row {row_num}: cannot parse {field}={v!r} as a number ({exc})"
        ) from exc


def _parse_iso_date(v, *, field: str, row_num: int, sheet: str,
                    required: bool = True) -> Optional[datetime]:
    if _is_blank(v):
        if required:
            raise BankStatementParseError(
                f"[{sheet}] Row {row_num}: missing required date field {field!r}"
            )
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime.combine(v, datetime.min.time())
    s = str(v).strip()
    s = re.sub(r"\.\d+$", "", s)
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise BankStatementParseError(
        f"[{sheet}] Row {row_num}: cannot parse {field}={v!r} as a date"
    )


_SENT_TO_RE = re.compile(
    r"^Sent to\s+(?P<name>.+?)\s+(?:[A-Z]{3})\s+\d+\s*$"
)
_RECEIVED_FROM_RE = re.compile(
    r"^Received from\s+(?P<name>.+?)\s+(?:[A-Z]{3})\s+\d+\s*$"
)


def _extract_counterparty(description: str) -> Optional[str]:
    if not description:
        return None
    m = _SENT_TO_RE.match(description.strip())
    if m:
        return m.group("name").strip()
    m = _RECEIVED_FROM_RE.match(description.strip())
    if m:
        return m.group("name").strip()
    return None


def parse_statement(file_path) -> tuple[StatementHeader, List[ParsedTransaction]]:
    """
    Parse an entire UB workbook. Returns (header, transactions). When the
    same Transaction ID appears in multiple sheets, the version with
    populated amounts wins; if all are blank, the first-seen wins.
    """
    path = Path(file_path)
    if not path.exists():
        raise BankStatementParseError(f"File not found: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise BankStatementParseError(
            f"Expected .xlsx or .xlsm, got {path.suffix}: {path}"
        )

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    header: Optional[StatementHeader] = None
    seen_ids: dict[str, ParsedTransaction] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if header is None:
            header = _parse_header_block(ws)

        marker_row = _find_transactions_marker(ws, sheet_name)
        col_map = _build_column_map(ws, header_row=marker_row + 1, sheet_name=sheet_name)

        for raw_row_idx, row in enumerate(
            ws.iter_rows(min_row=marker_row + 2, max_row=ws.max_row, values_only=True),
            start=marker_row + 2,
        ):
            if _is_blank(row[col_map["Transaction Date"]]):
                continue

            tx = _parse_row(row, col_map=col_map, row_num=raw_row_idx,
                            sheet_name=sheet_name)

            existing = seen_ids.get(tx.transaction_id)
            if existing is None:
                seen_ids[tx.transaction_id] = tx
            else:
                existing_has_amt = existing.debit_amount > 0 or existing.credit_amount > 0
                new_has_amt = tx.debit_amount > 0 or tx.credit_amount > 0
                if new_has_amt and not existing_has_amt:
                    seen_ids[tx.transaction_id] = tx

    if header is None:
        raise BankStatementParseError("Workbook has no readable header section.")

    return header, list(seen_ids.values())


def _parse_header_block(ws) -> StatementHeader:
    found = {}
    for row in ws.iter_rows(min_row=1, max_row=14, max_col=2, values_only=True):
        key, value = (row[0], row[1] if len(row) > 1 else None)
        if isinstance(key, str):
            found[key.strip().rstrip(":").lower()] = value
    return StatementHeader(
        account_name=found.get("account name"),
        account_number=str(found.get("account number")) if found.get("account number") else None,
        currency=found.get("currency"),
        period=found.get("period covered"),
    )


def _find_transactions_marker(ws, sheet_name: str) -> int:
    for r in range(1, 30):
        cell = ws.cell(row=r, column=1).value
        if isinstance(cell, str) and "TRANSACTIONS LIST" in cell:
            return r
    raise BankStatementParseError(
        f"[{sheet_name}] Could not find 'TRANSACTIONS LIST:' marker in first 30 rows."
    )


def _build_column_map(ws, *, header_row: int, sheet_name: str) -> dict:
    headers = [c.value for c in ws[header_row]]
    col_map: dict = {}
    for idx, h in enumerate(headers):
        if isinstance(h, str):
            col_map[h.strip()] = idx
    missing = [c for c in _REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise BankStatementParseError(
            f"[{sheet_name}] Header row {header_row} is missing required columns: {missing}. "
            f"Got: {list(col_map.keys())}"
        )
    return col_map


def _parse_row(row, *, col_map: dict, row_num: int,
               sheet_name: str) -> ParsedTransaction:
    def col(name: str):
        idx = col_map.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    tx_id = col("Transaction ID")
    if _is_blank(tx_id):
        raise BankStatementParseError(
            f"[{sheet_name}] Row {row_num}: missing Transaction ID"
        )
    tx_id = str(tx_id).strip()

    tx_date_dt = _parse_iso_date(col("Transaction Date"),
                                  field="Transaction Date",
                                  row_num=row_num, sheet=sheet_name)
    posted_dt = _parse_iso_date(col("Posted Date"),
                                 field="Posted Date",
                                 row_num=row_num, sheet=sheet_name,
                                 required=False)

    description = col("Transaction Description")
    if _is_blank(description):
        raise BankStatementParseError(
            f"[{sheet_name}] Row {row_num}: missing Transaction Description"
        )
    description = str(description).strip()

    debit  = _parse_decimal(col("Debits"),  field="Debits",
                            row_num=row_num, sheet=sheet_name)
    credit = _parse_decimal(col("Credits"), field="Credits",
                            row_num=row_num, sheet=sheet_name)
    net    = credit - debit

    ending_balance_raw = col("Ending Balance")
    ending_balance = (_parse_decimal(ending_balance_raw,
                                     field="Ending Balance",
                                     row_num=row_num, sheet=sheet_name)
                      if not _is_blank(ending_balance_raw) else None)

    def _str_or_none(v):
        if _is_blank(v):
            return None
        return str(v).strip()

    return ParsedTransaction(
        sheet_name        = sheet_name,
        transaction_id    = tx_id,
        transaction_date  = tx_date_dt.date(),
        posted_date       = posted_dt,
        description       = description,
        check_number      = _str_or_none(col("Check Number")),
        debit_amount      = debit,
        credit_amount     = credit,
        net_amount        = net,
        ending_balance    = ending_balance,
        reference_number  = _str_or_none(col("Reference Number")),
        remarks           = _str_or_none(col("Remarks")),
        branch            = _str_or_none(col("Branch")),
        biller_name       = _str_or_none(col("Biller Name")) if "Biller Name" in col_map else None,
        counterparty_name = _extract_counterparty(description),
    )
