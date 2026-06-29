"""
Combined GGR-reconciliation .xls parser.

The user's monthly file packages BOTH sources in one .xls workbook:
  Sheet 1: 'PER SYSTEM' - subsidiary ledger (account, voucher, debit, credit)
  Sheet 2: 'PER GGR'    - per-week per-outlet GGR/PAGCOR/Audit/Operator amounts

PER GGR sheet layout:
    R1: Period header
    R2: column header (Operator Name | Outlet Name | Outlet Code | GGR | PAGCOR | Audit | Operator)
    R3: 'WEEK 1 (mm/dd/yyyy to mm/dd/yyyy)'  in col 2
    R4..R29: outlet rows for that week (RAC PHIL CORP. on first row of week)
    Repeats per week.
"""
from __future__ import annotations

import re
import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional, Tuple

try:
    import xlrd
except ImportError as e:
    raise ImportError("xlrd is required: pip install xlrd") from e


class CombinedParseError(Exception):
    pass


# ---- PER GGR rows -----------------------------------------------------------

@dataclass
class GGRWeekRow:
    week_label: str           # 'Week 1', 'Week 2', ...
    week_number: int
    week_start: Optional[date]
    week_end: Optional[date]
    outlet_code: str
    outlet_name: str
    operator_name: Optional[str]
    ggr: Decimal
    pagcor_share: Decimal
    audit_fee: Decimal
    operator: Decimal


# ---- PER SYSTEM rows --------------------------------------------------------

@dataclass
class SystemEntry:
    account_code: str
    account_title: str
    voucher_date: Optional[date]
    voucher_no: str
    department: Optional[str]
    description: str
    week_label: Optional[str]    # parsed from 'WEEK N' inside description
    debit: Decimal
    credit: Decimal


@dataclass
class CombinedFile:
    period_label: Optional[str]
    ggr_rows: List[GGRWeekRow] = field(default_factory=list)
    system_entries: List[SystemEntry] = field(default_factory=list)


class _Cell:
    def __init__(self, value):
        self.value = value


class _MatrixSheet:
    """Small adapter so CSV/openpyxl sheets can be parsed like xlrd sheets."""

    def __init__(self, rows):
        self._rows = rows
        self.nrows = len(rows)

    def cell(self, r: int, c: int):
        if r < 0 or c < 0 or r >= len(self._rows):
            return _Cell("")
        row = self._rows[r]
        if c >= len(row):
            return _Cell("")
        return _Cell(row[c])


def _load_sheets(file_path) -> List[Tuple[str, object]]:
    """Load .xls/.xlsx/.xlsm/.csv into sheet-like objects.

    The detailed GGR screen now supports two separate uploads: one PER GGR/raw
    file and one PER SYSTEM file. This loader keeps the existing .xls parser
    behavior but also allows real Excel workbooks with separate sheets.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        wb = xlrd.open_workbook(path)
        return [(name, wb.sheet_by_name(name)) for name in wb.sheet_names()]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError as exc:
            raise CombinedParseError("openpyxl is required for .xlsx/.xlsm files") from exc
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheets = []
            for ws in wb.worksheets:
                rows = [list(row) for row in ws.iter_rows(values_only=True)]
                sheets.append((ws.title, _MatrixSheet(rows)))
        finally:
            wb.close()
        return sheets
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        return [(path.stem, _MatrixSheet(rows))]
    raise CombinedParseError(f"Unsupported file type: {suffix}. Use .xls, .xlsx, .xlsm, or .csv")


def _first_parseable_sheet(file_path, parser, label: str):
    errors = []
    for sheet_name, sheet in _load_sheets(file_path):
        try:
            return parser(sheet)
        except CombinedParseError as exc:
            errors.append(f"{sheet_name}: {exc}")
    raise CombinedParseError(f"{label}: no parseable sheet found. " + "; ".join(errors))


# ---- helpers ----------------------------------------------------------------

_WEEK_HEADER_RE = re.compile(
    r"WEEK\s*(\d+)\s*\(\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*to\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*\)",
    re.IGNORECASE,
)
_WEEK_IN_DESC_RE = re.compile(r"W(?:EEK|K)?[\s.#-]*(\d+)", re.IGNORECASE)


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal("0")
    s = str(v).strip()
    if s == "":
        return Decimal("0")
    try:
        return Decimal(s.replace(",", ""))
    except Exception:
        return Decimal("0")


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _xldate_to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and v > 0:
        try:
            return xlrd.xldate.xldate_as_datetime(v, 0).date()
        except Exception:
            return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---- main parse -------------------------------------------------------------

def parse_combined_xls(file_path) -> CombinedFile:
    path = Path(file_path)
    if not path.exists():
        raise CombinedParseError(f"File not found: {path}")
    if path.suffix.lower() != ".xls":
        raise CombinedParseError(f"Expected .xls, got {path.suffix}")

    wb = xlrd.open_workbook(path)
    sheet_names = wb.sheet_names()

    sys_sheet = next((s for s in sheet_names if "PER SYSTEM" in s.upper()), None)
    ggr_sheet = next((s for s in sheet_names if "PER GGR" in s.upper()), None)
    if sys_sheet is None or ggr_sheet is None:
        raise CombinedParseError(
            f"Workbook must contain 'PER SYSTEM' and 'PER GGR' sheets. "
            f"Got: {sheet_names}")

    cf = CombinedFile(period_label=None)
    cf.system_entries = _parse_per_system(wb.sheet_by_name(sys_sheet))
    cf.ggr_rows, cf.period_label = _parse_per_ggr(wb.sheet_by_name(ggr_sheet))
    return cf


def parse_two_file_ggr(raw_ggr_file, per_system_file) -> CombinedFile:
    """Parse separate detailed GGR source files.

    raw_ggr_file: PER GGR/raw GGR Excel or CSV file.
    per_system_file: PER SYSTEM/subsidiary-ledger Excel or CSV file.

    No sample data or fallback rows are generated. If either file cannot be
    parsed into the expected structure, CombinedParseError is raised.
    """
    raw_path = Path(raw_ggr_file)
    system_path = Path(per_system_file)
    if not raw_path.exists():
        raise CombinedParseError(f"Raw GGR file not found: {raw_path}")
    if not system_path.exists():
        raise CombinedParseError(f"Per-system file not found: {system_path}")

    ggr_rows, period_label = _first_parseable_sheet(raw_path, _parse_per_ggr, "Raw GGR file")
    system_entries = _first_parseable_sheet(system_path, _parse_per_system, "Per-system file")
    if not ggr_rows:
        raise CombinedParseError("Raw GGR file parsed successfully but contains no GGR rows")
    if not system_entries:
        raise CombinedParseError("Per-system file parsed successfully but contains no system entries")
    return CombinedFile(period_label=period_label, ggr_rows=ggr_rows, system_entries=system_entries)


# ---- PER SYSTEM parser ------------------------------------------------------

def _parse_per_system(ws) -> List[SystemEntry]:
    # Find header row (col 1 == 'A/C No.')
    header_row = None
    for r in range(min(15, ws.nrows)):
        v = ws.cell(r, 0).value
        if isinstance(v, str) and "A/C No" in v:
            header_row = r
            break
    if header_row is None:
        raise CombinedParseError("PER SYSTEM: 'A/C No.' header row not found")

    entries: List[SystemEntry] = []
    current_code = None
    current_title = None
    for r in range(header_row + 1, ws.nrows):
        ac_raw = ws.cell(r, 0).value
        title_raw = ws.cell(r, 1).value
        ac_text = _to_str(ac_raw)
        if isinstance(ac_raw, (int, float)) and ac_raw > 0:
            current_code = f"{int(ac_raw)}"
            current_title = _to_str(title_raw)
            continue
        if ac_text.isdigit():
            current_code = ac_text
            current_title = _to_str(title_raw)
            continue
        if current_code is None:
            continue
        v_date = _xldate_to_date(ws.cell(r, 3).value)
        voucher = _to_str(ws.cell(r, 4).value)
        dept = _to_str(ws.cell(r, 5).value) or None
        desc = _to_str(ws.cell(r, 7).value)
        debit = _to_decimal(ws.cell(r, 9).value)
        credit = _to_decimal(ws.cell(r, 10).value)
        if debit == 0 and credit == 0:
            continue

        week_label = None
        m = _WEEK_IN_DESC_RE.search(desc)
        if m:
            week_label = f"Week {int(m.group(1))}"

        entries.append(SystemEntry(
            account_code=current_code, account_title=current_title or current_code,
            voucher_date=v_date, voucher_no=voucher, department=dept,
            description=desc, week_label=week_label,
            debit=debit, credit=credit,
        ))
    return entries


# ---- PER GGR parser ---------------------------------------------------------

def _parse_per_ggr(ws):
    # Find header row (col 1 == 'Operator Name')
    header_row = None
    period_label = None
    for r in range(min(10, ws.nrows)):
        v = ws.cell(r, 0).value
        if isinstance(v, str):
            if "Period Covered" in v:
                period_label = v.split(":", 1)[-1].strip()
            if "Operator Name" in v:
                header_row = r
                break
    if header_row is None:
        raise CombinedParseError("PER GGR: 'Operator Name' header row not found")

    rows: List[GGRWeekRow] = []
    current_week_label = None
    current_week_no = 0
    current_week_start = None
    current_week_end = None
    current_operator = None

    for r in range(header_row + 1, ws.nrows):
        op = _to_str(ws.cell(r, 0).value)
        outlet_name = _to_str(ws.cell(r, 1).value)
        outlet_code = _to_str(ws.cell(r, 2).value)

        # Week separator row: outlet_name has 'WEEK N (...)'
        m = _WEEK_HEADER_RE.search(outlet_name)
        if m:
            current_week_no = int(m.group(1))
            current_week_label = f"Week {current_week_no}"
            try:
                current_week_start = date(int(m.group(4)), int(m.group(2)), int(m.group(3)))
                current_week_end   = date(int(m.group(7)), int(m.group(5)), int(m.group(6)))
            except Exception:
                current_week_start = current_week_end = None
            continue

        # Outlet rows must have an outlet code starting with 'LW'
        if not outlet_code.upper().startswith("LW"):
            continue
        if current_week_label is None:
            continue
        if op:
            current_operator = op

        rows.append(GGRWeekRow(
            week_label=current_week_label,
            week_number=current_week_no,
            week_start=current_week_start,
            week_end=current_week_end,
            outlet_code=outlet_code,
            outlet_name=outlet_name.strip().rstrip("\t").strip(),
            operator_name=current_operator,
            ggr=_to_decimal(ws.cell(r, 3).value),
            pagcor_share=_to_decimal(ws.cell(r, 4).value),
            audit_fee=_to_decimal(ws.cell(r, 5).value),
            operator=_to_decimal(ws.cell(r, 6).value),
        ))
    return rows, period_label
