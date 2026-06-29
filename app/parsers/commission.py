"""Sub-affiliate commission computation.

Computes commission shares from the raw Sub Affiliate CSV and optionally offsets
prior negative balances from an earlier commission workbook/export.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import csv
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CommissionParseError(Exception):
    pass


MONEY = Decimal("0.01")
RATES = {
    "shop": Decimal("0.25"),
    "dsp": Decimal("0.03"),
    "supervisor": Decimal("0.01"),
    "manager": Decimal("0.01"),
}
LABELS = {
    "shop": "Shop",
    "dsp": "DSP",
    "supervisor": "Supervisor",
    "manager": "Manager",
}


def _money(value: Decimal) -> Decimal:
    return value.quantize(MONEY, rounding=ROUND_HALF_UP)


def _dec(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace(",", "")
    if text in {"", "-"}:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _text(value) -> str:
    return "" if value is None else str(value).strip()


@dataclass
class CommissionShare:
    rate: Decimal
    positive: Decimal = Decimal("0")
    negative: Decimal = Decimal("0")
    carried: Decimal = Decimal("0")
    net: Decimal = Decimal("0")
    next_carry: Decimal = Decimal("0")


@dataclass
class CommissionRow:
    sub_affiliate: str
    sub_affiliate_name: str
    ggr: Decimal
    promotion: Decimal
    base: Decimal
    shares: Dict[str, CommissionShare] = field(default_factory=dict)
    dsp_tag: str = ""          # e.g. "dsp01" — populated when the source file has a DSP column

    @property
    def total_payable(self) -> Decimal:
        return sum((s.net for s in self.shares.values()), Decimal("0"))

    @property
    def total_negative_next_carry(self) -> Decimal:
        return sum((s.next_carry for s in self.shares.values() if s.next_carry < 0), Decimal("0"))


@dataclass
class CommissionResult:
    source_filename: str
    previous_filename: str = ""
    rows: List[CommissionRow] = field(default_factory=list)
    sup_mgr_total_basis: bool = False          # True = May+ (total-based netting, partner bears losses); False = Feb-Apr (per-store carry, same as shop/DSP)
    prior_total_carry: Dict[str, Decimal] = field(default_factory=dict)   # {"supervisor": Decimal("-250.00"), ...} — aggregate negative carry from the previous cutoff's TOTAL row (total-basis mode only)

    @property
    def totals(self) -> Dict[str, Decimal]:
        totals = {"ggr": Decimal("0"), "promotion": Decimal("0"), "base": Decimal("0"), "total_payable": Decimal("0"), "next_negative_carry": Decimal("0")}
        for key in RATES:
            totals[f"{key}_positive"] = Decimal("0")
            totals[f"{key}_negative"] = Decimal("0")
            totals[f"{key}_carried"] = Decimal("0")
            totals[f"{key}_net"] = Decimal("0")
            totals[f"{key}_next_carry"] = Decimal("0")
            totals[f"{key}_period_raw"] = Decimal("0")
            totals[f"{key}_prior_carry_in"] = Decimal("0")
        for row in self.rows:
            totals["ggr"] += row.ggr
            totals["promotion"] += row.promotion
            totals["base"] += row.base
            for key, share in row.shares.items():
                totals[f"{key}_positive"] += share.positive
                totals[f"{key}_negative"] += share.negative
                totals[f"{key}_carried"] += share.carried
                totals[f"{key}_net"] += share.net
                totals[f"{key}_next_carry"] += share.next_carry

        # May+ (sup_mgr_total_basis=True): supervisor/manager are total-based — offset happens
        # within the cut-off total (partner bears losses), so positive stores absorb negative stores.
        # Feb-Apr (sup_mgr_total_basis=False): per-store carry applies, same as shop/DSP.
        # Expose each role's "within this cut-off" figure (before any prior
        # period carry-in is folded in). For Shop/DSP, and for Supervisor/
        # Manager under the per-store carry mode, this is just the net.
        for key in RATES:
            totals[f"{key}_period_raw"] = totals[f"{key}_net"]

        if self.sup_mgr_total_basis:
            for key in ("supervisor", "manager"):
                # "Within this cut-off" raw total — before the prior cutoff's
                # negative carry-in is applied. Exposed separately so exports
                # can show the this-period / prior-carry / final breakdown
                # instead of only the post-offset figure.
                period_raw = _money(
                    totals[f"{key}_positive"] + totals[f"{key}_negative"] + totals[f"{key}_carried"]
                )
                totals[f"{key}_period_raw"] = period_raw

                # Pull in the negative aggregate carry from the previous
                # cutoff's TOTAL row so it offsets against this cutoff's
                # total before zeroing/netting.
                prior_total = _money(self.prior_total_carry.get(key, Decimal("0")))
                totals[f"{key}_prior_carry_in"] = prior_total
                totals[f"{key}_carried"] += prior_total
                raw_total = _money(period_raw + prior_total)
                totals[f"{key}_net"]        = raw_total if raw_total > 0 else Decimal("0")
                totals[f"{key}_next_carry"] = raw_total if raw_total < 0 else Decimal("0")

        totals["total_payable"] = sum((totals[f"{key}_net"] for key in RATES), Decimal("0"))
        totals["next_negative_carry"] = sum((totals[f"{key}_next_carry"] for key in RATES if totals[f"{key}_next_carry"] < 0), Decimal("0"))
        return totals

    @property
    def negative_carry_rows(self) -> List[CommissionRow]:
        return [r for r in self.rows if r.total_negative_next_carry < 0 or r.shares.get("shop", CommissionShare(Decimal("0"))).next_carry < 0]


def _norm_header(value) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\n\r\t_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9/ ]+", " ", text)
    return " ".join(text.split())


def _find_header(headers: List[str], candidates: List[str]) -> int:
    normalized = [_norm_header(h) for h in headers]
    aliases = [_norm_header(c) for c in candidates]
    for cand in aliases:
        if cand in normalized:
            return normalized.index(cand)
    for i, header in enumerate(normalized):
        for cand in aliases:
            if cand and (cand in header or header in cand):
                return i
    raise CommissionParseError(f"Missing required column. Expected one of: {', '.join(candidates)}. Detected: {', '.join(str(h) for h in headers if str(h).strip())}")


def _read_table_rows(path: Path) -> List[List[object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [row for row in csv.reader(f)]
    if suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            best = []
            for ws in wb.worksheets:
                rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
                if len(rows) > len(best):
                    best = rows
        finally:
            wb.close()
        return best
    if suffix == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise CommissionParseError("xlrd is required to read .xls commission files. Install requirements first.") from exc
        wb = xlrd.open_workbook(str(path))
        sheet = max(wb.sheets(), key=lambda sh: sh.nrows)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    raise CommissionParseError("Please upload Sub Affiliate report as .csv, .xlsx, .xlsm, or .xls.")


def _detect_commission_header(rows: List[List[object]]) -> tuple[int, int, int, int, int, int]:
    """Return (header_row_idx, idx_id, idx_name, idx_ggr, idx_promo, idx_dsp).
    idx_promo or idx_dsp will be -1 when the column is absent."""
    id_alias    = ["Sub Affiliate", "Sub Affiliate ID", "Affiliate", "Store", "Outlet",
                   "Site", "Venue", "Branch", "Location", "Account"]
    name_alias  = ["Sub Affiliate Name", "Affiliate Name", "Store Name", "Outlet Name",
                   "Site Name", "Venue Name", "Name", "Description"]
    ggr_alias   = ["Company Net Win (GGR)", "Company Net Win", "Net Win", "GGR",
                   "Gross Gaming Revenue", "Gross Gaming", "Win/Loss", "NGR"]
    promo_alias = ["Total Promotion Amount", "Promotion", "Promotions", "Promo Amount",
                   "Total Promo", "Bonus", "Bonus Amount", "Promo"]
    dsp_alias   = ["DSP", "DSP ID", "DSP Tag", "DSP Account", "DSP Name",
                   "Upline DSP", "Parent DSP", "Upline", "Sponsor DSP", "DSP Code"]

    def _try_dsp(row):
        try:
            return _find_header(row, dsp_alias)
        except CommissionParseError:
            return -1

    for row_idx, row in enumerate(rows[:50]):
        try:
            idx_id = _find_header(row, id_alias)
            try:
                idx_name = _find_header(row, name_alias)
            except CommissionParseError:
                idx_name = idx_id
            idx_ggr   = _find_header(row, ggr_alias)
            idx_promo = _find_header(row, promo_alias)
            return row_idx, idx_id, idx_name, idx_ggr, idx_promo, _try_dsp(row)
        except CommissionParseError:
            continue
    # Fallback: no promotion column
    for row_idx, row in enumerate(rows[:50]):
        try:
            idx_id = _find_header(row, id_alias)
            try:
                idx_name = _find_header(row, name_alias)
            except CommissionParseError:
                idx_name = idx_id
            idx_ggr = _find_header(row, ggr_alias)
            return row_idx, idx_id, idx_name, idx_ggr, -1, _try_dsp(row)
        except CommissionParseError:
            continue
    raise CommissionParseError(
        "Could not detect commission headers. "
        "Need Sub Affiliate/Store/Site and GGR/Net Win columns. Promotion and DSP are optional."
    )

def _extract_previous_carry(
    previous_path: Optional[Path],
) -> Tuple[Dict[str, Dict[str, Decimal]], Dict[str, Decimal]]:
    """Read per-store carry-forward amounts from the previous commission export.

    Strategy (most reliable first):
    1. Read the dedicated "Carry Forward" sheet — it has one row per store with
       explicit Share columns (Shop, DSP, Supervisor, Manager).
    2. Fall back to scanning the main commission sheet for "Carry to Next Month"
       columns positionally (net_col + 1).

    Returns a 2-tuple: (per_store_carries, total_row_carries).
    `total_row_carries` holds the negative aggregate Supervisor/Manager carry
    read from the previous export's TOTAL row — the only place that holds the
    real aggregate value when sup_mgr_total_basis is in effect (per-store rows
    always show 0 for those two roles in that mode).
    """
    if not previous_path or not previous_path.exists():
        return {}, {}
    if previous_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}, {}

    # Use data_only=True (no read_only) so styled cells return real values.
    # read_only mode can return None for cells that have borders/fills applied.
    wb = openpyxl.load_workbook(previous_path, data_only=True)
    try:
        all_sheets = list(wb.worksheets)

        # ── Strategy 1: dedicated "Carry Forward" sheet ──────────────────────
        # This sheet has one row per store with explicit per-share carry columns.
        cf_ws = next(
            (ws for ws in all_sheets
             if "carry" in ws.title.lower() and "forward" in ws.title.lower()),
            None,
        )
        if cf_ws is not None:
            rows = list(cf_ws.iter_rows(values_only=True))
            if len(rows) >= 2:
                hdr = [_norm_header(v) for v in (rows[0] or [])]
                id_col = next(
                    (i for i, h in enumerate(hdr)
                     if "sub" in h or "affiliate" in h),
                    0,
                )
                share_cols: Dict[str, int] = {}
                for key in ("shop", "dsp", "supervisor", "manager"):
                    col = next(
                        (i for i, h in enumerate(hdr)
                         if key in h and ("carry" in h or "next" in h)),
                        None,
                    )
                    if col is not None:
                        share_cols[key] = col

                carries: Dict[str, Dict[str, Decimal]] = {}
                total_row_carries: Dict[str, Decimal] = {}
                for row in rows[1:]:
                    if not row:
                        continue
                    sub_id = _text(row[id_col] if len(row) > id_col else None)
                    if not sub_id:
                        continue
                    if sub_id.upper() in {"TOTAL", "GRAND TOTAL"}:
                        # The TOTAL row holds the real aggregate negative carry
                        # for Supervisor/Manager under total-basis netting —
                        # capture it instead of discarding it, so the next
                        # cutoff can offset against it.
                        for key, cidx in share_cols.items():
                            raw = row[cidx] if len(row) > cidx else None
                            val = _dec(raw)
                            if val < 0:
                                total_row_carries[key] = _money(val)
                        continue
                    row_carries: Dict[str, Decimal] = {}
                    for key, cidx in share_cols.items():
                        raw = row[cidx] if len(row) > cidx else None
                        val  = _dec(raw)
                        row_carries[key] = _money(val if val < 0 else Decimal("0"))
                    if any(v < 0 for v in row_carries.values()):
                        carries[sub_id.lower()] = row_carries
                if carries or total_row_carries:
                    return carries, total_row_carries

        # ── Strategy 2: scan main commission sheet ───────────────────────────
        # Finds "Net Shop/DSP/..." columns then reads the column immediately after
        # (which is always "Carry to Next Month" in our layout).
        main_ws = next(
            (ws for ws in all_sheets
             if "summary" not in ws.title.lower()
             and "carry" not in ws.title.lower()),
            None,
        )
        if main_ws is None:
            return {}, {}

        net_cols:        Dict[str, int] = {}
        next_carry_cols: Dict[str, int] = {}
        header_row_idx = None

        for ridx, row in enumerate(
            main_ws.iter_rows(min_row=1, max_row=10, values_only=True)
        ):
            if not row:
                continue
            norm = [_norm_header(v) for v in row]
            for cidx, h in enumerate(norm):
                if "net" in h and "shop" in h and "shop" not in net_cols:
                    net_cols["shop"] = cidx;  header_row_idx = ridx
                if "net" in h and "dsp"  in h and "dsp"  not in net_cols:
                    net_cols["dsp"] = cidx
                    if header_row_idx is None: header_row_idx = ridx
                if "net" in h and "supervisor" in h and "supervisor" not in net_cols:
                    net_cols["supervisor"] = cidx
                if "net" in h and "manager" in h and "manager" not in net_cols:
                    net_cols["manager"] = cidx
                for key in ("shop", "dsp", "supervisor", "manager"):
                    if (("next" in h or "carry" in h) and key in h
                            and key not in next_carry_cols):
                        next_carry_cols[key] = cidx

        for key in ("shop", "dsp", "supervisor", "manager"):
            if key in net_cols and key not in next_carry_cols:
                next_carry_cols[key] = net_cols[key] + 1

        if not net_cols:
            return {}, {}

        data_start = (header_row_idx + 2) if header_row_idx is not None else 7
        carries = {}
        for values in main_ws.iter_rows(min_row=data_start, values_only=True):
            if not values:
                continue
            sub_id = _text(values[0])
            if not sub_id or sub_id.upper() in {"TOTAL", "GRAND TOTAL"}:
                continue
            row_carries: Dict[str, Decimal] = {}
            for key in ("shop", "dsp", "supervisor", "manager"):
                cidx = next_carry_cols.get(key) if key in next_carry_cols \
                       else net_cols.get(key)
                if cidx is None:
                    continue
                val = _dec(values[cidx] if len(values) > cidx else None)
                row_carries[key] = _money(val if val < 0 else Decimal("0"))
            carries[sub_id.lower()] = row_carries
        # Strategy 2's main sheet has no aggregate carry column for
        # Supervisor/Manager in the current layout, so there is no TOTAL-row
        # value to recover here — return an empty total_row_carries.
        return carries, {}
    finally:
        wb.close()


def _extract_dsp_group_carry(previous_path: Optional[Path]) -> Dict[str, "Decimal"]:
    """Read per-DSP-group carry-forward amounts from a previous commission carry file.

    Looks for a sheet whose title contains 'dsp' and 'group', then reads
    DSP Tag + carry columns.  Returns {dsp_tag.lower(): Decimal}.
    """
    if not previous_path or not previous_path.exists():
        return {}
    if previous_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return {}
    try:
        wb = openpyxl.load_workbook(previous_path, data_only=True, read_only=True)
        dsp_ws = None
        for ws in wb.worksheets:
            if "dsp" in ws.title.lower() and "group" in ws.title.lower():
                dsp_ws = ws
                break
        if dsp_ws is None:
            return {}
        rows = list(dsp_ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {}
        hdrs = [_norm_header(v) for v in (rows[0] or [])]
        tag_col  = next((i for i, h in enumerate(hdrs) if "dsp" in h), None)
        carry_col = next((i for i, h in enumerate(hdrs)
                          if ("carry" in h or "next" in h) and "dsp" not in h), None)
        if tag_col is None or carry_col is None:
            return {}
        result: Dict[str, Decimal] = {}
        for row in rows[1:]:
            if not row:
                continue
            tag = _text(row[tag_col] if len(row) > tag_col else None)
            if not tag or tag.upper() in {"TOTAL", "GRAND TOTAL", "(NO TAG)"}:
                continue
            val = _dec(row[carry_col] if len(row) > carry_col else None)
            if val < 0:
                result[tag.lower()] = _money(val)
        return result
    except Exception:
        return {}


def parse_commission_report(raw_path, previous_path=None,
                            sup_mgr_total_basis: bool = False) -> CommissionResult:
    raw = Path(raw_path)
    if not raw.exists():
        raise CommissionParseError(f"File not found: {raw}")

    previous = Path(previous_path) if previous_path else None
    prev_carry, prev_total_carry = _extract_previous_carry(previous)
    rows_data = _read_table_rows(raw)
    if not rows_data:
        raise CommissionParseError("Commission file is empty.")
    header_row, idx_id, idx_name, idx_ggr, idx_promo, idx_dsp = _detect_commission_header(rows_data)
    rows: List[CommissionRow] = []
    for rec in rows_data[header_row + 1:]:
        required_idx = max(idx_id, idx_name, idx_ggr, idx_promo if idx_promo >= 0 else 0)
        if not rec or len(rec) <= required_idx:
            continue
        sub_id = _text(rec[idx_id])
        if not sub_id or sub_id.upper() in {"TOTAL", "GRAND TOTAL"}:
            continue
        ggr = _money(_dec(rec[idx_ggr]))
        promo = _money(_dec(rec[idx_promo])) if idx_promo >= 0 else Decimal("0")
        if ggr == 0 and promo == 0:
            continue
        dsp_tag = _text(rec[idx_dsp]) if idx_dsp >= 0 and idx_dsp < len(rec) else ""
        base = _money(ggr - promo)
        previous_for_row = prev_carry.get(sub_id.lower(), {})
        shares: Dict[str, CommissionShare] = {}
        for key, rate in RATES.items():
            raw_amount = _money(base * rate)
            positive = raw_amount if raw_amount > 0 else Decimal("0")
            negative = raw_amount if raw_amount < 0 else Decimal("0")
            # Feb-Apr (sup_mgr_total_basis=False): all roles use per-store carry (same as shop/DSP).
            # May+ (sup_mgr_total_basis=True): supervisor/manager are total-based —
            #   store the raw per-row amount with no carry; the totals property does
            #   the within-cut-off netting at the aggregate level instead.
            if key in {"supervisor", "manager"} and sup_mgr_total_basis:
                shares[key] = CommissionShare(
                    rate, positive, negative, Decimal("0"), raw_amount, Decimal("0")
                )
                continue
            else:
                prior = _money(previous_for_row.get(key, Decimal("0")))
            if positive > 0:
                combined = _money(prior + positive)
                net = combined if combined > 0 else Decimal("0")
            else:
                net = Decimal("0")
            # Always show prior carry so the offset is visible as an audit trail.
            # Example: prior=-378.55, positive=500 → carried=-378.55, net=121.45
            # The user can see the -378.55 was applied even though it was fully absorbed.
            carried_display = prior
            next_carry = _money(prior + positive + negative)
            if next_carry > 0:
                next_carry = Decimal("0")
            shares[key] = CommissionShare(rate, positive, negative, carried_display, net, next_carry)
        rows.append(CommissionRow(sub_id, _text(rec[idx_name]) or sub_id, ggr, promo, base, shares,
                                  dsp_tag=dsp_tag))

    if not rows:
        raise CommissionParseError("No commission rows were found after reading the detected headers.")
    return CommissionResult(raw.name, previous.name if previous else "", rows,
                            sup_mgr_total_basis=sup_mgr_total_basis,
                            prior_total_carry=prev_total_carry)

def export_commission_result(result: CommissionResult) -> bytes:
    """
    Column layout (1-based):
      1–5  : ID, Name, DSP Tag, GGR, Promotion
      6–10 : SHOP 25%       (+, -, Offset Prev, Net, Carry Next)   net=col 9
      11–15: DSP 3%         (+, -, Offset Prev, Net, Carry Next)   net=col 14
      16   : SUPERVISOR 1%  Net only                               net=col 16
      17   : MANAGER 1%     Net only                               net=col 17
    Total = 17 columns.
    """
    import io as _io

    TOTAL_COLS = 17

    def _fill(hex_): return PatternFill("solid", fgColor=hex_)

    # (start_col, end_col, label, net_col, banner_hex, subhdr_hex, net_tint_hex)
    GROUPS = [
        (6,  10, "SHOP (25%)",      9,  "7B5800", "FFF0C2", "FFE08A"),
        (11, 15, "DSP (3%)",        14, "2E5F1A", "E5F2DC", "C8E6C9"),
        (16, 16, "SUPERVISOR (1%)", 16, "37474F", "ECEFF1", "CFD8DC"),
        (17, 17, "MANAGER (1%)",    17, "5D1E0A", "F8EAE6", "F5C6B8"),
    ]

    # col -> (subhdr_fill, net_tint_fill, is_net, banner_hex)
    COL_META: dict = {}
    for (sc, ec, _, nc, banner, subhdr, net_tint) in GROUPS:
        for c in range(sc, ec + 1):
            COL_META[c] = (_fill(subhdr), _fill(net_tint), c == nc, banner)

    # Border styles
    thin   = Side(style="thin",   color="D0D0D0")
    medium = Side(style="medium", color="888888")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_grp = Border(left=medium, right=thin, top=thin, bottom=thin)  # group separator

    SEP_COLS  = {sc for sc, *_ in GROUPS}   # cols where a thick left border starts
    BASE_COLS = 5                            # ID, Name, DSP Tag, GGR, Promotion

    MONEY_FMT = '#,##0.00;[Red](#,##0.00);-'
    WRAP_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _bdr(ci): return bdr_grp if ci in SEP_COLS else bdr

    # ── Workbook ───────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Commission"

    # ── Row 1 : Report title ───────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    r1 = ws.cell(1, 1, "SUB-AFFILIATE COMMISSION REPORT")
    r1.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    r1.fill      = _fill("1A3C6E")
    r1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2 : Company / period line (blank — user can type here) ─────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    r2 = ws.cell(2, 1, "")          # leave blank for the user to fill in period
    r2.fill      = _fill("EBF0FA")
    r2.font      = Font(size=10, color="1A3C6E", name="Calibri")
    r2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3 : Formula note ───────────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    r3 = ws.cell(3, 1,
        "Formula: (GGR − Total Promotion) × rate.  "
        + (
            "All roles (Shop, DSP, Supervisor, Manager) offset applied per store with carry to next cut-off."
            if not result.sup_mgr_total_basis
            else
            "Shop/DSP: per-store carry.  "
            "Supervisor/Manager: net from cut-off total — partner bears losses, offset within cut-off."
        ))
    r3.font      = Font(italic=True, size=8, color="606060", name="Calibri")
    r3.fill      = _fill("F4F6FB")
    r3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 20

    # ── Row 4 : Source note ────────────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TOTAL_COLS)
    src_text = f"Source: {result.source_filename}"
    if result.previous_filename:
        src_text += f"   |   Previous carry: {result.previous_filename}"
    r4 = ws.cell(4, 1, src_text)
    r4.font      = Font(size=8, color="808080", name="Calibri")
    r4.fill      = _fill("F4F6FB")
    r4.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 14

    # ── Row 5 : Group banners ──────────────────────────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=BASE_COLS)
    b5 = ws.cell(5, 1, "SUB-AFFILIATE")
    b5.font      = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
    b5.fill      = _fill("2C3E50")
    b5.alignment = WRAP_CTR

    for (sc, ec, label, _, banner, *__) in GROUPS:
        ws.merge_cells(start_row=5, start_column=sc, end_row=5, end_column=ec)
        bg = ws.cell(5, sc, label)
        bg.font      = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        bg.fill      = _fill(banner)
        bg.alignment = WRAP_CTR
    ws.row_dimensions[5].height = 18

    # ── Row 6 : Column sub-headers ─────────────────────────────────────────────
    # Col 3 = DSP Tag (new)
    col_headers = ["Sub Affiliate\nID", "Sub Affiliate\nName", "DSP\nTag", "GGR", "Promotion"]
    for key in ["shop", "dsp"]:
        lbl = LABELS[key]
        col_headers += [f"(+)\n{lbl}", f"(−)\n{lbl}", f"Offset\nPrev. Month",
                        f"Net\n{lbl}", f"Carry to\nNext Month"]
    for key in ["supervisor", "manager"]:
        col_headers += [f"Net\n{LABELS[key]}"]

    BASE_HDR_FILL = _fill("2C3E50")
    BASE_HDR_FONT = Font(bold=True, size=9, color="FFFFFF", name="Calibri")

    for ci, h in enumerate(col_headers, 1):
        c = ws.cell(6, ci, h)
        c.alignment = WRAP_CTR
        c.border    = _bdr(ci)
        if ci <= BASE_COLS:
            c.fill = BASE_HDR_FILL
            c.font = BASE_HDR_FONT
        else:
            subhdr_fill, net_tint_fill, is_net, banner = COL_META[ci]
            if is_net:
                c.fill = _fill(banner)
                c.font = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            else:
                c.fill = subhdr_fill
                c.font = Font(bold=True, size=9, color="1A1A1A", name="Calibri")
    ws.row_dimensions[6].height = 38

    # ── Rows 7+ : Data ─────────────────────────────────────────────────────────
    DATA_START = 7
    r = DATA_START
    STRIPE   = _fill("F7F9FC")   # even-row stripe (very light cool grey)
    WHITE    = _fill("FFFFFF")
    DSP_FILL = _fill("F0F4FF")   # DSP tag column — very light blue tint
    ID_FONT  = Font(size=10, name="Calibri")
    NUM_FONT = Font(size=10, name="Calibri")
    NET_FONT = Font(bold=True, size=10, name="Calibri")
    DSP_FONT = Font(size=10, color="1A3C6E", name="Calibri")   # navy text for DSP tag

    for row in result.rows:
        stripe   = (r - DATA_START) % 2 == 1
        row_fill = STRIPE if stripe else WHITE

        # col 1=ID, 2=Name, 3=DSP Tag, 4=GGR, 5=Promotion, then share cols
        vals = [row.sub_affiliate, row.sub_affiliate_name, row.dsp_tag,
                float(row.ggr), float(row.promotion)]
        for key in ["shop", "dsp"]:
            s = row.shares[key]
            vals += [float(s.positive), float(s.negative), float(s.carried),
                     float(s.net), float(s.next_carry)]
        for key in ["supervisor", "manager"]:
            vals += [float(row.shares[key].net)]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(r, ci, v)
            cell.border = _bdr(ci)
            if ci <= 2:
                cell.alignment = Alignment(vertical="center", horizontal="left")
                cell.fill = row_fill
                cell.font = ID_FONT
            elif ci == 3:                          # DSP Tag column
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.fill = DSP_FILL
                cell.font = DSP_FONT
            else:
                cell.alignment = Alignment(vertical="center", horizontal="right")
                cell.number_format = MONEY_FMT
                cell.font = NUM_FONT
                if ci <= BASE_COLS:
                    cell.fill = row_fill
                else:
                    _, net_tint_fill, is_net, _ = COL_META[ci]
                    cell.fill = net_tint_fill if is_net else row_fill
                    if is_net:
                        cell.font = NET_FONT
        r += 1

    # ── Totals row ─────────────────────────────────────────────────────────────
    totals    = result.totals
    total_row = r
    TOT_FILL  = _fill("1A3C6E")
    TOT_FONT  = Font(bold=True, size=10, color="FFFFFF", name="Calibri")

    ws.cell(r, 1, "TOTAL")
    ws.cell(r, 2, "")          # Name blank
    ws.cell(r, 3, "")          # DSP Tag blank
    ws.cell(r, 4, float(totals["ggr"]))
    ws.cell(r, 5, float(totals["promotion"]))
    col = 6
    for key in ["shop", "dsp"]:
        for suffix in ["positive", "negative", "carried", "net"]:
            ws.cell(r, col, float(totals[f"{key}_{suffix}"])); col += 1
        ws.cell(r, col, float(totals[f"{key}_next_carry"])); col += 1
    for key in ["supervisor", "manager"]:
        # Use the within-this-cut-off raw figure here so the TOTAL row equals
        # the sum of the per-store Net Supervisor/Net Manager cells above it.
        # (totals[f"{key}_net"] may be post-offset against a prior carry-in —
        # see the breakdown rows added below.)
        ws.cell(r, col, float(totals[f"{key}_period_raw"])); col += 1

    for ci in range(1, TOTAL_COLS + 1):
        c = ws.cell(total_row, ci)
        c.border = _bdr(ci)
        c.font   = TOT_FONT
        c.alignment = Alignment(vertical="center",
                                horizontal="center" if ci == 3
                                else ("right" if ci >= 4 else "left"))
        if ci >= 4:
            c.number_format = MONEY_FMT
        if ci <= BASE_COLS:
            c.fill = TOT_FILL
        else:
            _, _, _, banner = COL_META[ci]
            c.fill = _fill(banner)
    ws.row_dimensions[total_row].height = 20

    # ── Supervisor/Manager prior-carry breakdown (only when applicable) ────────
    # If there's a nonzero prior-period carry-in for Supervisor/Manager (i.e.
    # last cutoff's aggregate total ended negative for that role), show the
    # breakdown: this cut-off's raw total (already on the TOTAL row above),
    # the prior carry-in being offset, and the final for-payment/carry figure.
    sm_prior_carry_in = {
        k: totals.get(f"{k}_prior_carry_in", Decimal("0")) for k in ("supervisor", "manager")
    }
    if any(v != 0 for v in sm_prior_carry_in.values()):
        SUP_COL, MGR_COL = 16, 17
        SUP_TINT   = _fill("CFD8DC")
        MGR_TINT   = _fill("F5C6B8")
        LABEL_FONT = Font(italic=True, size=9, color="5A5A5A", name="Calibri")
        FINAL_FONT = Font(bold=True, size=10, color="1A1A1A", name="Calibri")
        FINAL_FILL = _fill("FFF6E0")

        offset_row = total_row + 1
        final_row  = total_row + 2

        ws.cell(offset_row, 1, "OFFSET PREV MONTH")
        ws.merge_cells(start_row=offset_row, start_column=1, end_row=offset_row, end_column=15)
        ws.cell(offset_row, SUP_COL, float(sm_prior_carry_in["supervisor"]))
        ws.cell(offset_row, MGR_COL, float(sm_prior_carry_in["manager"]))
        for ci in (1, SUP_COL, MGR_COL):
            c = ws.cell(offset_row, ci)
            c.border = _bdr(ci)
            c.font = LABEL_FONT
            if ci == 1:
                c.alignment = Alignment(vertical="center", horizontal="left")
            else:
                c.alignment = Alignment(vertical="center", horizontal="right")
                c.number_format = MONEY_FMT
                c.fill = SUP_TINT if ci == SUP_COL else MGR_TINT
        ws.row_dimensions[offset_row].height = 18

        sup_final = totals["supervisor_net"] + totals["supervisor_next_carry"]
        mgr_final = totals["manager_net"] + totals["manager_next_carry"]
        ws.cell(final_row, 1, "FOR PAYMENT / FOR OFFSETTING NEXT CUT OFF")
        ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=15)
        ws.cell(final_row, SUP_COL, float(sup_final))
        ws.cell(final_row, MGR_COL, float(mgr_final))
        for ci in (1, SUP_COL, MGR_COL):
            c = ws.cell(final_row, ci)
            c.border = _bdr(ci)
            c.font = FINAL_FONT
            c.fill = FINAL_FILL
            if ci == 1:
                c.alignment = Alignment(vertical="center", horizontal="left")
            else:
                c.alignment = Alignment(vertical="center", horizontal="right")
                c.number_format = MONEY_FMT
        ws.row_dimensions[final_row].height = 20

    # ── Freeze, filter, column widths ─────────────────────────────────────────
    ws.freeze_panes = "D7"                 # freeze ID + Name + DSP Tag
    ws.auto_filter.ref = f"A6:{get_column_letter(TOTAL_COLS)}6"
    for ci in range(1, TOTAL_COLS + 1):
        if ci == 1:   w = 16
        elif ci == 2: w = 22
        elif ci == 3: w = 10   # DSP Tag
        elif ci in {4, 5}: w = 14
        else:         w = 13
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.sheet_view.showGridLines = False   # cleaner corporate look

    # ── DSP Summary sheet ─────────────────────────────────────────────────────
    # Groups stores by DSP tag.  Each store shows its per-store Net DSP.
    # The DSP tag is purely for identification — carry is handled per store
    # in the main sheet (same as Shop), not at the group level.
    has_dsp_tags = any(row.dsp_tag for row in result.rows)
    if has_dsp_tags:
        from collections import defaultdict as _dd
        dsp_ws = wb.create_sheet("DSP Summary")
        dsp_ws.sheet_view.showGridLines = False

        DSP_BANNER   = _fill("2E5F1A")
        DSP_GRP_FONT = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        DSP_SUB_FONT = Font(size=10, name="Calibri")
        DSP_NET_FILL = _fill("C8E6C9")
        DSP_NET_FONT = Font(bold=True, size=10, name="Calibri")
        DSP_TOT_FILL = _fill("1A3C6E")
        DSP_TOT_FONT = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
        DSP_STRIPE   = _fill("F0F7ED")
        NUM_DSP_COLS = 4   # DSP/Store | ID | Name | Net DSP

        # Title
        dsp_ws.merge_cells(f"A1:{get_column_letter(NUM_DSP_COLS)}1")
        dt = dsp_ws.cell(1, 1, "DSP COMMISSION SUMMARY — 3%")
        dt.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        dt.fill      = DSP_BANNER
        dt.alignment = Alignment(horizontal="center", vertical="center")
        dsp_ws.row_dimensions[1].height = 26

        dsp_ws.merge_cells(f"A2:{get_column_letter(NUM_DSP_COLS)}2")
        ds = dsp_ws.cell(2, 1, src_text)
        ds.font = Font(italic=True, size=8, color="606060", name="Calibri")
        ds.fill = _fill("F4F6FB")
        ds.alignment = Alignment(horizontal="left", vertical="center")
        dsp_ws.row_dimensions[2].height = 14

        dsp_col_hdrs = ["DSP / Store", "Sub Affiliate ID", "Sub Affiliate Name", "Net DSP"]
        for ci, h in enumerate(dsp_col_hdrs, 1):
            c = dsp_ws.cell(3, ci, h)
            c.font      = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
            c.fill      = _fill("2C3E50")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = bdr
        dsp_ws.row_dimensions[3].height = 22

        # Group rows by DSP tag — include stores with positive Net DSP
        all_groups: dict = _dd(list)
        for row in result.rows:
            if row.shares["dsp"].net > 0:
                all_groups[row.dsp_tag or "(no tag)"].append(row)
        sorted_tags = sorted(all_groups.keys())

        dr = 4
        grand_total = Decimal("0")
        for tag in sorted_tags:
            store_rows = sorted(all_groups[tag], key=lambda x: x.shares["dsp"].net, reverse=True)
            tag_total  = sum(rw.shares["dsp"].net for rw in store_rows)
            grand_total += tag_total

            # Group header row
            for ci in range(1, NUM_DSP_COLS + 1):
                c = dsp_ws.cell(dr, ci)
                c.fill = DSP_BANNER; c.font = DSP_GRP_FONT; c.border = bdr
                c.alignment = Alignment(vertical="center",
                                        horizontal="left" if ci == 1 else "right")
            dsp_ws.cell(dr, 1, f"▸  {tag}")
            dsp_ws.cell(dr, 4, float(tag_total)).number_format = MONEY_FMT
            dsp_ws.row_dimensions[dr].height = 18
            dr += 1

            # Store rows (indented)
            for i, rw in enumerate(store_rows):
                sf = DSP_STRIPE if i % 2 == 1 else _fill("FFFFFF")
                for ci, val, fill, font, align in [
                    (1, None,                     sf,          DSP_SUB_FONT, "left"),
                    (2, rw.sub_affiliate,          sf,          DSP_SUB_FONT, "left"),
                    (3, rw.sub_affiliate_name,     sf,          DSP_SUB_FONT, "left"),
                    (4, float(rw.shares["dsp"].net), DSP_NET_FILL, DSP_NET_FONT, "right"),
                ]:
                    cell = dsp_ws.cell(dr, ci, val)
                    cell.fill = fill; cell.font = font; cell.border = bdr
                    cell.alignment = Alignment(vertical="center", horizontal=align,
                                               indent=(2 if ci in {2, 3} else 0))
                    if ci == 4:
                        cell.number_format = MONEY_FMT
                dsp_ws.row_dimensions[dr].height = 16
                dr += 1

        # Grand total
        for ci in range(1, NUM_DSP_COLS + 1):
            c = dsp_ws.cell(dr, ci)
            c.fill = DSP_TOT_FILL; c.font = DSP_TOT_FONT; c.border = bdr
            c.alignment = Alignment(vertical="center",
                                    horizontal="left" if ci == 1 else "right")
        dsp_ws.cell(dr, 1, "GRAND TOTAL")
        dsp_ws.cell(dr, 4, float(grand_total)).number_format = MONEY_FMT
        dsp_ws.row_dimensions[dr].height = 20

        dsp_ws.column_dimensions["A"].width = 18
        dsp_ws.column_dimensions["B"].width = 18
        dsp_ws.column_dimensions["C"].width = 26
        dsp_ws.column_dimensions["D"].width = 16
        dsp_ws.freeze_panes = "A4"

    # ── Carry Forward sheet ───────────────────────────────────────────────────
    # Section A: Per-store carries (Commission and Shop only; DSP is per-group below)
    carry = wb.create_sheet("Carry Forward")
    carry.sheet_view.showGridLines = False

    def _carry_hdr(ws, hdrs, fills, row_h=30):
        ws.append(hdrs)
        row_num = ws.max_row
        for ci in range(1, len(hdrs) + 1):   # only iterate the cols we actually wrote
            cell = ws.cell(row_num, ci)
            cell.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.fill      = _fill(fills[ci - 1])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = bdr
        ws.row_dimensions[row_num].height = row_h

    # Section A header
    sec_a_hdrs  = ["Sub Affiliate ID", "Sub Affiliate Name",
                   "Shop\nCarry to Next", "DSP\nCarry to Next",
                   "Supervisor\nCarry to Next", "Manager\nCarry to Next",
                   "Total Negative\nCarry Forward"]
    sec_a_fills = ["1A3C6E", "1A3C6E", "7B5800", "2E5F1A", "37474F", "5D1E0A", "1A3C6E"]
    _carry_hdr(carry, sec_a_hdrs, sec_a_fills)

    carry_rows = [r for r in result.rows
                  if any(r.shares[k].next_carry < 0
                         for k in ("shop", "dsp", "supervisor", "manager"))]
    for row in carry_rows:
        row_neg = sum(v for v in [
            row.shares["shop"].next_carry,
            row.shares["dsp"].next_carry,
            row.shares["supervisor"].next_carry,
            row.shares["manager"].next_carry,
        ] if v < 0)
        carry.append([
            row.sub_affiliate, row.sub_affiliate_name,
            float(row.shares["shop"].next_carry),
            float(row.shares["dsp"].next_carry),
            float(row.shares["supervisor"].next_carry),
            float(row.shares["manager"].next_carry),
            float(row_neg),
        ])
    carry.append([
        "TOTAL", "",
        float(totals["shop_next_carry"]),
        float(totals["dsp_next_carry"]),
        float(totals["supervisor_next_carry"]),
        float(totals["manager_next_carry"]),
        float(sum(totals[f"{k}_next_carry"] for k in ("shop", "dsp", "supervisor", "manager")
                  if totals[f"{k}_next_carry"] < 0)),
    ])
    sec_a_cols = 7
    for rw in carry.iter_rows(min_row=2, min_col=3, max_col=sec_a_cols):
        for cell in rw:
            cell.number_format = MONEY_FMT; cell.border = bdr
    for rw in carry.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in rw: cell.border = bdr
    last_a = carry.max_row
    for ci in range(1, sec_a_cols + 1):
        c = carry.cell(last_a, ci)
        c.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
        c.fill = _fill("2C3E50")

    # Column widths
    carry.column_dimensions["A"].width = 20
    carry.column_dimensions["B"].width = 26
    for ci in range(3, sec_a_cols + 1):
        carry.column_dimensions[get_column_letter(ci)].width = 20
    carry.freeze_panes = "A2"

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
