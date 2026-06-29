"""pcf_excel_parser.py — Dedicated parser for PCF PER EXCEL.xlsx.

Handles the complex two-part structure of the PCF petty-cash Excel workbook:

  LEFT SIDE  (cols 0–3, rows 1–N):
      Individual expense rows — date, description, amount.
      PS & AF summary rows (e.g. "PS & AF April 26-30 2026 week 5") are
      SKIPPED because they are totals of the cross-table below.

  RIGHT SIDE (cross-table, rows 12+):
      Per-branch PS (Pagcor Share) and AF (Audit Fee) amounts for each of
      the four payment weeks.  Column layout detected from row 11 labels:
          APR-W5 → bc=40, ps=41, af=42
          MAY-W1 → bc=44, ps=45, af=46
          MAY-W2 → bc=49, ps=50, af=51
          MAY-W3 → bc=54, ps=55, af=56

Returns list[GenericTransaction] with:
    debit  = 0
    credit = abs(amount)      ← outflow from the petty-cash fund
    amount = abs(amount)
so that Pool-A matching (sys_debits ↔ bank_credits) works correctly when
paired with parse_pcf_system().
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.parsers.bank_statement_generic import GenericTransaction, GenericStatementParseError


class PCFExcelParseError(GenericStatementParseError):
    pass


# Row index (0-based) where week labels appear in the cross-table header
_WEEK_LABEL_ROW = 11

# Fixed column layout for the four payment weeks.
# bc = branch-code col, ps = PS-amount col, af = AF-amount col
_WEEK_COLS: list[dict] = [
    {"label": "APR-W5", "bc": 40, "ps": 41, "af": 42},
    {"label": "MAY-W1", "bc": 44, "ps": 45, "af": 46},
    {"label": "MAY-W2", "bc": 49, "ps": 50, "af": 51},
    {"label": "MAY-W3", "bc": 54, "ps": 55, "af": 56},
]

_PSAF_KEYWORDS = ("ps & af", "ps&af")
_SKIP_DESC_STARTS = ("total:", "total ", "total\t", "shortpayment for lw14nac landbased ps")


def _is_psaf_summary(desc: str) -> bool:
    n = desc.strip().lower()
    return any(n.startswith(kw) for kw in _PSAF_KEYWORDS)


def _is_skip_row(desc: str) -> bool:
    n = desc.strip().lower()
    return any(n.startswith(kw) for kw in _SKIP_DESC_STARTS)


def _to_date_str(value: Any) -> str:
    """Convert an openpyxl cell value to ISO date string, or '' on failure."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        s = str(value).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
    except Exception:
        pass
    return ""


def _to_amount(value: Any) -> Decimal | None:
    """Parse a numeric cell to a positive Decimal, or None if blank/zero."""
    if value is None or str(value).strip() == "":
        return None
    try:
        d = Decimal(str(value))
        return d if d > 0 else abs(d) if d < 0 else None
    except (InvalidOperation, Exception):
        return None


# ── Week-date detection ───────────────────────────────────────────────────────

def _detect_week_dates(rows: list[tuple]) -> dict[str, str]:
    """
    Scan the left-side description column (col 1) for PS & AF summary rows.
    Returns {week_label: iso_date_str}.
    """
    dates: dict[str, str] = {}
    for row in rows:
        desc = str(row[1] if len(row) > 1 else "").strip()
        if not _is_psaf_summary(desc):
            continue
        d = _to_date_str(row[0] if row else None)
        if not d:
            continue
        u = desc.upper()
        if ("WEEK 5" in u or "APR" in u) and "APR-W5" not in dates:
            dates["APR-W5"] = d
        elif "WEEK 1" in u and "APR-W5" not in u and "MAY-W1" not in dates:
            dates["MAY-W1"] = d
        elif "WEEK 2" in u and "MAY-W2" not in dates:
            dates["MAY-W2"] = d
        elif "WEEK 3" in u and "MAY-W3" not in dates:
            dates["MAY-W3"] = d
    return dates


# ── Main entry point ──────────────────────────────────────────────────────────

def parse_pcf_excel(path: str | Path) -> list[GenericTransaction]:
    """Parse PCF PER EXCEL.xlsx and return a list of GenericTransaction objects."""
    path = Path(path)
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        raise PCFExcelParseError(f"Cannot open {path.name}: {exc}") from exc

    ws = wb.active
    rows: list[tuple] = [row for row in ws.iter_rows(values_only=True)]
    wb.close()

    if not rows:
        raise PCFExcelParseError("PCF Excel file is empty.")

    week_dates = _detect_week_dates(rows)
    out: list[GenericTransaction] = []
    row_no = 0

    # ── Part 1: Individual expense rows (left side) ───────────────────────────
    for i, row in enumerate(rows):
        # Row 0 is the column-header row ("Petty Cash Beg. Balance" etc.)
        if i == 0:
            continue
        # Skip fully empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        raw_date = row[0] if len(row) > 0 else None
        raw_desc = row[1] if len(row) > 1 else None
        raw_amt  = row[3] if len(row) > 3 else None

        desc = str(raw_desc or "").strip()
        if not desc:
            continue

        # Skip PS & AF totals (they are broken out into per-branch entries below)
        if _is_psaf_summary(desc):
            continue

        # Skip total / shortpayment header rows
        if _is_skip_row(desc):
            continue

        amount = _to_amount(raw_amt)
        if amount is None:
            continue

        date_str = _to_date_str(raw_date)
        row_no += 1
        out.append(GenericTransaction(
            row_number=row_no,
            date=date_str,
            description=desc,
            reference=None,
            debit=Decimal("0"),
            credit=amount,
            amount=amount,
            balance=None,
            bank_profile="PCF_EXCEL",
        ))

    # ── Part 2: Branch PS & AF cross-table ───────────────────────────────────
    for week in _WEEK_COLS:
        label  = week["label"]
        bc_col = week["bc"]
        ps_col = week["ps"]
        af_col = week["af"]
        w_date = week_dates.get(label, "")

        for i, row in enumerate(rows):
            if i <= _WEEK_LABEL_ROW:
                continue  # skip header + data above the cross-table
            if len(row) <= af_col:
                continue

            bc_val = row[bc_col]
            if bc_val is None or str(bc_val).strip() in ("", "TOTAL:", "TOTAL"):
                continue
            branch = str(bc_val).strip()

            # PS entry
            ps_amt = _to_amount(row[ps_col])
            if ps_amt:
                row_no += 1
                out.append(GenericTransaction(
                    row_number=row_no,
                    date=w_date,
                    description=f"{branch} PS {label}",
                    reference=label,
                    debit=Decimal("0"),
                    credit=ps_amt,
                    amount=ps_amt,
                    balance=None,
                    bank_profile="PCF_EXCEL",
                ))

            # AF entry
            af_amt = _to_amount(row[af_col])
            if af_amt:
                row_no += 1
                out.append(GenericTransaction(
                    row_number=row_no,
                    date=w_date,
                    description=f"{branch} AF {label}",
                    reference=label,
                    debit=Decimal("0"),
                    credit=af_amt,
                    amount=af_amt,
                    balance=None,
                    bank_profile="PCF_EXCEL",
                ))

    if not out:
        raise PCFExcelParseError("No transaction rows found in PCF Excel file.")

    return out
