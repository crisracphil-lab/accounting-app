"""GGR Weekly Summary parser and exporter.

Parses the RAC Phil Corp monthly GGR report (25-column format).
Exports the original file content with a per-outlet monthly SUMMARY appended.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
import re
from typing import Any, Dict, List, Optional

class GGRSummaryParseError(Exception):
    pass

_WEEK_RE = re.compile(
    r"WEEK\s*(\d+)(?:\s*\(\s*(\d{1,2})/(\d{1,2})/(\d{2,4})"
    r"\s*(?:to|-)\s*(\d{1,2})/(\d{1,2})/(\d{2,4})\s*\))?",
    re.I,
)

NCOLS     = 25   # total columns in the RAC report
NUM_START = 3    # first numeric column (col D), 0-based index
NUM_END   = 23   # last numeric column  (col X), 0-based index, inclusive
NUM_COUNT = NUM_END - NUM_START + 1   # = 21


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class GGRWeekSubtotal:
    """One weekly SUB TOTAL row in raw 25-column form (used for web display)."""
    week_label:  str
    week_number: int
    week_start:  Optional[date]
    week_end:    Optional[date]
    values: List[Any] = field(default_factory=list)

    @property
    def ggr_total(self)    -> Decimal: return _dec(self.values[19] if len(self.values) > 19 else 0)
    @property
    def pagcor_total(self) -> Decimal: return _dec(self.values[20] if len(self.values) > 20 else 0)
    @property
    def audit_total(self)  -> Decimal: return _dec(self.values[21] if len(self.values) > 21 else 0)
    @property
    def operator(self)     -> Decimal: return _dec(self.values[22] if len(self.values) > 22 else 0)

    @property
    def date_range(self) -> str:
        if self.week_start and self.week_end:
            return f"{self.week_start.isoformat()} to {self.week_end.isoformat()}"
        return ""


@dataclass
class GGROutletTotal:
    """Per-outlet monthly total — sum of all weekly data rows for this outlet."""
    operator_name: str
    outlet_name:   str
    outlet_code:   str
    sums: List[float] = field(default_factory=list)  # NUM_COUNT floats (cols D-X)


@dataclass
class GGRSummary:
    period_label:    str
    source_filename: str
    sheet_name:      str
    week_subtotals:  List[GGRWeekSubtotal] = field(default_factory=list)
    outlet_totals:   List[GGROutletTotal]  = field(default_factory=list)
    raw_rows:        List[List[Any]]       = field(default_factory=list)
    header_row_idx:  int = 5

    @property
    def weeks(self):
        return self.week_subtotals

    @property
    def total(self):
        class _T:
            pass
        t = _T()
        t.ggr          = sum(w.ggr_total    for w in self.week_subtotals)
        t.pagcor_share = sum(w.pagcor_total  for w in self.week_subtotals)
        t.audit_fee    = sum(w.audit_total   for w in self.week_subtotals)
        t.operator     = sum(w.operator      for w in self.week_subtotals)
        t.outlet_count = len(self.week_subtotals)
        return t

    @property
    def venues(self):
        return {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()

def _norm(v: Any) -> str:
    text = _s(v).lower()
    text = re.sub(r"[\n\r\t_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9/ ]+", " ", text)
    return " ".join(text.split())

def _dec(v: Any) -> Decimal:
    if v is None or _s(v) == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(round(v, 10)))
    t = _s(v).replace(",", "").replace("₱", "")
    if t.startswith("(") and t.endswith(")"):
        t = "-" + t[1:-1]
    t = re.sub(r"[^0-9.\-]", "", t)
    if t in {"", "-", "."}:
        return Decimal("0")
    try:
        return Decimal(t)
    except Exception:
        return Decimal("0")

def _flt(v: Any) -> float:
    try:
        return float(_dec(v))
    except Exception:
        return 0.0

def _parse_week_header(text: str):
    m = _WEEK_RE.search(text or "")
    if not m:
        return None
    no = int(m.group(1))
    start = end = None
    if m.group(2):
        y1 = int(m.group(4))
        y2 = int(m.group(7))
        y1 += 2000 if y1 < 100 else 0
        y2 += 2000 if y2 < 100 else 0
        try:
            start = date(y1, int(m.group(2)), int(m.group(3)))
            end   = date(y2, int(m.group(5)), int(m.group(6)))
        except Exception:
            pass
    return no, _s(text).strip(), start, end


# ---------------------------------------------------------------------------
# File reader
# ---------------------------------------------------------------------------

def _read_summary_sheet(path: Path):
    suf = path.suffix.lower()

    if suf == ".xls":
        try:
            import xlrd
        except ImportError as exc:
            raise GGRSummaryParseError(
                "xlrd is required for .xls files. Run: pip install xlrd"
            ) from exc
        wb = xlrd.open_workbook(str(path))
        sh = next((s for s in wb.sheets() if "summary" in s.name.lower()), None)
        if sh is None:
            sh = max(wb.sheets(), key=lambda s: s.nrows * max(s.ncols, 1))
        rows = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return sh.name, rows

    if suf in {".xlsx", ".xlsm"}:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            ws = next((s for s in wb.worksheets if "summary" in s.title.lower()), None)
            if ws is None:
                ws = max(wb.worksheets,
                         key=lambda s: (s.max_row or 0) * max(s.max_column or 1, 1))
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            return ws.title, rows
        finally:
            wb.close()

    raise GGRSummaryParseError(
        f"Unsupported file type: {suf}. Upload a .xlsx, .xlsm, or .xls file."
    )


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_ggr_weekly_summary(file_path) -> GGRSummary:
    path = Path(file_path)
    if not path.exists():
        raise GGRSummaryParseError(f"File not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xls"}:
        raise GGRSummaryParseError(
            "Please upload the GGR workbook as .xlsx, .xlsm, or .xls."
        )

    sheet_name, rows = _read_summary_sheet(path)

    # Detect period label
    period = ""
    for row in rows[:20]:
        vals = " ".join(_s(c) for c in row[:10])
        if "Period Covered" in vals:
            period = vals.split(":", 1)[-1].strip()
            break

    # Detect header row index
    header_row_idx = 5
    for i, row in enumerate(rows[:20]):
        joined = " ".join(_norm(c) for c in row)
        if "gross gaming revenue" in joined and "pagcor" in joined:
            header_row_idx = i
            break

    # Walk rows
    current_week_label = ""
    current_week_no    = 0
    current_week_start: Optional[date] = None
    current_week_end:   Optional[date] = None
    week_subtotals: List[GGRWeekSubtotal] = []
    outlet_map: Dict[str, GGROutletTotal] = {}  # preserves insertion order

    for row_idx, row in enumerate(rows):
        # Pad to NCOLS
        while len(row) < NCOLS:
            row.append(None)

        col0 = _s(row[0])
        col1 = _s(row[1])
        col2 = _s(row[2])

        # Skip header rows
        if row_idx <= header_row_idx:
            continue

        col1_up = col1.upper()

        # WEEK header
        pw = _parse_week_header(col1)
        if pw:
            current_week_no, current_week_label, current_week_start, current_week_end = pw
            continue

        # SUB TOTAL row
        if col1_up.startswith("SUB TOTAL"):
            if current_week_no > 0:
                week_subtotals.append(GGRWeekSubtotal(
                    week_label  = current_week_label,
                    week_number = current_week_no,
                    week_start  = current_week_start,
                    week_end    = current_week_end,
                    values      = list(row),
                ))
            current_week_no    = 0
            current_week_label = ""
            current_week_start = None
            current_week_end   = None
            continue

        # Skip GRAND TOTAL, SUMMARY, blank rows
        if col1_up in {"GRAND TOTAL", "SUMMARY"} or col1_up.startswith("GRAND TOTAL"):
            continue
        if not col0 and not col1 and not col2:
            continue

        # Data row: accumulate per-outlet monthly totals
        key = col2.strip() if col2.strip() else col1.strip()
        if not key:
            continue

        if key not in outlet_map:
            outlet_map[key] = GGROutletTotal(
                operator_name = col0,
                outlet_name   = col1,
                outlet_code   = col2,
                sums          = [0.0] * NUM_COUNT,
            )
        ot = outlet_map[key]
        for c in range(NUM_START, NUM_END + 1):
            ot.sums[c - NUM_START] += _flt(row[c])

    if not week_subtotals:
        raise GGRSummaryParseError(
            "No weekly SUB TOTAL rows found. "
            "Check that the file has 'WEEK N' headers and 'SUB TOTAL' rows in the Summary sheet."
        )

    return GGRSummary(
        period_label    = period,
        source_filename = path.name,
        sheet_name      = sheet_name,
        week_subtotals  = week_subtotals,
        outlet_totals   = list(outlet_map.values()),
        raw_rows        = rows,
        header_row_idx  = header_row_idx,
    )


# ---------------------------------------------------------------------------
# Exporter
# ---------------------------------------------------------------------------

def export_ggr_weekly_summary(summary: GGRSummary) -> bytes:
    """Reproduce the original sheet and append the per-outlet monthly SUMMARY section."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = summary.sheet_name or "Summary"

    # ── Styles ────────────────────────────────────────────────────────────────
    thin = Side(style="thin", color="BFBFBF")
    THIN = Border(left=thin, right=thin, top=thin, bottom=thin)

    NORM  = Font(name="Calibri", size=9)
    BOLD  = Font(name="Calibri", bold=True, size=9)
    WHBLD = Font(name="Calibri", bold=True, size=9, color="FFFFFF")

    FILL_SUMHDR = PatternFill("solid", fgColor="375623")  # dark green banner
    FILL_OUTLET = PatternFill("solid", fgColor="E2EFDA")  # light green outlet rows
    FILL_GRAND  = PatternFill("solid", fgColor="A9D18E")  # medium green grand total

    NUM_FMT = '#,##0.00;[Red](#,##0.00);"-"'

    def num_cell(cell, val, fill=None, bold=False):
        try:
            cell.value = float(val) if val is not None else None
        except Exception:
            cell.value = val
        cell.font          = BOLD if bold else NORM
        cell.number_format = NUM_FMT
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.border        = THIN
        if fill:
            cell.fill = fill

    def txt_cell(cell, val, fill=None, bold=False):
        cell.value     = val
        cell.font      = BOLD if bold else NORM
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border    = THIN
        if fill:
            cell.fill = fill

    # ── Write original rows ───────────────────────────────────────────────────
    for r_idx, row in enumerate(summary.raw_rows, start=1):
        for c_idx, val in enumerate(row[:NCOLS], start=1):
            c = ws.cell(r_idx, c_idx, val)
            c.font = NORM

    # Column widths
    widths = [22, 38, 10, 14, 14, 14, 16, 16, 16, 14, 12,
              14, 14, 14, 16, 16, 16, 14, 12, 16, 14, 12, 14, 18, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D7"

    # ── SUMMARY banner ────────────────────────────────────────────────────────
    # Appended immediately after last original row (no blank gap — matches expected output)
    r = len(summary.raw_rows) + 1

    for c_idx in range(1, NCOLS + 1):
        cell = ws.cell(r, c_idx)
        cell.fill   = FILL_SUMHDR
        cell.border = THIN
    ws.cell(r, 2).value     = "SUMMARY"
    ws.cell(r, 2).font      = WHBLD
    ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 16
    r += 1

    # ── Per-outlet rows ───────────────────────────────────────────────────────
    for ot in summary.outlet_totals:
        txt_cell(ws.cell(r, 1), "",            fill=FILL_OUTLET)
        txt_cell(ws.cell(r, 2), ot.outlet_name, fill=FILL_OUTLET)
        txt_cell(ws.cell(r, 3), ot.outlet_code, fill=FILL_OUTLET)
        for c_idx in range(NUM_START, NUM_END + 1):
            val = ot.sums[c_idx - NUM_START]
            num_cell(ws.cell(r, c_idx + 1), val, fill=FILL_OUTLET)
        # col Y (index 24, 1-based col 25) — blank
        ws.cell(r, NCOLS).fill   = FILL_OUTLET
        ws.cell(r, NCOLS).border = THIN
        ws.row_dimensions[r].height = 15
        r += 1

    # ── Grand total row (no label — matches expected output R169) ─────────────
    txt_cell(ws.cell(r, 1), "", fill=FILL_GRAND, bold=True)
    txt_cell(ws.cell(r, 2), "", fill=FILL_GRAND, bold=True)
    txt_cell(ws.cell(r, 3), "", fill=FILL_GRAND, bold=True)
    for c_idx in range(NUM_START, NUM_END + 1):
        col_sum = sum(ot.sums[c_idx - NUM_START] for ot in summary.outlet_totals)
        num_cell(ws.cell(r, c_idx + 1), col_sum, fill=FILL_GRAND, bold=True)
    ws.cell(r, NCOLS).fill   = FILL_GRAND
    ws.cell(r, NCOLS).border = THIN
    ws.row_dimensions[r].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
