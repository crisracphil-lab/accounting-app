"""
Subsidiary Ledger parser for RAC Phil Corp / PGI format.

Accepts:
  (a) .xls  — the PER SYSTEM export (sheet 1, col 1 = 'A/C No.')
  (b) .xlsx/.xlsm — same column layout

Column layout (1-indexed):
  1  A/C No.          — account code (numeric on header rows)
  2  Acct Title
  3  A/C Alias
  4  Voucher Date
  5  Voucher No.      — e.g. TR10-26040000007-0010
  6  Department No.
  7  Department Name
  8  Description
  9  Remark
  10 Debit (Local Curr.)
  11 Credit (Local Curr.)
  12 Balance

Journal-entry reconstruction:
  Voucher numbers have the form  BASE-NNNN  where NNNN is a 4-digit line suffix.
  All postings sharing the same BASE across ALL accounts form one journal entry.
  The bank/cash account appears on the CREDIT side for outgoing payments;
  the expense/asset account appears on the DEBIT side.
  Use build_voucher_map() to get the full entry, then find_expense_accounts()
  to extract non-cash debit legs — those are the account titles to classify.
"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple


class SubsidiaryLedgerParseError(Exception):
    pass


@dataclass
class LedgerPosting:
    voucher_date: str
    voucher_no: str
    description: str
    remark: str
    debit: Decimal
    credit: Decimal
    balance: Decimal

    @property
    def text(self) -> str:
        return " ".join(part for part in (self.description, self.remark) if part).strip()

    @property
    def amount(self) -> Decimal:
        return self.debit if self.debit != 0 else self.credit

    @property
    def voucher_base(self) -> str:
        """Strip the 4-digit line suffix (e.g. '-0010') to get the journal entry key."""
        # Pattern: ends with -NNNN where NNNN is exactly 4 digits
        m = re.match(r'^(.+)-(\d{4})$', self.voucher_no.strip())
        return m.group(1) if m else self.voucher_no.strip()


@dataclass
class AccountActivity:
    code: str
    title: str
    opening_balance: Decimal
    period_debit: Decimal
    period_credit: Decimal
    closing_balance: Decimal
    rows: int
    postings: List[LedgerPosting] = field(default_factory=list)

    @property
    def net_change(self) -> Decimal:
        """closing - opening (signed; positive = increase in net position)."""
        return self.closing_balance - self.opening_balance


@dataclass
class SubsidiaryLedger:
    period_label: Optional[str] = None
    accounts: List[AccountActivity] = field(default_factory=list)
    rows_total: int = 0


# ── Cash / bank account code prefixes — never used as the "expense" leg ──────
_CASH_PREFIXES = (
    "1101", "1102", "1103", "1104", "1105", "1106", "1107",
    "1108", "1109", "1110", "1111", "1112", "1113", "1114", "1115",
)

def _is_cash_or_bank(code: str) -> bool:
    return any(code.startswith(p) for p in _CASH_PREFIXES)


@dataclass
class VoucherEntry:
    """A complete journal entry reconstructed from all accounts sharing a voucher base."""
    voucher_base: str
    description: str                    # from the first posting found
    debit_legs: List[Tuple[str, str, Decimal]]   # [(code, title, amount)]
    credit_legs: List[Tuple[str, str, Decimal]]  # [(code, title, amount)]

    @property
    def expense_accounts(self) -> List[Tuple[str, str, Decimal]]:
        """Non-cash/bank debit accounts — the actual expense/asset being recorded."""
        return [(c, t, a) for c, t, a in self.debit_legs if not _is_cash_or_bank(c)]

    @property
    def total_amount(self) -> Decimal:
        return sum(a for _, _, a in self.debit_legs)


def build_voucher_map(ledger: "SubsidiaryLedger") -> Dict[str, VoucherEntry]:
    """
    Group all postings across every account by voucher base.
    Returns a dict  voucher_base → VoucherEntry  containing every DR and CR leg.
    This lets you look up any voucher and instantly see which expense accounts
    were debited when the bank/cash was credited (i.e. a payment was made).
    """
    raw: Dict[str, List[Tuple[str, str, "LedgerPosting"]]] = defaultdict(list)
    for acct in ledger.accounts:
        for posting in acct.postings:
            if posting.voucher_no:
                raw[posting.voucher_base].append((acct.code, acct.title, posting))

    result: Dict[str, VoucherEntry] = {}
    for vbase, entries in raw.items():
        dr_legs: List[Tuple[str, str, Decimal]] = []
        cr_legs: List[Tuple[str, str, Decimal]] = []
        first_desc = ""
        seen_dr: Dict[Tuple[str,str], Decimal] = defaultdict(Decimal)
        seen_cr: Dict[Tuple[str,str], Decimal] = defaultdict(Decimal)
        for code, title, posting in entries:
            if not first_desc and posting.description:
                first_desc = posting.description
            if posting.debit > 0:
                seen_dr[(code, title)] += posting.debit
            if posting.credit > 0:
                seen_cr[(code, title)] += posting.credit
        dr_legs = [(c, t, a) for (c, t), a in seen_dr.items()]
        cr_legs = [(c, t, a) for (c, t), a in seen_cr.items()]
        result[vbase] = VoucherEntry(
            voucher_base=vbase,
            description=first_desc,
            debit_legs=dr_legs,
            credit_legs=cr_legs,
        )
    return result


# Accounts that appear on almost every payment as a trivial fee —
# never the meaningful expense being classified.
_TRIVIAL_ACCOUNTS = {
    "6603",   # Bank Charges (flat PHP 25 per transaction)
    "1218002", # Input VAT Services (always paired with the actual expense)
    "1218001", # Input VAT Goods
    "2311",   # Expanded Withholding Tax Payable
    "2312",   # Final Withholding Tax
}


def find_expense_account_for_description(
    description: str,
    voucher_map: Dict[str, VoucherEntry],
    top_n: int = 1,
) -> Optional[Tuple[str, str, float]]:
    """
    Given a free-text description (e.g. from a payment request or receipt),
    find the most likely non-cash, non-trivial expense account.

    Strategy:
    1. Tokenise both the query and each voucher description.
    2. Score each matching voucher by keyword Jaccard overlap.
    3. For each matched voucher, award that score to the PRIMARY expense account
       (largest-amount non-cash, non-trivial DR leg) — not equally to all legs.
    4. Return (account_code, account_title, confidence_0_to_1).

    Returns None if no match found.
    """
    if not description:
        return None

    # Stopwords that appear everywhere and add no signal
    _STOP = {
        'FOR', 'THE', 'AND', 'APR', 'MAR', 'FEB', 'JAN', 'MAY', 'JUN',
        'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC', '2025', '2026', '2024',
        'LKW', 'BKL', 'RAC', 'PGI', 'DFO', 'BCK', 'BCL',
    }

    def tokens(s: str) -> set:
        raw = {w.upper() for w in re.findall(r'[A-Za-z]{3,}', s)}
        return raw - _STOP

    desc_tokens = tokens(description)
    if not desc_tokens:
        return None

    acct_scores: Dict[Tuple[str, str], float] = defaultdict(float)

    for entry in voucher_map.values():
        # Only consider non-trivial expense accounts
        meaningful = [
            (c, t, a) for c, t, a in entry.expense_accounts
            if c not in _TRIVIAL_ACCOUNTS
        ]
        if not meaningful:
            continue

        entry_tokens = tokens(entry.description)
        if not entry_tokens:
            continue

        overlap = len(desc_tokens & entry_tokens)
        if overlap == 0:
            continue

        jaccard = overlap / len(desc_tokens | entry_tokens)

        # Award the full score only to the PRIMARY (largest amount) expense account
        primary = max(meaningful, key=lambda x: x[2])
        code, title, amount = primary
        # Weight by log of amount so large payments dominate more
        import math
        weight = math.log1p(float(amount))
        acct_scores[(code, title)] += jaccard * weight

    if not acct_scores:
        return None

    # Sort by score descending
    ranked = sorted(acct_scores.items(), key=lambda x: -x[1])
    (code, title), best_score = ranked[0]

    # Normalise: cap confidence at 0.95 to reflect uncertainty
    max_possible = max(v for _, v in ranked)
    confidence = min(best_score / max_possible * 0.95, 0.95) if max_possible > 0 else 0.0
    return (code, title, confidence)


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal("0")
    s = str(v).strip().replace(",", "")
    if s == "":
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _looks_like_subsidiary_ledger_headers(headers: List[str]) -> bool:
    joined = " ".join(h.lower() for h in headers)
    has_account_header = "a/c no" in joined
    has_ledger_columns = "debit" in joined and "credit" in joined and "balance" in joined
    # Some internally generated tests/minimal exports leave the later debit/credit
    # headers blank but still use the subsidiary-ledger column layout through col L.
    return has_account_header and (has_ledger_columns or len(headers) >= 12)


def parse_subsidiary_ledger(file_path) -> SubsidiaryLedger:
    path = Path(file_path)
    if not path.exists():
        raise SubsidiaryLedgerParseError(f"File not found: {path}")
    suf = path.suffix.lower()
    if suf == ".xls":
        return _parse_xls(path)
    if suf in (".xlsx", ".xlsm"):
        return _parse_xlsx(path)
    raise SubsidiaryLedgerParseError(f"Expected .xls or .xlsx, got {suf}")


def _parse_xls(path) -> SubsidiaryLedger:
    try:
        import xlrd
    except ImportError as e:
        raise SubsidiaryLedgerParseError("xlrd required for .xls files") from e
    wb = xlrd.open_workbook(path)
    # Use the first sheet that has 'A/C No.' header. Skip 'PER GGR'.
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        for r in range(min(15, ws.nrows)):
            headers = [_to_str(ws.cell(r, c).value) for c in range(ws.ncols)]
            if _looks_like_subsidiary_ledger_headers(headers):
                return _parse_ledger_table_xls(ws, header_row=r)
    raise SubsidiaryLedgerParseError(
        "Could not find 'A/C No.' header in any sheet")


def _parse_ledger_table_xls(ws, header_row: int) -> SubsidiaryLedger:
    led = SubsidiaryLedger()
    current: Optional[AccountActivity] = None
    for r in range(header_row + 1, ws.nrows):
        ac_raw = ws.cell(r, 0).value
        title_raw = ws.cell(r, 1).value
        # Account header row: numeric A/C No.
        if isinstance(ac_raw, (int, float)) and ac_raw > 0:
            if current:
                led.accounts.append(current)
            opening = _to_decimal(ws.cell(r, 11).value)  # col 12 = balance
            current = AccountActivity(
                code=f"{int(ac_raw)}",
                title=_to_str(title_raw),
                opening_balance=opening,
                period_debit=Decimal("0"),
                period_credit=Decimal("0"),
                closing_balance=opening,
                rows=0,
            )
            continue
        if current is None:
            continue
        debit = _to_decimal(ws.cell(r, 9).value)
        credit = _to_decimal(ws.cell(r, 10).value)
        bal = _to_decimal(ws.cell(r, 11).value)
        if debit == 0 and credit == 0 and bal == 0:
            continue
        current.period_debit += debit
        current.period_credit += credit
        if bal != 0:
            current.closing_balance = bal
        current.postings.append(LedgerPosting(
            voucher_date=_to_str(ws.cell(r, 3).value),
            voucher_no=_to_str(ws.cell(r, 4).value),
            description=_to_str(ws.cell(r, 7).value),
            remark=_to_str(ws.cell(r, 8).value),
            debit=debit,
            credit=credit,
            balance=bal,
        ))
        current.rows += 1
        led.rows_total += 1
    if current:
        led.accounts.append(current)
    return led


def _parse_xlsx(path) -> SubsidiaryLedger:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(15, ws.max_row + 1)):
            headers = [_to_str(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
            if _looks_like_subsidiary_ledger_headers(headers):
                return _parse_ledger_table_xlsx(ws, header_row=r)
    raise SubsidiaryLedgerParseError(
        "Could not find 'A/C No.' header in any sheet")


def _parse_ledger_table_xlsx(ws, header_row: int) -> SubsidiaryLedger:
    led = SubsidiaryLedger()
    current: Optional[AccountActivity] = None
    for r in range(header_row + 1, ws.max_row + 1):
        ac_raw = ws.cell(row=r, column=1).value
        title_raw = ws.cell(row=r, column=2).value
        if isinstance(ac_raw, (int, float)) and ac_raw > 0:
            if current:
                led.accounts.append(current)
            opening = _to_decimal(ws.cell(row=r, column=12).value)
            current = AccountActivity(
                code=f"{int(ac_raw)}",
                title=_to_str(title_raw),
                opening_balance=opening,
                period_debit=Decimal("0"),
                period_credit=Decimal("0"),
                closing_balance=opening,
                rows=0,
            )
            continue
        if current is None:
            continue
        debit = _to_decimal(ws.cell(row=r, column=10).value)
        credit = _to_decimal(ws.cell(row=r, column=11).value)
        bal = _to_decimal(ws.cell(row=r, column=12).value)
        if debit == 0 and credit == 0 and bal == 0:
            continue
        current.period_debit += debit
        current.period_credit += credit
        if bal != 0:
            current.closing_balance = bal
        current.postings.append(LedgerPosting(
            voucher_date=_to_str(ws.cell(row=r, column=4).value),
            voucher_no=_to_str(ws.cell(row=r, column=5).value),
            description=_to_str(ws.cell(row=r, column=8).value),
            remark=_to_str(ws.cell(row=r, column=9).value),
            debit=debit,
            credit=credit,
            balance=bal,
        ))
        current.rows += 1
        led.rows_total += 1
    if current:
        led.accounts.append(current)
    return led
