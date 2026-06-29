"""
Financial Statement parser (RAC Phil format).

Reads a .xls workbook with these sheets:
  IS3  - Income Statement
  BS   - Balance Sheet
  ...  - other sheets ignored

Each FS sheet has:
  Row 1-4: header / metadata
  Row 5:   column header (A/C No. | Acct Title | <period cols> | Total | % | Inc./Dec. | Reasons/Remarks)
  Row 6+:  account rows (numeric A/C No.) + section headers (no A/C No.) + blank spacers

The parser captures every meaningful row with its values and any
existing Reasons/Remarks text the user has filled in.
"""
from __future__ import annotations

import json
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import List, Optional


class FinancialStatementParseError(Exception):
    pass


@dataclass
class FSRow:
    sheet: str                  # 'IS' | 'BS'
    row_number: int             # original spreadsheet row (1-based)
    account_code: Optional[str]
    account_title: str
    columns: dict               # {column_label: value} for amount columns
    inc_dec: Optional[Decimal]
    remarks: Optional[str]

    def as_dict(self) -> dict:
        return {
            "sheet": self.sheet, "row_number": self.row_number,
            "account_code": self.account_code, "account_title": self.account_title,
            "columns": {k: str(v) if isinstance(v, Decimal) else v
                        for k, v in self.columns.items()},
            "inc_dec": str(self.inc_dec) if self.inc_dec is not None else None,
            "remarks": self.remarks,
        }


@dataclass
class FSWorkbook:
    period_label: Optional[str]
    company_name: Optional[str]
    is_columns: List[str] = field(default_factory=list)
    bs_columns: List[str] = field(default_factory=list)
    is_rows: List[FSRow] = field(default_factory=list)
    bs_rows: List[FSRow] = field(default_factory=list)


def _to_decimal(v) -> Optional[Decimal]:
    if v is None:
        return None
    s = str(v).strip().replace(",", "")
    if s == "":
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _to_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def parse_fs(file_path) -> FSWorkbook:
    path = Path(file_path)
    if not path.exists():
        raise FinancialStatementParseError(f"File not found: {path}")
    suf = path.suffix.lower()
    if suf == ".xls":
        return _parse_xls(path)
    if suf in (".xlsx", ".xlsm"):
        return _parse_xlsx(path)
    raise FinancialStatementParseError(f"Expected .xls or .xlsx, got {suf}")


def _parse_xls(path) -> FSWorkbook:
    try:
        import xlrd
    except ImportError as e:
        raise FinancialStatementParseError("xlrd required for .xls files") from e
    wb = xlrd.open_workbook(path)

    fs = FSWorkbook(period_label=None, company_name=None)

    # IS sheet (try IS3 first, then any sheet starting with IS)
    is_sheet = next((s for s in wb.sheet_names() if s.upper() == "IS3"), None)
    if is_sheet is None:
        is_sheet = next((s for s in wb.sheet_names() if s.upper().startswith("IS")), None)
    if is_sheet:
        fs.is_columns, fs.is_rows, fs.company_name = _parse_xls_sheet(
            wb.sheet_by_name(is_sheet), "IS")

    # BS sheet
    bs_sheet = next((s for s in wb.sheet_names() if s.upper() == "BS"), None)
    if bs_sheet:
        bs_cols, bs_rows, bs_company = _parse_xls_sheet(wb.sheet_by_name(bs_sheet), "BS")
        fs.bs_columns = bs_cols
        fs.bs_rows = bs_rows
        if not fs.company_name:
            fs.company_name = bs_company

    if not fs.is_rows and not fs.bs_rows:
        raise FinancialStatementParseError(
            f"Neither IS nor BS sheet found, or both empty. Sheets: {wb.sheet_names()}")
    return fs


def _parse_xls_sheet(ws, sheet_label: str):
    # Find header row containing 'A/C No.'
    header_row = None
    for r in range(min(15, ws.nrows)):
        v = ws.cell(r, 0).value
        if isinstance(v, str) and "A/C No" in v:
            header_row = r
            break
    if header_row is None:
        return [], [], None

    headers = [_to_str(ws.cell(header_row, c).value) for c in range(ws.ncols)]
    # Find Reasons/Remarks column index (rightmost match)
    remarks_idx = None
    for i, h in enumerate(headers):
        if "Reason" in h or "Remarks" in h:
            remarks_idx = i
    inc_dec_idx = None
    for i, h in enumerate(headers):
        if "Inc" in h and "Dec" in h:
            inc_dec_idx = i

    # Try to find company name in rows 1-2 column A or C
    company = None
    for r in range(min(3, ws.nrows)):
        for c in range(min(4, ws.ncols)):
            v = _to_str(ws.cell(r, c).value)
            if v and ("Corp" in v or "Inc" in v or "RAC" in v):
                company = v
                break
        if company:
            break

    rows: List[FSRow] = []
    for r in range(header_row + 1, ws.nrows):
        ac_raw = ws.cell(r, 0).value
        title = _to_str(ws.cell(r, 1).value)
        if not title and (ac_raw is None or _to_str(ac_raw) == ""):
            continue  # blank
        account_code = None
        if isinstance(ac_raw, (int, float)) and ac_raw > 0:
            # Integer-ish formatting
            if float(ac_raw).is_integer():
                account_code = str(int(ac_raw))
            else:
                account_code = str(ac_raw)
        elif _to_str(ac_raw) and _to_str(ac_raw) != "":
            account_code = _to_str(ac_raw) or None

        # Capture amount columns (everything between col 2 and remarks/inc_dec)
        cols = {}
        end_amount_col = remarks_idx if remarks_idx is not None else len(headers)
        for c in range(2, end_amount_col):
            label = headers[c] or f"col{c+1}"
            val = ws.cell(r, c).value
            d = _to_decimal(val)
            if d is not None:
                cols[label] = d
            else:
                s = _to_str(val)
                if s and s.strip():
                    cols[label] = s
        inc_dec = None
        if inc_dec_idx is not None:
            inc_dec = _to_decimal(ws.cell(r, inc_dec_idx).value)
        remarks = None
        if remarks_idx is not None:
            remarks = _to_str(ws.cell(r, remarks_idx).value) or None

        rows.append(FSRow(sheet=sheet_label, row_number=r + 1,
                          account_code=account_code, account_title=title,
                          columns=cols, inc_dec=inc_dec, remarks=remarks))
    return headers, rows, company


def _parse_xlsx(path) -> FSWorkbook:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    fs = FSWorkbook(period_label=None, company_name=None)
    is_sheet = next((s for s in wb.sheetnames if s.upper() == "IS3"), None)
    if is_sheet is None:
        is_sheet = next((s for s in wb.sheetnames if s.upper().startswith("IS")), None)
    if is_sheet:
        fs.is_columns, fs.is_rows, fs.company_name = _parse_xlsx_sheet(wb[is_sheet], "IS")
    bs_sheet = next((s for s in wb.sheetnames if s.upper() == "BS"), None)
    if bs_sheet:
        c, r, comp = _parse_xlsx_sheet(wb[bs_sheet], "BS")
        fs.bs_columns = c
        fs.bs_rows = r
        if not fs.company_name:
            fs.company_name = comp
    if not fs.is_rows and not fs.bs_rows:
        raise FinancialStatementParseError(
            f"Neither IS nor BS sheet found. Got: {wb.sheetnames}")
    return fs


def _parse_xlsx_sheet(ws, sheet_label: str):
    header_row = None
    for r in range(1, min(16, ws.max_row + 1)):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "A/C No" in v:
            header_row = r
            break
    if header_row is None:
        return [], [], None
    headers = [_to_str(ws.cell(row=header_row, column=c).value)
               for c in range(1, ws.max_column + 1)]
    remarks_idx = None
    for i, h in enumerate(headers):
        if "Reason" in h or "Remarks" in h:
            remarks_idx = i
    inc_dec_idx = None
    for i, h in enumerate(headers):
        if "Inc" in h and "Dec" in h:
            inc_dec_idx = i
    company = None
    for r in range(1, min(4, ws.max_row + 1)):
        for c in range(1, min(5, ws.max_column + 1)):
            v = _to_str(ws.cell(row=r, column=c).value)
            if v and ("Corp" in v or "Inc" in v or "RAC" in v):
                company = v
                break
        if company:
            break
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        ac_raw = ws.cell(row=r, column=1).value
        title = _to_str(ws.cell(row=r, column=2).value)
        if not title and (ac_raw is None or _to_str(ac_raw) == ""):
            continue
        account_code = None
        if isinstance(ac_raw, (int, float)) and ac_raw > 0:
            account_code = str(int(ac_raw)) if float(ac_raw).is_integer() else str(ac_raw)
        elif _to_str(ac_raw):
            account_code = _to_str(ac_raw)
        cols = {}
        end_col = (remarks_idx if remarks_idx is not None else len(headers))
        for c in range(2, end_col):  # 0-based -> col index 2 means column C
            label = headers[c] or f"col{c+1}"
            val = ws.cell(row=r, column=c+1).value  # convert 0-based to 1-based
            d = _to_decimal(val)
            if d is not None:
                cols[label] = d
            else:
                s = _to_str(val)
                if s and s.strip():
                    cols[label] = s
        inc_dec = (_to_decimal(ws.cell(row=r, column=inc_dec_idx + 1).value)
                   if inc_dec_idx is not None else None)
        remarks = None
        if remarks_idx is not None:
            remarks = _to_str(ws.cell(row=r, column=remarks_idx + 1).value) or None
        rows.append(FSRow(sheet=sheet_label, row_number=r,
                          account_code=account_code, account_title=title,
                          columns=cols, inc_dec=inc_dec, remarks=remarks))
    return headers, rows, company
