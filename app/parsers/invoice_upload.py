from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


class InvoiceUploadParseError(ValueError):
    pass


@dataclass
class ParsedInvoice:
    invoice_number: str | None
    invoice_date: str | None
    supplier_name: str
    description: str | None
    gross_amount: Decimal
    vat_amount: Decimal | None
    ewt_amount: Decimal | None
    net_amount: Decimal
    due_date: str | None


HEADER_ALIASES = {
    "invoice_number": {"invoice no", "invoice number", "invoice #", "invoice", "receipt no", "receipt number", "reference", "ref no", "document no"},
    "invoice_date": {"date", "invoice date", "receipt date", "document date", "posting date"},
    "supplier_name": {"supplier", "vendor", "payee", "beneficiary", "counterparty", "supplier name", "vendor name"},
    "description": {"description", "particulars", "remarks", "memo", "details"},
    "gross_amount": {"gross amount", "gross", "total amount", "amount", "invoice amount", "receipt amount"},
    "vat_amount": {"vat", "input vat", "vat amount", "tax"},
    "ewt_amount": {"ewt", "withholding", "withholding tax", "expanded withholding tax"},
    "net_amount": {"net amount", "net", "amount paid", "payment amount", "payable amount"},
    "due_date": {"due date", "payment due", "maturity date"},
}


def _norm(value) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").replace("-", " ").split())


def _to_decimal(value, field: str, row_no: int, required: bool = False) -> Decimal | None:
    if value is None or str(value).strip() == "":
        if required:
            raise InvoiceUploadParseError(f"Missing {field} on row {row_no}")
        return None
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise InvoiceUploadParseError(f"Invalid {field} {value!r} on row {row_no}") from exc


def _to_date(value, field: str, row_no: int) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise InvoiceUploadParseError(f"Invalid {field} {value!r} on row {row_no}")


def _find_header(ws) -> tuple[int, dict[str, int]]:
    for row_no in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row_no, col).value for col in range(1, ws.max_column + 1)]
        normalized = [_norm(v) for v in values]
        mapping: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for idx, header in enumerate(normalized):
                if header in aliases:
                    mapping[field] = idx + 1
                    break
        if "supplier_name" in mapping and ("gross_amount" in mapping or "net_amount" in mapping):
            return row_no, mapping
    raise InvoiceUploadParseError("Could not find invoice header row. Required columns include supplier/vendor and amount/net/gross amount.")


def parse_invoice_file(path: str | Path) -> list[ParsedInvoice]:
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise InvoiceUploadParseError("Invoice upload supports .xlsx and .xlsm files only.")
    wb = load_workbook(path, data_only=True, read_only=True)
    invoices: list[ParsedInvoice] = []
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        try:
            header_row, mapping = _find_header(ws)
        except InvoiceUploadParseError:
            continue
        for row_no in range(header_row + 1, ws.max_row + 1):
            def val(field: str):
                col = mapping.get(field)
                return ws.cell(row_no, col).value if col else None
            supplier = str(val("supplier_name") or "").strip()
            if not supplier:
                continue
            gross = _to_decimal(val("gross_amount"), "gross amount", row_no)
            net = _to_decimal(val("net_amount"), "net amount", row_no)
            if gross is None and net is None:
                continue
            vat = _to_decimal(val("vat_amount"), "VAT", row_no)
            ewt = _to_decimal(val("ewt_amount"), "EWT", row_no)
            if net is None:
                net = gross or Decimal("0")
                if ewt:
                    net -= ewt
            if gross is None:
                gross = net
                if ewt:
                    gross += ewt
            invoices.append(ParsedInvoice(
                invoice_number=str(val("invoice_number") or "").strip() or None,
                invoice_date=_to_date(val("invoice_date"), "invoice date", row_no),
                supplier_name=supplier,
                description=str(val("description") or "").strip() or None,
                gross_amount=gross,
                vat_amount=vat,
                ewt_amount=ewt,
                net_amount=net,
                due_date=_to_date(val("due_date"), "due date", row_no),
            ))
    if not invoices:
        raise InvoiceUploadParseError("No invoice rows were found in the workbook.")
    return invoices
