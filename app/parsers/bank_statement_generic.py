from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class GenericStatementParseError(ValueError):
    pass


@dataclass
class GenericTransaction:
    row_number: int
    date: str
    description: str
    reference: str | None
    debit: Decimal
    credit: Decimal
    amount: Decimal
    balance: Decimal | None
    bank_profile: str | None = None


# Generic, header-based parser. It does not use bank-specific labels in the UI
# and it never creates synthetic fallback rows. Real uploaded files must contain
# recognizable columns for date, description/reference, and amount or debit/credit.
COMMON_ALIASES = {
    "date": {
        "date", "transaction date", "txn date", "posting date", "post date", "posted date", "value date",
        "effective date", "doc date", "document date", "entry date", "je date", "journal date",
        "payment date", "invoice date", "receipt date", "clearing date", "report date", "period date", "check date", "cheque date", "transaction posted", "date posted",
        "booking date", "record date", "posted on", "created date", "created at",
        "voucher date", "journal entry date",
    },
    "description": {
        "description", "particulars", "details", "remarks", "memo", "narration", "narrative",
        "transaction details", "transaction description", "payee", "beneficiary", "supplier", "supplier name",
        "vendor", "vendor name", "customer", "account title", "account name", "name", "line description",
        "journal description", "merchant", "counterparty", "paid to", "received from", "purpose", "item",
        "transaction", "explanation", "comment", "details of transaction", "particular", "payer", "client",
    },
    "reference": {
        "reference", "ref", "ref no", "reference no", "reference number", "transaction id", "transaction no",
        "trace no", "trace number", "serial no", "doc no", "document no", "document number",
        "je no", "journal no", "journal entry no", "voucher no", "payment ref", "payment reference",
        "invoice no", "invoice number", "receipt no", "receipt number", "or no", "official receipt",
        "control no", "batch no", "record id", "id", "external id", "source id", "or number", "si number", "po number", "check no", "cheque no", "check number", "cheque number",
    },
    "debit": {
        "debit", "debits", "withdrawal", "withdrawals", "outflow", "dr", "debit amount", "dr amount",
        "charges", "disbursement", "disbursements", "expense", "less", "deduction",
    },
    "credit": {
        "credit", "credits", "deposit", "deposits", "inflow", "cr", "credit amount", "cr amount",
        "receipt", "receipts", "collection", "collections", "received", "income", "proceeds", "additions", "add",
    },
    "amount": {
        "amount", "net amount", "transaction amount", "total amount", "gross amount", "payment amount",
        "paid amount", "receipt amount", "invoice amount", "book amount", "system amount", "statement amount",
        "source amount", "comparison amount", "base amount", "value", "peso amount", "php amount",
        "balance movement", "movement", "variance amount",
        # Accounting system export column names (e.g. PCF PER SYSTEM.XLS)
        "amt local curr", "amt in local curr", "amt in trans curr", "amt in transaction curr",
        "amount local curr", "amount in local curr", "local amount", "local curr amount",
    },
    "amount_type": {
        "type", "transaction type", "dr/cr", "debit/credit", "debit credit", "dc", "d/c", "cr/dr",
        "direction", "sign", "movement type",
    },
    "balance": {
        "balance", "ending balance", "running balance", "closing balance", "available balance", "remaining balance",
    },
}

PROFILE_ALIASES = {field: set(values) for field, values in COMMON_ALIASES.items()}

def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\n\r\t_\-]+", " ", text)
    text = re.sub(r"[^a-z0-9/ ]+", " ", text)
    return " ".join(text.split())


def _header_matches(header: str, alias: str) -> bool:
    if not header or not alias:
        return False
    if header == alias:
        return True
    # Allow common decorated headers: "Debit (PHP)", "Credit Amount PHP".
    return alias in header and (len(alias) >= 5 or header.startswith(alias + " "))


def _column_for(headers: list[Any], aliases: set[str], used: set[int]) -> int | None:
    normalized = [_norm(h) for h in headers]
    normalized_aliases = {_norm(a) for a in aliases}
    for idx, header in enumerate(normalized):
        if idx not in used and header in normalized_aliases:
            return idx
    for idx, header in enumerate(normalized):
        if idx in used:
            continue
        for alias in normalized_aliases:
            if _header_matches(header, alias):
                return idx
    return None


def _detect_profile(headers: list[Any]) -> str | None:
    # Kept for backwards compatibility with older records/tests. The parser is
    # intentionally generic and does not expose bank-specific profiles.
    return None


def _mapping(headers: list[Any]) -> tuple[dict[str, int], str | None]:
    used: set[int] = set()
    mapping: dict[str, int] = {}
    profile_name = None

    field_order = ["date", "description", "reference", "debit", "credit", "amount", "amount_type", "balance"]
    for field in field_order:
        aliases = set(PROFILE_ALIASES.get(field, set()))
        col = _column_for(headers, aliases, used)
        if col is not None:
            mapping[field] = col
            # debit/credit/amount columns should not be reused; description and reference may be reused if needed.
            if field not in {"description", "reference"}:
                used.add(col)

    if "description" not in mapping and "reference" in mapping:
        mapping["description"] = mapping["reference"]
    if "reference" not in mapping and "description" in mapping:
        mapping["reference"] = mapping["description"]

    if "date" not in mapping or "description" not in mapping or not ({"amount", "debit", "credit"} & set(mapping)):
        found = ", ".join(_norm(h) for h in headers if _norm(h))
        raise GenericStatementParseError(
            "Required columns not found. Need date, description/reference, and amount or debit/credit columns. "
            f"Detected headers: {found or 'none'}"
        )
    return mapping, profile_name


def _decimal(value: Any, row_no: int, field: str, default: Decimal | None = Decimal("0")) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return default
    text = str(value).strip()
    upper = text.upper()
    sign = Decimal("1")
    if " DR" in f" {upper}" or upper.endswith("DR"):
        sign = Decimal("-1")
    if " CR" in f" {upper}" or upper.endswith("CR"):
        sign = Decimal("1")
    text = re.sub(r"(?i)\b(PHP|PHP\.|PESO|PESOS|DR|CR)\b", "", text)
    text = text.replace(",", "").replace("₱", "").strip()
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", "."}:
        return default
    try:
        return Decimal(text) * sign
    except InvalidOperation as exc:
        raise GenericStatementParseError(f"Invalid {field} {value!r} on row {row_no}") from exc


def _date(value: Any, row_no: int) -> str:
    if value is None or str(value).strip() == "":
        return ""  # blank dates are tolerated; reconciliation matches by amount only
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial date fallback for CSV exports that preserved serial numbers.
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception as exc:
            raise GenericStatementParseError(f"Invalid date {value!r} on row {row_no}") from exc
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    # Strip time component if present (e.g. "5/30/2026 0:33", "2026-05-30 14:05")
    text_date_only = re.split(r"\s+\d{1,2}:\d{2}", text)[0].strip()

    for candidate in (text, text_date_only):
        for fmt in (
            "%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y",
            "%m/%d/%y", "%d/%m/%y", "%m-%d-%y", "%d-%m-%y", "%b %d %Y", "%d %b %Y",
            "%B %d %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y", "%d-%B-%Y", "%b-%d-%Y", "%B-%d-%Y",
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
            "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
        ):
            try:
                return datetime.strptime(candidate, fmt).date().isoformat()
            except ValueError:
                pass
    raise GenericStatementParseError(f"Invalid date {value!r} on row {row_no}")


def _type_sign(value: Any) -> int | None:
    # Handle numeric Dr/Cr indicators from accounting system exports (1.0 = debit, -1.0 = credit)
    if isinstance(value, (int, float)):
        if value == 1.0:
            return -1  # debit → outflow
        if value == -1.0:
            return 1   # credit → inflow
        return None
    text = _norm(value).replace("/", " ")
    if not text:
        return None
    if text in {"dr", "debit", "withdrawal", "withdrawals", "payment", "payments", "outflow"}:
        return -1
    if text in {"cr", "credit", "deposit", "deposits", "receipt", "receipts", "inflow"}:
        return 1
    return None


def _looks_like_total_row(row: list[Any]) -> bool:
    joined = " ".join(_norm(c) for c in row if c is not None)
    return any(token in joined for token in ("opening balance", "closing balance", "total debit", "total credit", "grand total", "beginning balance"))



def _is_probable_date(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return False
    try:
        _date(value, 0)
        return True
    except Exception:
        return False


def _is_probable_number(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return False
    try:
        return _decimal(value, 0, "amount", None) is not None
    except Exception:
        return False


def _infer_mapping_from_data(rows: list[list[Any]]) -> tuple[int, dict[str, int], str | None] | None:
    """Fallback for messy bank/reconciliation files with weak or unusual headers."""
    best = None
    for header_idx in range(0, min(len(rows), 30)):
        sample = rows[header_idx + 1: header_idx + 21]
        if not sample:
            continue
        max_cols = max((len(r) for r in sample), default=0)
        date_scores = []
        num_scores = []
        text_scores = []
        for col in range(max_cols):
            vals = [r[col] for r in sample if col < len(r)]
            date_scores.append(sum(1 for v in vals if _is_probable_date(v)))
            num_scores.append(sum(1 for v in vals if _is_probable_number(v)))
            text_scores.append(sum(1 for v in vals if str(v or "").strip() and not _is_probable_number(v) and not _is_probable_date(v)))
        if not date_scores:
            continue
        date_col = max(range(max_cols), key=lambda c: date_scores[c])
        amount_candidates = [c for c in range(max_cols) if c != date_col and num_scores[c] >= 2]
        text_candidates = [c for c in range(max_cols) if c != date_col and text_scores[c] >= 2]
        if date_scores[date_col] >= 2 and amount_candidates and text_candidates:
            amount_col = max(amount_candidates, key=lambda c: num_scores[c])
            desc_col = max(text_candidates, key=lambda c: text_scores[c])
            ref_candidates = [c for c in text_candidates if c != desc_col]
            mapping = {"date": date_col, "description": desc_col, "reference": (ref_candidates[0] if ref_candidates else desc_col), "amount": amount_col}
            score = date_scores[date_col] + num_scores[amount_col] + text_scores[desc_col]
            if best is None or score > best[0]:
                best = (score, header_idx, mapping)
    if best:
        return best[1], best[2], None
    return None

def _build(rows: list[list[Any]]) -> list[GenericTransaction]:
    if not rows:
        raise GenericStatementParseError("The file is empty.")
    header_idx = None
    mapping = None
    profile_name = None
    for idx, row in enumerate(rows[:50]):
        try:
            mapping, profile_name = _mapping(row)
            header_idx = idx
            break
        except GenericStatementParseError:
            continue
    if mapping is None or header_idx is None:
        inferred = _infer_mapping_from_data(rows)
        if inferred is None:
            raise GenericStatementParseError("Could not detect a valid header row. Please make sure the file has recognizable date, description/reference, and amount/debit/credit columns.")
        header_idx, mapping, profile_name = inferred

    out: list[GenericTransaction] = []
    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        def v(field: str):
            col = mapping.get(field)
            return row[col] if col is not None and col < len(row) else None

        if all((c is None or str(c).strip() == "") for c in row):
            continue
        if _looks_like_total_row(row):
            continue
        desc = str(v("description") or "").strip()
        ref = str(v("reference") or "").strip() or None
        if not desc and not ref:
            continue
        debit = _decimal(v("debit"), idx, "debit") or Decimal("0")
        credit = _decimal(v("credit"), idx, "credit") or Decimal("0")
        amount_val = _decimal(v("amount"), idx, "amount", None)
        if amount_val is not None:
            sign = _type_sign(v("amount_type"))
            if sign == -1 and amount_val > 0:
                amount_val = -amount_val
            elif sign == 1 and amount_val < 0:
                amount_val = abs(amount_val)
            if amount_val < 0:
                debit = abs(amount_val)
                credit = Decimal("0")
            else:
                debit = Decimal("0")
                credit = amount_val
            amount = amount_val
        else:
            amount = credit - debit
        out.append(GenericTransaction(
            row_number=idx,
            date=_date(v("date"), idx),
            description=desc or ref or "",
            reference=ref,
            debit=abs(debit),
            credit=abs(credit),
            amount=amount,
            balance=_decimal(v("balance"), idx, "balance", None),
            bank_profile=profile_name,
        ))
    if not out:
        raise GenericStatementParseError("No transaction rows found.")
    return out


def _xlsx_rows(path: Path) -> list[list[Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    all_errors = []
    for ws in wb.worksheets:
        rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        try:
            return _build(rows)
        except GenericStatementParseError as exc:
            all_errors.append(f"{ws.title}: {exc}")
    raise GenericStatementParseError("; ".join(all_errors) or "No readable worksheets.")


def _xls_rows(path: Path) -> list[GenericTransaction]:
    try:
        import xlrd
    except ImportError as exc:
        raise GenericStatementParseError("xlrd is required to parse .xls bank statements") from exc
    wb = xlrd.open_workbook(str(path))
    errors = []
    for sheet in wb.sheets():
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            values: list[Any] = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode))
                else:
                    values.append(cell.value)
            rows.append(values)
        try:
            return _build(rows)
        except GenericStatementParseError as exc:
            errors.append(f"{sheet.name}: {exc}")
    raise GenericStatementParseError("; ".join(errors) or "No readable worksheets.")


def _is_pcf_system_xls(path: Path) -> bool:
    """Return True if an .xls file looks like the PCF accounting-system export
    (two sheets: Detail + Header, with Dr/Cr numeric flags in Detail)."""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(path))
        names_lower = {s.lower() for s in wb.sheet_names()}
        if "detail" not in names_lower or "header" not in names_lower:
            return False
        dsh = wb.sheet_by_name(next(s for s in wb.sheet_names() if s.lower() == "detail"))
        # Confirm Dr/Cr column contains numeric 1.0/-1.0 values
        for r in range(3, min(dsh.nrows, 10)):
            val = dsh.cell(r, 1).value
            if val in (1.0, -1.0):
                return True
    except Exception:
        pass
    return False


def _is_pcf_excel_xlsx(path: Path) -> bool:
    """Return True if an .xlsx file looks like the PCF petty-cash Excel workbook.

    Detection strategy (two independent signals, either is sufficient):
    1. Row 0, col 0 contains "Petty Cash" (the fixed header cell).
    2. Row 11 (0-based), cols 41/45/50/55 contain week labels like "WEEK 5".
    """
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        # Signal 1: header cell in row 0, col 0
        first_rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if first_rows:
            cell0 = str(first_rows[0][0] if first_rows[0] else "").lower()
            if "petty cash" in cell0:
                wb.close()
                return True
        # Signal 2: week labels in row 12 (1-based) = row 11 (0-based)
        week_rows = list(ws.iter_rows(min_row=12, max_row=12, values_only=True))
        wb.close()
        if week_rows:
            row11 = week_rows[0]
            # Cols 41, 45, 50, 55 (0-based) hold the week labels
            for col_idx in (41, 45, 50, 55):
                if col_idx < len(row11) and row11[col_idx] is not None:
                    label = str(row11[col_idx]).upper()
                    if "WEEK" in label:
                        return True
    except Exception:
        pass
    return False


def parse_generic_statement(path: str | Path) -> list[GenericTransaction]:
    path = Path(path)
    suffix = path.suffix.lower()

    # ── Auto-detect PCF-specific formats ─────────────────────────────────────
    stem = path.stem.lower()
    if suffix == ".xls":
        # Detect by filename first, then by file structure
        if ("pcf" in stem and "system" in stem) or _is_pcf_system_xls(path):
            from app.parsers.pcf_system_parser import parse_pcf_system
            return parse_pcf_system(path)
    elif suffix in {".xlsx", ".xlsm"}:
        if ("pcf" in stem and "excel" in stem) or _is_pcf_excel_xlsx(path):
            from app.parsers.pcf_excel_parser import parse_pcf_excel
            return parse_pcf_excel(path)

    # ── Generic format handling ───────────────────────────────────────────────
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as f:
            return _build([row for row in csv.reader(f)])
    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_rows(path)
    if suffix == ".xls":
        return _xls_rows(path)
    raise GenericStatementParseError("Supported formats are .xlsx, .xlsm, .xls, and .csv.")
