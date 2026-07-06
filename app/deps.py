"""
app/deps.py — Shared dependencies for all BookPoint routers.

Contains:
  - BASE_DIR, templates (Jinja2Templates singleton), directory constants
  - Jinja2 filter registrations (money, english_date)
  - All shared helper functions used by multiple routers

Import this module in each router:
    from app.deps import templates, BASE_DIR, _current_user, ...

Never import from app.main or any router here to avoid circular imports.
"""
from __future__ import annotations

import base64
import hashlib
import json
import os
import smtplib
from datetime import date as _date, timedelta as _timedelta
from decimal import Decimal, InvalidOperation
from email.message import EmailMessage
from pathlib import Path
from types import SimpleNamespace

from cryptography.fernet import Fernet, InvalidToken
from fastapi import HTTPException, Request
from fastapi.templating import Jinja2Templates

from app.auth import require_user, SECRET
from app.db import db

# ---------------------------------------------------------------------------
# Paths and singleton objects
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent

templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

COMMISSION_CARRY_DIR = Path(
    os.environ.get("COMMISSION_CARRY_DIR", str(BASE_DIR.parent / "data" / "commission_carry"))
)
COMMISSION_CARRY_DIR.mkdir(parents=True, exist_ok=True)

REQUEST_UPLOAD_DIR = (
    Path(os.environ.get("ACCOUNTING_UPLOAD_DIR", str(BASE_DIR.parent / "uploads")))
    / "request_attachments"
)

# ---------------------------------------------------------------------------
# Jinja2 filters
# ---------------------------------------------------------------------------

date = _date  # alias used in filter functions


def _fmt_money(v) -> str:
    if v is None or v == "":
        return ""
    try:
        d = Decimal(str(v))
    except (ArithmeticError, ValueError):
        return str(v)
    return f"{d:,.2f}"


templates.env.filters["money"] = _fmt_money


def _fmt_english_date(v) -> str:
    if not v:
        return "-"
    raw = str(v).strip()
    try:
        d = _date.fromisoformat(raw[:10])
        return d.strftime("%B %d, %Y")
    except Exception:
        return raw


templates.env.filters["english_date"] = _fmt_english_date

# ---------------------------------------------------------------------------
# Simple data helpers
# ---------------------------------------------------------------------------


def _get_expense_accounts():
    with db() as conn:
        return conn.execute(
            """SELECT id, code, name FROM chart_of_accounts
               WHERE is_active = 1 AND type = 'expense'
               ORDER BY code"""
        ).fetchall()


def _validate_iso_date_or_400(value: str, field: str = "date") -> str:
    value = (value or "").strip()
    try:
        _date.fromisoformat(value)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field}; use YYYY-MM-DD") from exc
    return value


def _validate_time_or_none(value: str, field: str):
    value = (value or "").strip()
    if not value:
        return None
    parts = value.split(":")
    if len(parts) < 2:
        raise HTTPException(status_code=400, detail=f"Invalid {field}; use HH:MM")
    try:
        hh, mm = int(parts[0]), int(parts[1])
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field}; use HH:MM") from exc
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        raise HTTPException(status_code=400, detail=f"Invalid {field}; use HH:MM")
    return f"{hh:02d}:{mm:02d}"


def _list_companies():
    with db() as conn:
        return conn.execute("SELECT * FROM companies WHERE is_active = 1 ORDER BY name").fetchall()


def _list_departments(company_id: int | None = None):
    with db() as conn:
        if company_id:
            return conn.execute(
                "SELECT d.*, c.name AS company_name FROM departments d LEFT JOIN companies c ON c.id = COALESCE(d.company_id, 1) WHERE d.is_active = 1 AND COALESCE(d.company_id, 1) = ? ORDER BY d.name",
                (company_id,),
            ).fetchall()
        return conn.execute(
            "SELECT d.*, c.name AS company_name FROM departments d LEFT JOIN companies c ON c.id = COALESCE(d.company_id, 1) WHERE d.is_active = 1 ORDER BY c.name, d.name"
        ).fetchall()


def _department_belongs_to_company(department_id: int | None, company_id: int | None) -> bool:
    if not department_id:
        return True
    with db() as conn:
        row = conn.execute(
            "SELECT id FROM departments WHERE id = ? AND is_active = 1 AND COALESCE(company_id, 1) = ?",
            (department_id, company_id or 1),
        ).fetchone()
    return row is not None


# ---------------------------------------------------------------------------
# User role helpers
# ---------------------------------------------------------------------------


def _current_user(request: Request):
    return require_user(request)


def _is_accounting_user(user) -> bool:
    """Return True when the logged-in user belongs to Accounting.

    Admin and accountant roles are treated as Accounting staff. Department users
    can also access Accounting-only features when their department is named
    Accounting.
    """
    if user is None:
        return False
    if user["role"] in {"admin", "accountant"}:
        return True
    try:
        department_id = user["department_id"]
    except KeyError:
        department_id = None
    if not department_id:
        return False
    with db() as conn:
        row = conn.execute(
            "SELECT name FROM departments WHERE id = ? AND is_active = 1", (department_id,)
        ).fetchone()
    return bool(row and (row["name"] or "").strip().lower() == "accounting")


def _is_operations_manager(user) -> bool:
    if user is None:
        return False
    # Some User rows/sessions (e.g. minted before this flag existed, or via a
    # code path that selects a partial column set) may not carry this key at
    # all. Treat a missing flag the same as "not an operations manager"
    # instead of letting a bare KeyError crash the request — mirrors the
    # defensive pattern already used above in _is_accounting_user.
    try:
        flag = user["is_operations_manager"]
    except KeyError:
        flag = 0
    return bool(int(flag or 0) == 1)


def _operations_access_company_ids(conn, user) -> list[int]:
    if not _is_operations_manager(user):
        return []
    rows = conn.execute(
        """SELECT c.id FROM companies c
           JOIN operations_manager_company_access a ON a.company_id = c.id
           WHERE a.user_id = ? AND c.is_active = 1 ORDER BY c.name""",
        (user["id"],),
    ).fetchall()
    if rows:
        return [int(r["id"]) for r in rows]
    return [int(user["company_id"] or 1)]


def _operations_access_companies(conn, user):
    ids = _operations_access_company_ids(conn, user)
    if not ids:
        return []
    placeholders = ",".join("?" for _ in ids)
    return conn.execute(
        f"SELECT * FROM companies WHERE id IN ({placeholders}) AND is_active = 1 ORDER BY name",
        ids,
    ).fetchall()


def _can_access_request(conn, user, pr) -> bool:
    try:
        is_draft = int(pr["is_draft"] or 0)
    except (KeyError, IndexError, TypeError):
        is_draft = 0
    if is_draft:
        return pr["requester_user_id"] == user["id"]
    req_company_id = int(pr["company_id"] or 1)
    if user["role"] == "admin" and not _is_operations_manager(user):
        return True
    if _is_operations_manager(user):
        return req_company_id in _operations_access_company_ids(conn, user) and _needs_operations_approval(
            pr["request_type"], pr["amount"]
        )
    if _is_accounting_user(user):
        return req_company_id == int(user["company_id"] or 1)
    return pr["requester_user_id"] == user["id"]


def _request_amount(value) -> Decimal:
    try:
        return Decimal(str(value).replace(",", ""))
    except Exception:
        return Decimal("0")


def _needs_operations_approval(request_type: str, amount) -> bool:
    return request_type == "reimbursement" and _request_amount(amount) > Decimal("50000")


def require_accounting_calendar_access(request: Request):
    user = getattr(request.state, "user", None)
    if not _is_accounting_user(user):
        raise HTTPException(status_code=403, detail="Calendar is available to Accounting users only")
    return user


# ---------------------------------------------------------------------------
# Credential encryption
# ---------------------------------------------------------------------------

_SENSITIVE_SETTING_KEYS: frozenset[str] = frozenset({"smtp_password"})
_ENC_PREFIX = "enc1:"


def _get_settings_fernet() -> Fernet:
    """Return a Fernet instance keyed from the app secret (deterministic, stable)."""
    key_bytes = hashlib.pbkdf2_hmac(
        "sha256", SECRET.encode("utf-8"), b"bookpoint-settings-v1", 200_000, dklen=32
    )
    return Fernet(base64.b64encode(key_bytes))


def _encrypt_setting(value: str) -> str:
    """Encrypt a sensitive setting value for storage.  Empty strings are stored as-is."""
    if not value:
        return value
    return _ENC_PREFIX + _get_settings_fernet().encrypt(value.encode("utf-8")).decode("ascii")


def _decrypt_setting(value: str) -> str:
    """Decrypt a sensitive setting value.  Falls back to the raw value for legacy plain-text rows."""
    if not value:
        return value
    if not value.startswith(_ENC_PREFIX):
        return value
    try:
        return _get_settings_fernet().decrypt(value[len(_ENC_PREFIX):].encode("ascii")).decode("utf-8")
    except (InvalidToken, Exception):
        return ""


def _get_app_settings(conn=None) -> dict:
    def _inner(c) -> dict:
        try:
            rows = c.execute("SELECT key, value FROM app_settings").fetchall()
            result: dict = {}
            for r in rows:
                k, v = r["key"], r["value"] or ""
                result[k] = _decrypt_setting(v) if k in _SENSITIVE_SETTING_KEYS else v
            return result
        except Exception as _e:
            print(f"[WARN] _get_app_settings: could not read settings: {_e}")
            return {}

    if conn is not None:
        return _inner(conn)
    with db() as _conn:
        return _inner(_conn)


def _save_app_settings(settings: dict) -> None:
    with db() as conn:
        for key, value in settings.items():
            stored = _encrypt_setting(value) if key in _SENSITIVE_SETTING_KEYS else value
            conn.execute(
                """INSERT INTO app_settings (key, value, updated_at) VALUES (?, ?, datetime('now'))
                   ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = datetime('now')""",
                (key, stored),
            )


# ---------------------------------------------------------------------------
# Email helpers
# ---------------------------------------------------------------------------


def _send_email_notification(
    to_email: str,
    subject: str,
    body: str,
    attachment_path: str | None = None,
    attachment_name: str | None = None,
) -> None:
    """Send email when SMTP is configured in the admin UI or environment.

    Optionally attaches a file (e.g. payment receipt) to the message.
    Falls back silently if SMTP is not configured or sending fails.
    """
    settings = _get_app_settings()
    host = (settings.get("smtp_host") or os.environ.get("SMTP_HOST", "")).strip()
    if not host or not to_email:
        return
    msg = EmailMessage()
    msg["Subject"] = subject
    username = (settings.get("smtp_user") or os.environ.get("SMTP_USER", "")).strip()
    msg["From"] = (
        settings.get("smtp_from") or os.environ.get("SMTP_FROM") or username or ""
    ).strip()
    msg["To"] = to_email
    msg.set_content(body)
    if attachment_path:
        try:
            att_path = Path(attachment_path)
            if att_path.exists() and att_path.stat().st_size <= 10 * 1024 * 1024:
                import mimetypes

                ctype, _ = mimetypes.guess_type(str(att_path))
                maintype, subtype = (ctype or "application/octet-stream").split("/", 1)
                with att_path.open("rb") as fh:
                    msg.add_attachment(
                        fh.read(),
                        maintype=maintype,
                        subtype=subtype,
                        filename=attachment_name or att_path.name,
                    )
        except Exception:
            pass
    try:
        port = int(
            (settings.get("smtp_port") or os.environ.get("SMTP_PORT", "587")).strip() or "587"
        )
    except Exception:
        port = 587
    password = settings.get("smtp_password") or os.environ.get("SMTP_PASSWORD", "")
    tls_value = (settings.get("smtp_tls") or os.environ.get("SMTP_TLS", "1")).strip().lower()
    use_tls = tls_value not in {"0", "false", "no", "off"}
    try:
        with smtplib.SMTP(host, port, timeout=10) as smtp:
            if use_tls:
                smtp.starttls()
            if username:
                smtp.login(username, password)
            smtp.send_message(msg)
    except Exception:
        return


def _email_for_user(user_row) -> str:
    if not user_row:
        return ""
    try:
        return (user_row["email"] or "").strip()
    except KeyError:
        return ""


def _absolute_app_link(link_url: str) -> str:
    base_url = os.environ.get("APP_BASE_URL", "").strip().rstrip("/")
    if not base_url:
        return link_url
    if link_url.startswith("http://") or link_url.startswith("https://"):
        return link_url
    return base_url + (link_url if link_url.startswith("/") else f"/{link_url}")


# ---------------------------------------------------------------------------
# Notification helpers
# ---------------------------------------------------------------------------


def _notify(
    conn,
    user_id: int,
    title: str,
    message: str,
    link_url: str = "/requests",
    send_email: bool = True,
) -> None:
    conn.execute(
        "INSERT INTO notifications (user_id, title, message, link_url) VALUES (?, ?, ?, ?)",
        (user_id, title, message, link_url),
    )
    if send_email:
        user_row = conn.execute(
            "SELECT username, email FROM users WHERE id = ? AND is_active = 1", (user_id,)
        ).fetchone()
        to_email = _email_for_user(user_row) if user_row else ""
        if to_email:
            open_link = _absolute_app_link(link_url)
            _send_email_notification(
                to_email,
                f"BookPoint: {title}",
                f"{message}\n\nOpen request: {open_link}",
            )


def _notify_once(
    conn,
    user_id: int,
    title: str,
    message: str,
    link_url: str = "/requests",
    send_email: bool = True,
    window_minutes: int = 5,
) -> None:
    """Insert a notification only once within a short window."""
    existing = conn.execute(
        """SELECT id FROM notifications
           WHERE user_id = ? AND title = ? AND message = ? AND COALESCE(link_url, '') = COALESCE(?, '')
             AND created_at >= datetime('now', ?)
           ORDER BY id DESC LIMIT 1""",
        (user_id, title, message, link_url, f"-{int(window_minutes)} minutes"),
    ).fetchone()
    if existing:
        return
    _notify(conn, user_id, title, message, link_url, send_email=send_email)


def _notify_requesting_department(
    conn,
    request_id: int,
    company_id: int,
    department_id: int | None,
    requester_user_id: int,
    title: str,
    message: str,
) -> None:
    """Notify the requester and other active users assigned to the same requesting department."""
    link = f"/requests/{request_id}"
    _notify_once(conn, requester_user_id, title, message, link, send_email=False)
    if not department_id:
        return
    users = conn.execute(
        """SELECT id FROM users
           WHERE is_active = 1
             AND COALESCE(company_id, 1) = ?
             AND department_id = ?
             AND id <> ?""",
        (company_id, department_id, requester_user_id),
    ).fetchall()
    for u in users:
        _notify_once(conn, u["id"], title, message, link, send_email=False)


def _notify_accounting(
    conn,
    title: str,
    message: str,
    link_url: str,
    company_id: int | None = None,
) -> None:
    params = []
    where_company = ""
    if company_id:
        where_company = " AND COALESCE(company_id, 1) = ?"
        params.append(company_id)
    for u in conn.execute(
        f"SELECT id FROM users WHERE is_active = 1 AND role IN ('admin','accountant') {where_company}",
        params,
    ).fetchall():
        _notify_once(conn, u["id"], title, message, link_url)


def _notify_operations_managers(
    conn,
    title: str,
    message: str,
    link_url: str,
    company_id: int | None = None,
) -> None:
    params = []
    where_access = ""
    if company_id:
        where_access = " AND (COALESCE(u.company_id, 1) = ? OR EXISTS (SELECT 1 FROM operations_manager_company_access a WHERE a.user_id = u.id AND a.company_id = ?))"
        params.extend([company_id, company_id])
    for u in conn.execute(
        f"SELECT DISTINCT u.id FROM users u WHERE u.is_active = 1 AND u.is_operations_manager = 1 AND u.notify_operations_approvals = 1 {where_access}",
        params,
    ).fetchall():
        _notify_once(conn, u["id"], title, message, link_url)


# ---------------------------------------------------------------------------
# File / data utilities
# ---------------------------------------------------------------------------


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _recent_month_options(count: int = 18):
    today = _date.today().replace(day=1)
    opts = []
    y, m = today.year, today.month
    for _ in range(count):
        value = f"{y:04d}-{m:02d}"
        label = _date(y, m, 1).strftime("%B %Y")
        opts.append({"value": value, "label": label})
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return opts


def _week_range(week_str: str) -> tuple[str, str]:
    """Parse 'YYYY-Www' (e.g. '2026-W20') and return (monday_iso, sunday_iso)."""
    import re as _re

    m = _re.fullmatch(r"(\d{4})-W(\d{2})", (week_str or "").strip())
    if not m:
        raise HTTPException(status_code=400, detail="week must be YYYY-Www, e.g. 2026-W20")
    yr, wk = int(m.group(1)), int(m.group(2))
    monday = _date.fromisocalendar(yr, wk, 1)
    sunday = _date.fromisocalendar(yr, wk, 7)
    return monday.isoformat(), sunday.isoformat()


def _recent_week_options(count: int = 26):
    """Return the last N ISO weeks as [{"value": "YYYY-Www", "label": "..."}]."""
    opts = []
    d = _date.today()
    d = d - _timedelta(days=d.weekday())
    for _ in range(count):
        iso = d.isocalendar()
        value = f"{iso[0]:04d}-W{iso[1]:02d}"
        end = d + _timedelta(days=6)
        label = f"Week {iso[1]} · {d.strftime('%b %d')}–{end.strftime('%b %d, %Y')}"
        opts.append({"value": value, "label": label})
        d -= _timedelta(weeks=1)
    return opts


def _parse_iso(s):
    if not s:
        return None
    try:
        return _date.fromisoformat(s.strip())
    except Exception:
        return None


def _build_audit_rows(raw_rows):
    rows_with_details = []
    for r in raw_rows:
        d = dict(r)
        if d["details_json"]:
            try:
                d["details"] = json.loads(d["details_json"])
            except Exception:
                d["details"] = d["details_json"]
        else:
            d["details"] = None
        rows_with_details.append(d)
    return rows_with_details


# ---------------------------------------------------------------------------
# Account-titles helpers
# ---------------------------------------------------------------------------


def _norm_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _parse_account_titles_file(path: Path):
    """Parse an account-titles file and return a list of (code, name, type, is_active) tuples."""
    from app.services.historical_journal_learning import (
        _is_rac_format,
        _read_rows as _hjl_read_rows,
        extract_chart_of_accounts_from_journal,
    )

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xlsm", ".xls"}:
        raise HTTPException(
            status_code=400,
            detail="Supported formats are .xlsx, .xlsm, .xls, and .csv",
        )

    try:
        rows = _hjl_read_rows(path)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read file: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=422, detail="Uploaded file is empty")

    if _is_rac_format(rows):
        try:
            tuples = extract_chart_of_accounts_from_journal(path)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        if not tuples:
            raise HTTPException(
                status_code=422,
                detail="No account codes found in the RAC Phil Corp journal file",
            )
        return tuples

    # ── Subsidiary ledger format: header row contains "A/C No." ───────────────
    # The PER SYSTEM .xls export has alternating account-header rows (col 0 =
    # numeric account code, col 1 = account title) and detail rows (col 0 blank).
    # We detect this layout by finding "A/C No." in the first 15 rows and then
    # extract one (code, title) per account-header row.
    def _is_subsidiary_ledger(rows_sample):
        for row in rows_sample:
            joined = " ".join(str(c or "").strip() for c in row[:4])
            if "A/C No" in joined:
                return True
        return False

    def _infer_account_type(code: str) -> str:
        c = str(code).strip().lstrip("0") or "0"
        if c.startswith("1"):
            return "asset"
        if c.startswith("2"):
            return "liability"
        if c.startswith("3"):
            return "equity"
        if c.startswith("4"):
            return "income"
        return "expense"  # 5xxx, 6xxx, 7xxx, etc.

    if _is_subsidiary_ledger(rows[:15]):
        try:
            from app.parsers.subsidiary_ledger import parse_subsidiary_ledger
            ledger = parse_subsidiary_ledger(path)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not parse subsidiary ledger: {exc}") from exc
        seen: set = set()
        tuples = []
        for acct in ledger.accounts:
            code = acct.code.strip()
            title = acct.title.strip()
            if not code or not title or code in seen:
                continue
            seen.add(code)
            tuples.append((code, title, _infer_account_type(code), 1))
        if not tuples:
            raise HTTPException(
                status_code=422,
                detail="No account codes found in the subsidiary ledger file",
            )
        return tuples

    header_idx = None
    mapping = {}
    for idx, row in enumerate(rows[:20]):
        headers = [_norm_header(c) for c in row]

        def find(names, _h=headers):
            for i, h in enumerate(_h):
                if h in names:
                    return i
            return None

        code_i = find({"code", "account code", "acct code", "account no", "account number", "gl code"})
        name_i = find({"name", "account name", "account title", "title", "description"})
        type_i = find({"type", "account type", "category", "normal type"})
        active_i = find({"active", "is active", "status"})
        if code_i is not None and name_i is not None:
            header_idx = idx
            mapping = {"code": code_i, "name": name_i, "type": type_i, "active": active_i}
            break
    if header_idx is None:
        raise HTTPException(
            status_code=422,
            detail=(
                "Could not detect required columns: account code and account title. "
                "Please ensure the file has headers for account code and account title, "
                "or upload a journal export file directly to extract account codes automatically."
            ),
        )
    parsed = []
    valid_types = {"asset", "liability", "equity", "income", "expense"}
    type_map = {
        "assets": "asset",
        "liabilities": "liability",
        "capital": "equity",
        "revenue": "income",
        "revenues": "income",
        "expenses": "expense",
    }
    for row_no, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):

        def val(key, _row=row, _m=mapping):
            i = _m.get(key)
            return _row[i] if i is not None and i < len(_row) else None

        code = str(val("code") or "").strip()
        name = str(val("name") or "").strip()
        if not code and not name:
            continue
        if not code or not name:
            raise HTTPException(
                status_code=422,
                detail=f"Row {row_no}: account code and title are required",
            )
        t = _norm_header(val("type")) or "expense"
        t = type_map.get(t, t)
        if t not in valid_types:
            raise HTTPException(
                status_code=422,
                detail=f"Row {row_no}: invalid account type {t!r}; use asset, liability, equity, income, or expense",
            )
        active_raw = _norm_header(val("active"))
        is_active = 0 if active_raw in {"inactive", "no", "false", "0", "disabled"} else 1
        parsed.append((code, name, t, is_active))
    if not parsed:
        raise HTTPException(status_code=422, detail="No account rows found after the header")
    return parsed


# ---------------------------------------------------------------------------
# Journal-basis helper
# ---------------------------------------------------------------------------


def _parse_flexible_journal_basis(path: Path):
    """Read ledger/journal files with varied headers and return a ledger-like object."""
    import csv as _csv

    from openpyxl import load_workbook as _load_workbook

    def norm(v):
        import re as _re

        return _re.sub(r"[^a-z0-9]+", " ", str(v or "").strip().lower()).strip()

    def dec(v):
        try:
            return Decimal(str(v or "0").replace(",", "").replace("₱", "").strip() or "0")
        except Exception:
            return Decimal("0")

    def rows_from_file():
        suf = path.suffix.lower()
        if suf == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                return [r for r in _csv.reader(f)]
        if suf in {".xlsx", ".xlsm"}:
            wb = _load_workbook(path, data_only=True, read_only=True)
            ws = max(wb.worksheets, key=lambda w: w.max_row * max(w.max_column, 1))
            return [[c for c in row] for row in ws.iter_rows(values_only=True)]
        if suf == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(path))
            sh = max(wb.sheets(), key=lambda x: x.nrows * x.ncols)
            return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return []

    rows = rows_from_file()
    aliases = {
        "code": {"account code", "account no", "acct code", "gl code", "code"},
        "title": {"account title", "account name", "account", "description", "particulars"},
        "debit": {"debit", "dr", "debit amount", "period debit"},
        "credit": {"credit", "cr", "credit amount", "period credit"},
    }
    mapping = None
    header_idx = 0
    for i, row in enumerate(rows[:50]):
        n = [norm(x) for x in row]
        found = {}
        for field, opts in aliases.items():
            for idx, h in enumerate(n):
                if h in opts or any(o in h for o in opts if len(o) >= 5):
                    found[field] = idx
                    break
        if {"code", "title", "debit", "credit"}.issubset(found):
            mapping = found
            header_idx = i
            break
    if mapping is None:
        raise ValueError("Could not detect account code, account title, debit, and credit columns.")
    accounts = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= max(mapping.values()):
            continue
        code = str(row[mapping["code"]] or "").strip()
        title = str(row[mapping["title"]] or "").strip()
        if not code or code.lower() in {"total", "grand total"}:
            continue
        debit = dec(row[mapping["debit"]])
        credit = dec(row[mapping["credit"]])
        if debit == 0 and credit == 0:
            continue
        acc = accounts.setdefault(
            code,
            SimpleNamespace(
                code=code,
                title=title or code,
                opening_balance=Decimal("0"),
                period_debit=Decimal("0"),
                period_credit=Decimal("0"),
                closing_balance=Decimal("0"),
                rows=0,
            ),
        )
        acc.title = title or acc.title
        acc.period_debit += debit
        acc.period_credit += credit
        acc.closing_balance += debit - credit
        acc.rows += 1
    if not accounts:
        raise ValueError("No account activity rows were found.")
    return SimpleNamespace(accounts=list(accounts.values()))


# ---------------------------------------------------------------------------
# Payments helpers
# ---------------------------------------------------------------------------

ALL_PAYMENT_STATUSES = (
    "Transaction Successful",
    "Transaction Released",
    "Scheduled",
    "Cancelled",
    "Transaction Failed",
    "Rejected",
)
ALL_REMITTANCE_TYPES = (
    "instaPay",
    "PESONet",
    "Internal Account Transfer",
    "Own Account Transfer",
    "SWIFT",
)


def _payments_query(conn, *, status="", remittance_type="", q="", limit=500):
    where = []
    params = []
    if status:
        where.append("transaction_status = ?")
        params.append(status)
    if remittance_type:
        where.append("remittance_type = ?")
        params.append(remittance_type)
    if q:
        where.append("(beneficiary_name LIKE ? OR remarks LIKE ? OR tran_id LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])
    sql = "SELECT * FROM payment_instructions"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY transaction_date DESC, id DESC LIMIT ?"
    params.append(limit)
    return conn.execute(sql, params).fetchall()


# ---------------------------------------------------------------------------
# Commission helpers
# ---------------------------------------------------------------------------


def _commission_history():
    """Return commission run history rows, empty list on error."""
    try:
        with db() as conn:
            return conn.execute(
                "SELECT * FROM commission_runs ORDER BY id DESC LIMIT 50"
            ).fetchall()
    except Exception:
        return []


def _commission_saved_carries():
    """Return saved carry-forward files list, newest first."""
    try:
        with db() as conn:
            return conn.execute(
                "SELECT * FROM commission_carry_files ORDER BY id DESC LIMIT 30"
            ).fetchall()
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Request form helpers
# ---------------------------------------------------------------------------


def _get_request_form_accounts(conn):
    """Return chart of accounts grouped by type for the request form dropdown."""
    rows = conn.execute(
        "SELECT id, code, name, type FROM chart_of_accounts WHERE is_active = 1 ORDER BY type, code"
    ).fetchall()
    groups: dict[str, list] = {}
    for r in rows:
        groups.setdefault(r["type"], []).append(r)
    ordered = {}
    for t in ("expense", "asset", "liability", "equity", "income"):
        if t in groups:
            ordered[t] = groups.pop(t)
    ordered.update(groups)
    return ordered


def _parse_expense_rows_from_journal(path: Path) -> list[dict]:
    """
    Parse an uploaded journal/ledger file and return a flat list of expense rows.

    Supports two layouts:
    A) Structured ledger:  account_code | account_title | debit | credit
       → Each non-zero debit row becomes one expense line.
    B) Flat expense list:  date | description/particulars | amount
       → Each row becomes one expense line (no GL code).

    Returns a list of dicts:
        {description, amount (str), account_code, account_title, confidence (float)}
    """
    import csv as _csv

    from openpyxl import load_workbook as _load_workbook

    def _dec(v):
        try:
            return Decimal(str(v or "0").replace(",", "").replace("₱", "").strip() or "0")
        except Exception:
            return Decimal("0")

    def _norm(v):
        import re as _re

        return _re.sub(r"[^a-z0-9 ]+", " ", str(v or "").strip().lower()).strip()

    def _read_raw():
        suf = path.suffix.lower()
        if suf == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as f:
                return list(_csv.reader(f))
        if suf in {".xlsx", ".xlsm"}:
            wb = _load_workbook(path, data_only=True, read_only=True)
            ws = max(wb.worksheets, key=lambda w: w.max_row * max(w.max_column, 1))
            return [[c for c in row] for row in ws.iter_rows(values_only=True)]
        if suf == ".xls":
            import xlrd

            wb = xlrd.open_workbook(str(path))
            sh = max(wb.sheets(), key=lambda x: x.nrows * x.ncols)
            return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        return []

    raw = _read_raw()
    if not raw:
        raise ValueError("File is empty or unreadable.")

    # --- Try to detect structured ledger layout ---
    code_aliases = {"account code", "acct code", "gl code", "code", "account no"}
    title_aliases = {"account title", "account name", "account", "description", "particulars"}
    debit_aliases = {"debit", "dr", "debit amount", "period debit", "amount"}
    credit_aliases = {"credit", "cr", "credit amount", "period credit"}

    mapping = None
    header_idx = 0
    for i, row in enumerate(raw[:30]):
        n = [_norm(x) for x in row]
        found = {}
        for field, opts in [
            ("code", code_aliases),
            ("title", title_aliases),
            ("debit", debit_aliases),
            ("credit", credit_aliases),
        ]:
            for j, h in enumerate(n):
                if h in opts or any(o in h for o in opts if len(o) >= 4):
                    found[field] = j
                    break
        if {"title", "debit"}.issubset(found):
            mapping = found
            header_idx = i
            break

    expense_rows = []

    if mapping:
        # Layout A: structured ledger
        for row in raw[header_idx + 1:]:
            if len(row) <= max(mapping.values()):
                continue
            code = str(row[mapping.get("code", -1)] if mapping.get("code") is not None and mapping.get("code", -1) < len(row) else "").strip()
            title = str(row[mapping["title"]] if mapping["title"] < len(row) else "").strip()
            debit = _dec(row[mapping["debit"]] if mapping["debit"] < len(row) else 0)
            if debit <= 0 or not title:
                continue
            if title.lower() in {"total", "grand total", "subtotal"}:
                continue
            # Match against chart of accounts in DB
            account_code = code
            account_title = title
            confidence = 0.9 if code else 0.6
            with db() as conn:
                match = conn.execute(
                    "SELECT id, code, name FROM chart_of_accounts WHERE is_active = 1 AND (code = ? OR lower(trim(name)) = lower(trim(?))) LIMIT 1",
                    (code, title),
                ).fetchone()
                if match:
                    account_code = match["code"]
                    account_title = match["name"]
                    confidence = 0.95
            expense_rows.append(
                {
                    "description": title,
                    "amount": str(debit),
                    "account_code": account_code,
                    "account_title": account_title,
                    "confidence": confidence,
                }
            )
    else:
        # Layout B: flat list with date | description | amount
        desc_col = amount_col = None
        for i, row in enumerate(raw[:20]):
            n = [_norm(x) for x in row]
            d_i = next((j for j, h in enumerate(n) if any(k in h for k in ("desc", "particular", "narration", "memo", "detail"))), None)
            a_i = next((j for j, h in enumerate(n) if any(k in h for k in ("amount", "debit", "dr"))), None)
            if d_i is not None and a_i is not None:
                desc_col = d_i
                amount_col = a_i
                header_idx = i
                break
        if desc_col is not None and amount_col is not None:
            for row in raw[header_idx + 1:]:
                if len(row) <= max(desc_col, amount_col):
                    continue
                desc = str(row[desc_col] or "").strip()
                amt = _dec(row[amount_col])
                if not desc or amt <= 0:
                    continue
                expense_rows.append(
                    {
                        "description": desc,
                        "amount": str(amt),
                        "account_code": "",
                        "account_title": "",
                        "confidence": 0.5,
                    }
                )

    return expense_rows
