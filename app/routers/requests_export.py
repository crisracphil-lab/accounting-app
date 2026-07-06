"""app/routers/requests_export.py — Excel export routes for payment requests."""
from __future__ import annotations

import calendar as _calendar
import io
from collections import defaultdict
from datetime import date
from datetime import timedelta as _timedelta

from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.db import db
from app.deps import (
    _current_user,
    _is_accounting_user,
    _is_operations_manager,
    _operations_access_company_ids,
)

router = APIRouter()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _safe_float(val) -> float:
    """Convert a DB amount value to float, tolerating comma-formatted strings."""
    if val is None:
        return 0.0
    return float(str(val).replace(",", "").strip() or 0)


def _co_id(conn, value) -> int:
    """Return a usable company_id integer, falling back to first active company."""
    cid = int(value) if value else 0
    if cid:
        return cid
    row = conn.execute("SELECT id FROM companies WHERE is_active=1 ORDER BY id LIMIT 1").fetchone()
    return row["id"] if row else 0


# ── Readability helpers for the monthly export (scoped to this file only —
#    none of this touches app/services/excel_styles.py, so the closing-books,
#    financial-statement, and open-items exports are unaffected) ─────────────

PESO_FMT = '"₱"#,##0.00'

_STATUS_STYLES = {
    "paid":           (PatternFill("solid", fgColor="C6EFCE"), Font(size=12, color="006100", name="Calibri", bold=True)),
    "approved":       (PatternFill("solid", fgColor="D9E8F5"), Font(size=12, color="1E3A5F", name="Calibri", bold=True)),
    "partially_paid": (PatternFill("solid", fgColor="D9E8F5"), Font(size=12, color="1E3A5F", name="Calibri", bold=True)),
    "rejected":       (PatternFill("solid", fgColor="FFC7CE"), Font(size=12, color="9C0006", name="Calibri", bold=True)),
    "cancelled":      (PatternFill("solid", fgColor="FFC7CE"), Font(size=12, color="9C0006", name="Calibri", bold=True)),
}
_STATUS_DEFAULT_STYLE = (PatternFill("solid", fgColor="FFEB9C"), Font(size=12, color="9C6500", name="Calibri", bold=True))

_THIN_SIDE_LOCAL = Side(style="thin", color="B0B8C5")
_TOTAL_BORDER_LOCAL = Border(
    top=Side(style="medium", color="1E3A5F"),
    bottom=Side(style="medium", color="1E3A5F"),
    left=_THIN_SIDE_LOCAL, right=_THIN_SIDE_LOCAL,
)

# Bumped up from the shared excel_styles.py defaults (9-10pt) to at least
# 12pt for readability, per request. Kept local to this file so the other
# reports that share excel_styles.py aren't affected.
DATA_FONT_SIZE = 12
HEADER_FONT_LOCAL = Font(bold=True, color="FFFFFF", size=DATA_FONT_SIZE, name="Calibri")


def _apply_status_style(cell, raw_status: str, center_alignment) -> None:
    fill, font = _STATUS_STYLES.get((raw_status or "").lower(), _STATUS_DEFAULT_STYLE)
    cell.fill = fill
    cell.font = font
    cell.alignment = center_alignment


@router.get("/requests/export.xlsx")
def requests_export_xlsx(request: Request, month: str = "", company_id: int = 0) -> Response:
    user = _current_user(request)
    if not (_is_accounting_user(user) or _is_operations_manager(user)):
        raise HTTPException(status_code=403, detail="Export access required")
    month = (month or date.today().strftime("%Y-%m")).strip()
    if len(month) != 7 or month[4] != "-":
        raise HTTPException(status_code=400, detail="Month must be YYYY-MM")
    start = f"{month}-01"
    year_i, month_i = map(int, month.split("-"))
    last_day = _calendar.monthrange(year_i, month_i)[1]
    end = f"{month}-{last_day:02d}"
    with db() as conn:
        selected_company_id = _co_id(conn, user["company_id"])
        if _is_operations_manager(user):
            allowed_company_ids = _operations_access_company_ids(conn, user)
            if company_id and company_id in allowed_company_ids:
                selected_company_id = company_id
            elif allowed_company_ids:
                selected_company_id = allowed_company_ids[0]
        ops_export_filter = (
            " AND pr.request_type = 'reimbursement' AND CAST(REPLACE(pr.amount, ',', '') AS REAL) > 50000"
            if _is_operations_manager(user) else ""
        )
        rows = conn.execute(
            "SELECT pr.id, c.name AS company, d.name AS department, pr.request_type,"
            "       u.full_name, u.username, pr.payee_name, pr.supplier_name, pr.description,"
            "       pr.amount, pr.due_date, pr.status, pr.created_at, pr.paid_at,"
            "       pr.accounting_notes, pr.operations_approved_at, ou.full_name AS operations_approved_by,"
            "       coa.code AS account_code, coa.name AS account_title"
            "  FROM payment_requests pr"
            "  JOIN users u ON u.id = pr.requester_user_id"
            "  LEFT JOIN departments d ON d.id = pr.department_id"
            "  LEFT JOIN companies c ON c.id = pr.company_id"
            "  LEFT JOIN users ou ON ou.id = pr.operations_approved_by_user_id"
            "  LEFT JOIN chart_of_accounts coa ON coa.id = pr.account_id"
            "  WHERE pr.company_id = ? AND date(pr.created_at) BETWEEN ? AND ?"
            + ops_export_filter +
            "  ORDER BY pr.created_at, pr.id",
            (selected_company_id, start, end),
        ).fetchall()

        # Fetch each request's individual line items (own account + amount) so
        # that requests with more than one account title can be reported
        # per-line-item, instead of squashing every account into a single
        # GROUP_CONCAT'd string (the cause of the export display bug).
        request_ids = [r["id"] for r in rows]
        line_items = []
        if request_ids:
            placeholders = ",".join("?" * len(request_ids))
            line_items = conn.execute(
                "SELECT li.request_id, li.description, li.amount,"
                "       coa2.code AS account_code, coa2.name AS account_title"
                "  FROM request_line_items li"
                "  LEFT JOIN chart_of_accounts coa2 ON coa2.id = li.account_id"
                "  WHERE li.request_id IN (" + placeholders + ")",
                tuple(request_ids),
            ).fetchall()
        line_items_by_request: dict = defaultdict(list)
        for li in line_items:
            line_items_by_request[li["request_id"]].append(li)

    from app.services.excel_styles import (
        CENTER,
        HDR_FILL,
        INT_FMT,
        RIGHT,
        add_corp_header,
        add_total_row,
        auto_col_width,
        finalize_workbook,
        style_data_rows,
        write_column_headers,
    )
    wb = Workbook()
    period = f"Period: {month}"

    ws = wb.active
    ws.title = "Requests"
    headers = [
        "ID", "Company", "Department", "Type", "Requester",
        "Payee", "Supplier", "Description",
        "Account Code", "Account Title",
        "Amount", "Due Date", "Status", "Created", "Paid At",
        "GM Approved At", "GM Approved By", "Notes",
    ]
    num_cols = len(headers)
    data_row = add_corp_header(ws, "Payment Requests", period, num_cols)
    write_column_headers(ws, data_row, headers)
    ws.row_dimensions[data_row].height = 20
    for c in range(1, num_cols + 1):
        ws.cell(data_row, c).font = HEADER_FONT_LOCAL
    data_row += 1
    # write_column_headers() sets freeze_panes via ws.cell(row+1, 1); merely
    # referencing a cell in openpyxl instantiates it, which silently bumps
    # ws.max_row and leaves a permanent blank row before the real data below
    # the header. Delete that phantom row, then set freeze_panes ourselves
    # with a plain string reference (no side effect) so the header rows AND
    # the first two columns (ID, Company) both stay frozen while scrolling.
    ws.delete_rows(data_row, 1)
    ws.freeze_panes = f"{get_column_letter(3)}{data_row}"

    total_amount = 0.0
    status_by_row: dict = {}
    row_cursor = data_row
    for r in rows:
        amt = _safe_float(r["amount"])
        total_amount += amt
        base = [
            r["id"], r["company"], r["department"],
            (r["request_type"] or "").replace("_", " ").title(),
            r["full_name"] or r["username"],
            r["payee_name"], r["supplier_name"], r["description"],
        ]
        tail = [
            r["due_date"],
            (r["status"] or "").replace("_", " ").title(),
            (r["created_at"] or "")[:16],
            (r["paid_at"] or "")[:10],
            (r["operations_approved_at"] or "")[:10],
            r["operations_approved_by"], r["accounting_notes"],
        ]
        req_li = line_items_by_request.get(r["id"], [])
        if req_li:
            # More than one account title: emit one row per line item, each
            # with its own account code/title and amount.
            for li in req_li:
                ws.append(base + [
                    li["account_code"] or "",
                    li["account_title"] or "Unclassified",
                    _safe_float(li["amount"]),
                ] + tail)
                status_by_row[row_cursor] = r["status"] or ""
                row_cursor += 1
        else:
            ws.append(base + [
                r["account_code"] or "",
                r["account_title"] or "Unclassified",
                amt,
            ] + tail)
            status_by_row[row_cursor] = r["status"] or ""
            row_cursor += 1

    last_data = ws.max_row
    style_data_rows(ws, data_row, last_data, num_cols, money_cols={11})
    add_total_row(ws, last_data + 1, num_cols, {11: total_amount})

    # Bump the data font up a size, give every row a touch more breathing
    # room, colour-code the Status column, and switch the Amount column to a
    # peso-prefixed format. Done after style_data_rows()/add_total_row() so
    # these overrides aren't wiped out by the shared helpers.
    for r in range(data_row, last_data + 2):
        ws.row_dimensions[r].height = 22
        for c in range(1, num_cols + 1):
            cell = ws.cell(r, c)
            cell.font = Font(bold=bool(cell.font and cell.font.bold), size=DATA_FONT_SIZE, name="Calibri")
    for r in range(data_row, last_data + 1):
        ws.cell(r, 11).number_format = PESO_FMT
        _apply_status_style(ws.cell(r, 13), status_by_row.get(r, ""), CENTER)
    ws.cell(last_data + 1, 11).number_format = PESO_FMT

    auto_col_width(ws)
    ws.column_dimensions["H"].width = 36
    ws.column_dimensions["J"].width = 32

    ws2 = wb.create_sheet("By Account")
    pivot: dict = defaultdict(lambda: {"code": "", "total": 0.0, "count": 0})
    seen_line_item_buckets: set = set()
    for r in rows:
        if (r["status"] or "").lower() in ("cancelled", "rejected"):
            continue
        req_li = line_items_by_request.get(r["id"], [])
        if req_li:
            # Split the request's amount across each of its actual accounts
            # instead of bucketing the full request amount under one fake,
            # concatenated "account" — this is what broke per-account totals
            # whenever a request had more than one line item.
            for li in req_li:
                key = li["account_title"] or "Unclassified"
                pivot[key]["code"] = li["account_code"] or ""
                pivot[key]["total"] += _safe_float(li["amount"])
                bucket_key = (r["id"], key)
                if bucket_key not in seen_line_item_buckets:
                    seen_line_item_buckets.add(bucket_key)
                    pivot[key]["count"] += 1
        else:
            key = r["account_title"] or "Unclassified"
            pivot[key]["code"] = r["account_code"] or ""
            pivot[key]["total"] += _safe_float(r["amount"])
            pivot[key]["count"] += 1
    pivot = dict(pivot)

    grand_total = sum(v["total"] for v in pivot.values())
    headers2 = ["Account Code", "Account Title", "No. of Requests", "Total Amount", "% of Total"]
    num_cols2 = len(headers2)
    hdr_row2 = add_corp_header(ws2, "Requests by Account Title", period, num_cols2)
    write_column_headers(ws2, hdr_row2, headers2)
    ws2.row_dimensions[hdr_row2].height = 20
    for c in range(1, num_cols2 + 1):
        ws2.cell(hdr_row2, c).font = HEADER_FONT_LOCAL
    # Same phantom blank-row cleanup as the Requests sheet above.
    ws2.delete_rows(hdr_row2 + 1, 1)
    ws2.freeze_panes = f"A{hdr_row2 + 1}"

    sorted_pivot = sorted(pivot.items(), key=lambda x: -x[1]["total"])
    for title, data in sorted_pivot:
        pct = (data["total"] / grand_total * 100) if grand_total else 0
        ws2.append([data["code"], title, data["count"], data["total"], pct / 100])

    last_data2 = ws2.max_row
    # Consistent borders/font/alignment/alternating fill for every data row —
    # previously this sheet only got alternating fill and nothing else, which
    # is what made it look unfinished next to the Requests sheet.
    style_data_rows(ws2, hdr_row2 + 1, last_data2, num_cols2, money_cols={4}, int_cols={3})
    for r in range(hdr_row2 + 1, last_data2 + 1):
        ws2.cell(r, 4).number_format = PESO_FMT
        ws2.cell(r, 5).number_format = "0.00%"
        for c in range(1, num_cols2 + 1):
            ws2.cell(r, c).font = Font(size=DATA_FONT_SIZE, name="Calibri")

    total_row2 = last_data2 + 1
    ws2.append(["", "TOTAL", sum(v["count"] for v in pivot.values()), grand_total, 1.0])
    for cell in ws2[total_row2]:
        cell.fill = HDR_FILL
        cell.font = Font(bold=True, size=DATA_FONT_SIZE, color="FFFFFF", name="Calibri")
        cell.border = _TOTAL_BORDER_LOCAL
    ws2.cell(total_row2, 3).alignment = RIGHT
    ws2.cell(total_row2, 3).number_format = INT_FMT
    ws2.cell(total_row2, 4).alignment = RIGHT
    ws2.cell(total_row2, 4).number_format = PESO_FMT
    ws2.cell(total_row2, 5).alignment = RIGHT
    ws2.cell(total_row2, 5).number_format = "0.00%"

    # A touch more row height throughout, matching the Requests sheet.
    for r in range(hdr_row2 + 1, total_row2 + 1):
        ws2.row_dimensions[r].height = 22

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 14

    # Move the account summary sheet to the front so it's the first
    # thing anyone sees when they open the export.
    wb.move_sheet("By Account", offset=-1)

    filename = f"bookpoint_requests_{month}.xlsx"
    return StreamingResponse(
        io.BytesIO(finalize_workbook(wb)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/requests/export-range.xlsx")
def requests_export_weekly(request: Request, start: str = "", end: str = "", company_id: int = 0) -> Response:
    """Export payment requests for a custom date range as a 3-sheet workbook."""
    user = _current_user(request)
    if not (_is_accounting_user(user) or _is_operations_manager(user)):
        raise HTTPException(status_code=403, detail="Export access required")

    today = date.today()
    if not start:
        start = (today - _timedelta(days=today.weekday())).isoformat()
    if not end:
        end = (today - _timedelta(days=today.weekday()) + _timedelta(days=6)).isoformat()

    monday_iso = start.strip()
    sunday_iso = end.strip()

    period_days = (date.fromisoformat(sunday_iso) - date.fromisoformat(monday_iso)).days + 1
    prev_monday = (date.fromisoformat(monday_iso) - _timedelta(days=period_days)).isoformat()
    prev_sunday = (date.fromisoformat(monday_iso) - _timedelta(days=1)).isoformat()

    with db() as conn:
        selected_company_id = _co_id(conn, user["company_id"])
        if _is_operations_manager(user):
            allowed_company_ids = _operations_access_company_ids(conn, user)
            if company_id and company_id in allowed_company_ids:
                selected_company_id = company_id
            elif allowed_company_ids:
                selected_company_id = allowed_company_ids[0]

        ops_filter = (
            " AND pr.request_type = 'reimbursement' AND CAST(REPLACE(pr.amount,',','') AS REAL) > 50000"
            if _is_operations_manager(user) else ""
        )

        def fetch_rows(s: str, e: str):
            return conn.execute(
                "SELECT pr.id, c.name AS company, d.name AS department, pr.request_type,"
                "       u.full_name, u.username, pr.payee_name, pr.supplier_name,"
                "       pr.description, pr.amount, pr.due_date, pr.status,"
                "       pr.created_at, pr.paid_at, pr.accounting_notes,"
                "       coa.code AS account_code, coa.name AS account_title"
                "  FROM payment_requests pr"
                "  JOIN users u ON u.id = pr.requester_user_id"
                "  LEFT JOIN departments d ON d.id = pr.department_id"
                "  LEFT JOIN companies c ON c.id = pr.company_id"
                "  LEFT JOIN chart_of_accounts coa ON coa.id = pr.account_id"
                "  WHERE pr.company_id = ?"
                "    AND date(pr.created_at) BETWEEN ? AND ?"
                "    AND pr.status NOT IN ('cancelled', 'rejected')"
                + ops_filter +
                "  ORDER BY pr.created_at, pr.id",
                (selected_company_id, s, e),
            ).fetchall()

        this_rows = fetch_rows(monday_iso, sunday_iso)
        prev_rows = fetch_rows(prev_monday, prev_sunday)

        def fetch_line_items(rows_list):
            ids = [r["id"] for r in rows_list]
            if not ids:
                return []
            ph = ",".join("?" * len(ids))
            return conn.execute(
                "SELECT li.request_id, li.description, li.amount,"
                "       coa.code AS account_code, coa.name AS account_title"
                "  FROM request_line_items li"
                "  LEFT JOIN chart_of_accounts coa ON coa.id = li.account_id"
                "  WHERE li.request_id IN (" + ph + ")",
                tuple(ids),
            ).fetchall()

        this_li = fetch_line_items(this_rows)
        prev_li = fetch_line_items(prev_rows)

        def build_pivot(rows_list, li_list):
            has_li = {r["request_id"] for r in li_list}
            pivot: dict = defaultdict(lambda: {"code": "", "total": 0.0, "count": 0})
            for li in li_list:
                key = li["account_title"] or "Unclassified"
                pivot[key]["code"] = li["account_code"] or ""
                pivot[key]["total"] += _safe_float(li["amount"])
            seen: set = set()
            for li in li_list:
                key = li["account_title"] or "Unclassified"
                k2 = (li["request_id"], key)
                if k2 not in seen:
                    seen.add(k2)
                    pivot[key]["count"] += 1
            for r in rows_list:
                if r["id"] in has_li:
                    continue
                key = r["account_title"] or "Unclassified"
                pivot[key]["code"] = r["account_code"] or ""
                pivot[key]["total"] += _safe_float(r["amount"])
                pivot[key]["count"] += 1
            return dict(pivot)


    from app.services.excel_styles import (
        ALT_FILL,
        BOLD_FONT,
        HDR_FILL,
        MONEY_FMT,
        add_corp_header,
        auto_col_width,
        finalize_workbook,
        write_column_headers,
    )

    wb = Workbook()
    period_label = f"{monday_iso} to {sunday_iso}"

    ws1 = wb.active
    ws1.title = "Requests"
    headers1 = ["ID", "Date", "Company", "Department", "Type", "Requester",
                 "Payee / Supplier", "Overall Description",
                 "Line Item Description", "Account Code", "Account Title",
                 "Amount", "Due Date", "Status", "Notes"]
    num_cols1 = len(headers1)
    hdr_row1 = add_corp_header(ws1, "Payment Requests — Weekly", period_label, num_cols1)
    write_column_headers(ws1, hdr_row1, headers1)

    li_by_req: dict = defaultdict(list)
    for li in this_li:
        li_by_req[li["request_id"]].append(li)

    excel_row = hdr_row1 + 1
    for r in this_rows:
        req_li = li_by_req.get(r["id"], [])
        base = [
            r["id"],
            (r["created_at"] or "")[:10],
            r["company"], r["department"],
            (r["request_type"] or "").replace("_", " ").title(),
            r["full_name"] or r["username"],
            r["payee_name"],
            r["description"],
        ]
        status_label = (r["status"] or "").replace("_", " ").title()
        fill = ALT_FILL if excel_row % 2 == 0 else None
        if req_li:
            for li in req_li:
                ws1.append(base + [
                    li["description"] or "",
                    li["account_code"] or "",
                    li["account_title"] or "Unclassified",
                    _safe_float(li["amount"]),
                    r["due_date"] or "",
                    status_label,
                    r["accounting_notes"] or "",
                ])
                if fill:
                    for cell in ws1[excel_row]:
                        cell.fill = fill
                excel_row += 1
        else:
            ws1.append(base + [
                "",
                r["account_code"] or "",
                r["account_title"] or "Unclassified",
                _safe_float(r["amount"]),
                r["due_date"] or "",
                status_label,
                r["accounting_notes"] or "",
            ])
            if fill:
                for cell in ws1[excel_row]:
                    cell.fill = fill
            excel_row += 1

    for row in ws1.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            cell.number_format = MONEY_FMT
    auto_col_width(ws1)
    ws1.column_dimensions["H"].width = 40
    ws1.column_dimensions["I"].width = 35
    ws1.column_dimensions["K"].width = 30

    ws2 = wb.create_sheet("By Account")
    pivot = build_pivot(this_rows, this_li)
    grand_total = sum(v["total"] for v in pivot.values())
    headers2 = ["Account Code", "Account Title", "No. of Requests", "Total Amount", "% of Week Total"]
    num_cols2 = len(headers2)
    hdr_row2 = add_corp_header(ws2, "Requests by Account — Weekly", period_label, num_cols2)
    write_column_headers(ws2, hdr_row2, headers2)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18
    sorted_pivot = sorted(pivot.items(), key=lambda x: -x[1]["total"])
    for i, (title, data) in enumerate(sorted_pivot, hdr_row2 + 1):
        pct = (data["total"] / grand_total * 100) if grand_total else 0
        ws2.append([data["code"], title, data["count"], data["total"], pct / 100])
        if i % 2 == 0:
            for cell in ws2[i]:
                cell.fill = ALT_FILL
    total_row = ws2.max_row + 1
    ws2.append(["", "TOTAL", sum(v["count"] for v in pivot.values()), grand_total, 1.0])
    for cell in ws2[total_row]:
        cell.fill = HDR_FILL
        cell.font = BOLD_FONT
    for row in ws2.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = MONEY_FMT
    for row in ws2.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "0.0%"

    ws3 = wb.create_sheet("vs Last Week")
    prev_pivot = build_pivot(prev_rows, prev_li)
    all_titles = sorted(set(pivot.keys()) | set(prev_pivot.keys()))
    this_label = f"{monday_iso} – {sunday_iso}"
    prev_label = f"{prev_monday} – {prev_sunday}"
    headers3 = ["Account Code", "Account Title",
                 f"This Week Amount ({this_label})", "This Week Count",
                 f"Prior Week Amount ({prev_label})", "Prior Week Count",
                 "Change (Amount)", "Change %"]
    num_cols3 = len(headers3)
    hdr_row3 = add_corp_header(ws3, "Week-over-Week Comparison", period_label, num_cols3)
    write_column_headers(ws3, hdr_row3, headers3)
    ws3.row_dimensions[hdr_row3].height = 32
    for i, title in enumerate(all_titles, hdr_row3 + 1):
        this_data = pivot.get(title, {"code": prev_pivot.get(title, {}).get("code", ""), "total": 0.0, "count": 0})
        prev_data = prev_pivot.get(title, {"code": "", "total": 0.0, "count": 0})
        code = this_data.get("code") or prev_data.get("code") or ""
        this_amt = this_data["total"]
        prev_amt = prev_data["total"]
        change_amt = this_amt - prev_amt
        change_pct = ((this_amt - prev_amt) / prev_amt) if prev_amt else (1.0 if this_amt else 0.0)
        ws3.append([code, title, this_amt, this_data["count"], prev_amt, prev_data["count"], change_amt, change_pct])
        if i % 2 == 0:
            for cell in ws3[i]:
                cell.fill = ALT_FILL
        change_cell = ws3.cell(row=i, column=7)
        pct_cell = ws3.cell(row=i, column=8)
        if change_amt > 0:
            change_cell.font = Font(color="C00000")
            pct_cell.font   = Font(color="C00000")
        elif change_amt < 0:
            change_cell.font = Font(color="006100")
            pct_cell.font   = Font(color="006100")

    gr = ws3.max_row + 1
    this_grand = sum(v["total"] for v in pivot.values())
    prev_grand = sum(v["total"] for v in prev_pivot.values())
    change_grand = this_grand - prev_grand
    change_grand_pct = ((this_grand - prev_grand) / prev_grand) if prev_grand else (1.0 if this_grand else 0.0)
    ws3.append(["", "TOTAL",
                 this_grand, sum(v["count"] for v in pivot.values()),
                 prev_grand, sum(v["count"] for v in prev_pivot.values()),
                 change_grand, change_grand_pct])
    for cell in ws3[gr]:
        cell.font = BOLD_FONT
        cell.fill = HDR_FILL

    for col in (3, 5, 7):
        for row in ws3.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.number_format = MONEY_FMT
    for row in ws3.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = '+0.0%;-0.0%;0.0%'
    auto_col_width(ws3)
    ws3.column_dimensions["B"].width = 35
    ws3.row_dimensions[1].height = 36

    # Move the account summary sheet to the front so it's the first
    # thing anyone sees when they open the export.
    wb.move_sheet("By Account", offset=-1)

    filename = f"bookpoint_requests_{monday_iso}_to_{sunday_iso}.xlsx"
    return StreamingResponse(
        io.BytesIO(finalize_workbook(wb)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
