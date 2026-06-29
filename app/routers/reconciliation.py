"""app/routers/reconciliation.py — GGR reconciliation, commission, bank-reconcile, open-items, ops."""
from __future__ import annotations

import contextlib
import io
import logging
import shutil
import tempfile
import uuid as _uuid
from collections import defaultdict
from datetime import date as _date
from datetime import timedelta as _timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse

from app.db import db, log_action
from app.deps import (
    COMMISSION_CARRY_DIR,
    _current_user,
    _is_accounting_user,
    _is_operations_manager,
    _operations_access_companies,
    _parse_iso,
    templates,
)
from app.parsers.bank_statement_generic import GenericStatementParseError
from app.parsers.combined_ggr_xls import CombinedParseError, parse_combined_xls, parse_two_file_ggr
from app.parsers.commission import (
    CommissionParseError,
    export_commission_result,
    parse_commission_report,
)
from app.parsers.ggr_excel import GGRParseError, parse_ggr_excel
from app.parsers.ggr_summary import (
    GGRSummaryParseError,
    export_ggr_weekly_summary,
    parse_ggr_weekly_summary,
)
from app.parsers.system_ledger import SystemLedgerParseError, parse_system_ledger
from app.services.bank_reconciliation import ReconciliationError, reconcile_files
from app.services.bank_system_reconciliation import reconcile_bank_system
from app.services.detailed_reconciliation import reconcile_detailed
from app.services.open_items import OpenItemsError, create_open_items_run, export_open_items_xlsx
from app.services.reconciliation import ggr_template, reconcile

logger = logging.getLogger(__name__)
router = APIRouter()

# In-process result caches (transient per-session; not persisted across restarts)
_commission_cache: dict = {}   # token -> CommissionResult
_ggr_summary_cache: dict = {}  # token -> GGRSummary
_detailed_cache: dict = {}     # token -> DetailedResult
_bank_system_cache: dict = {}  # token -> BankSystemResult


# ---------- GGR Reconciliation -----------------------------------------------

@router.get("/reconcile", response_class=HTMLResponse)
def reconcile_form(request: Request) -> Response:
    return templates.TemplateResponse(request, "reconcile_upload.html",
                                       {"error": None,
                                        "default_start": "2026-04-01",
                                        "default_end": "2026-04-27",
                                        "default_account": "4111"})


@router.post("/reconcile/ggr", response_class=HTMLResponse)
async def reconcile_ggr(request: Request,
                        excel_file: UploadFile = File(...),
                        system_file: UploadFile = File(...),
                        period_start: str = Form(""),
                        period_end: str = Form(""),
                        system_account_code: str = Form("4111")) -> Response:
    if not excel_file.filename.lower().endswith((".xlsx", ".xlsm")):
        return templates.TemplateResponse(request, "reconcile_upload.html",
                                           {"error": "Excel file must be .xlsx",
                                            "default_start": period_start,
                                            "default_end": period_end,
                                            "default_account": system_account_code})
    if not system_file.filename.lower().endswith(".xls"):
        return templates.TemplateResponse(request, "reconcile_upload.html",
                                           {"error": "System file must be .xls",
                                            "default_start": period_start,
                                            "default_end": period_end,
                                            "default_account": system_account_code})

    tmp_dir = Path(tempfile.mkdtemp())
    excel_path = tmp_dir / Path(excel_file.filename).name
    system_path = tmp_dir / Path(system_file.filename).name
    try:
        with excel_path.open("wb") as out:
            shutil.copyfileobj(excel_file.file, out)
        with system_path.open("wb") as out:
            shutil.copyfileobj(system_file.file, out)

        ps = _parse_iso(period_start)
        pe = _parse_iso(period_end)

        try:
            ex = parse_ggr_excel(excel_path, period_start=ps, period_end=pe)
        except GGRParseError as exc:
            return templates.TemplateResponse(request, "reconcile_upload.html",
                                               {"error": f"Excel parse error: {exc}",
                                                "default_start": period_start,
                                                "default_end": period_end,
                                                "default_account": system_account_code})
        try:
            sysl = parse_system_ledger(system_path, period_start=ps, period_end=pe)
        except SystemLedgerParseError as exc:
            return templates.TemplateResponse(request, "reconcile_upload.html",
                                               {"error": f"System parse error: {exc}",
                                                "default_start": period_start,
                                                "default_end": period_end,
                                                "default_account": system_account_code})

        period_label = f"{period_start or 'beginning'} to {period_end or 'end'}"
        result = reconcile(ex, sysl,
                           template_name=f"GGR Reconciliation ({period_label})",
                           rules=ggr_template(system_account_code.strip()),
                           excel_filename=excel_file.filename,
                           system_filename=system_file.filename)

        # Audit log
        with db() as conn:
            log_action(conn, "reconcile", "ggr", None, {
                "excel": excel_file.filename,
                "system": system_file.filename,
                "period_start": period_start,
                "period_end": period_end,
                "system_account_code": system_account_code,
                "all_within_tolerance": result.all_within_tolerance,
            })

        return templates.TemplateResponse(request, "reconcile_result.html",
                                           {"result": result})
    finally:
        for p in (excel_path, system_path):
            with contextlib.suppress(OSError):
                p.unlink(missing_ok=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ---------- GGR Weekly Summary -----------------------------------------------

@router.get("/ggr-summary", response_class=HTMLResponse)
def ggr_summary_form(request: Request) -> Response:
    return templates.TemplateResponse(request, "ggr_summary_upload.html", {"error": None})


@router.post("/ggr-summary", response_class=HTMLResponse)
async def ggr_summary_post(request: Request, ggr_file: UploadFile = File(...)) -> Response:
    allowed = (".xlsx", ".xlsm", ".xls")
    if not ggr_file.filename.lower().endswith(allowed):
        return templates.TemplateResponse(request, "ggr_summary_upload.html", {
            "error": "Please upload the monthly GGR workbook as .xlsx, .xlsm, or .xls."
        })

    tmp_dir = Path(tempfile.mkdtemp())
    tmp_path = tmp_dir / Path(ggr_file.filename).name
    try:
        with tmp_path.open("wb") as out:
            shutil.copyfileobj(ggr_file.file, out)
        try:
            summary = parse_ggr_weekly_summary(tmp_path)
        except GGRSummaryParseError as exc:
            return templates.TemplateResponse(request, "ggr_summary_upload.html", {
                "error": f"GGR summary parse error: {exc}"
            })
        except Exception as exc:
            return templates.TemplateResponse(request, "ggr_summary_upload.html", {
                "error": f"GGR summary file could not be processed: {exc}"
            })

        with db() as conn:
            log_action(conn, "ggr_summary", "ggr_report", None, {
                "filename": ggr_file.filename,
                "period_label": summary.period_label,
                "weeks": len(summary.weeks),
                "total_ggr": str(summary.total.ggr),
            })

        token = _uuid.uuid4().hex
        _ggr_summary_cache[token] = summary
        while len(_ggr_summary_cache) > 5:
            _ggr_summary_cache.pop(next(iter(_ggr_summary_cache)))
        return templates.TemplateResponse(request, "ggr_summary_result.html",
                                           {"summary": summary, "download_token": token})
    finally:
        with contextlib.suppress(OSError):
            tmp_path.unlink(missing_ok=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.get("/ggr-summary/export/{token}")
def ggr_summary_export(token: str) -> Response:
    summary = _ggr_summary_cache.get(token)
    if summary is None:
        raise HTTPException(404, "Result expired or not found - re-run the GGR weekly summary")
    data = export_ggr_weekly_summary(summary)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="GGR_Weekly_Summary_{token[:8]}.xlsx"'},
    )


# ---------- Commission --------------------------------------------------------

@router.get("/commission", response_class=HTMLResponse)
def commission_form(request: Request) -> Response:
    return templates.TemplateResponse(
        request, "commission_upload.html",
        {"error": None, "history": _commission_history(), "saved_carries": _commission_saved_carries()},
    )


@router.post("/commission", response_class=HTMLResponse)
async def commission_post(request: Request,
                          raw_file: UploadFile = File(...),
                          previous_file: UploadFile = File(None),
                          previous_carry_id: int = Form(None)) -> Response:
    allowed_raw = (".csv", ".xlsx", ".xlsm", ".xls")
    allowed_prev = (".xlsx", ".xlsm")

    def _err(msg):
        return templates.TemplateResponse(request, "commission_upload.html", {
            "error": msg, "history": _commission_history(), "saved_carries": _commission_saved_carries()
        })

    # Guard against missing or empty filename
    raw_filename = (raw_file.filename or "").strip()
    prev_filename = (previous_file.filename or "").strip() if previous_file else ""

    if not raw_filename:
        return _err("Please select a file to upload.")
    if not raw_filename.lower().endswith(allowed_raw):
        return _err("Please upload the raw Sub Affiliate report as .csv, .xlsx, .xlsm, or .xls.")
    if prev_filename and not prev_filename.lower().endswith(allowed_prev):
        return _err("Previous commission carry file must be .xlsx or .xlsm.")

    tmp_dir = Path(tempfile.mkdtemp())
    raw_path = tmp_dir / Path(raw_filename).name
    previous_path = None
    try:
        with raw_path.open("wb") as out:
            shutil.copyfileobj(raw_file.file, out)

        # Option A: user uploaded a new previous file
        if prev_filename:
            previous_path = tmp_dir / Path(prev_filename).name
            with previous_path.open("wb") as out:
                shutil.copyfileobj(previous_file.file, out)
        # Option B: user selected a saved carry-forward from the program
        elif previous_carry_id:
            try:
                with db() as conn:
                    carry_row = conn.execute(
                        "SELECT stored_filename, label FROM commission_carry_files WHERE id = ?",
                        (previous_carry_id,)
                    ).fetchone()
                if carry_row:
                    stored_path = COMMISSION_CARRY_DIR / carry_row["stored_filename"]
                    if stored_path.exists():
                        # Copy to temp dir so the parser can work with it safely
                        previous_path = tmp_dir / carry_row["stored_filename"]
                        shutil.copy2(stored_path, previous_path)
                        prev_filename = f"[saved] {carry_row['label']}"
            except Exception:
                logger.warning("Carry load failure — proceeding without previous carry file", exc_info=True)
        from datetime import date as _today_date
        _today = _today_date.today()
        _apply_carry = (_today.year > 2026) or (_today.year == 2026 and _today.month >= 5)
        try:
            result = parse_commission_report(
                raw_path, previous_path,
                sup_mgr_total_basis=_apply_carry,
            )
        except CommissionParseError as exc:
            return _err(str(exc))
        except Exception as exc:
            return _err(f"Commission upload could not be processed: {exc}")

        # Merge DSP tags from DB for rows that have no tag from the CSV.
        # Also upsert any new sub-affiliates seen in this run (blank tag) so
        # the user can assign them from /commission/dsp without re-uploading.
        try:
            with db() as conn:
                dsp_map = {
                    r["sub_id"].lower(): r["dsp_tag"]
                    for r in conn.execute(
                        "SELECT sub_id, dsp_tag FROM sub_affiliate_dsp"
                    ).fetchall()
                }
                for row in result.rows:
                    if not row.dsp_tag:
                        row.dsp_tag = dsp_map.get(row.sub_affiliate.lower(), "")
                    # Register unseen sub-affiliates so they appear in the DSP manager
                    conn.execute(
                        """INSERT INTO sub_affiliate_dsp (sub_id, sub_name, dsp_tag)
                           VALUES (?, ?, ?)
                           ON CONFLICT(sub_id) DO UPDATE SET
                             sub_name = CASE WHEN excluded.sub_name != '' THEN excluded.sub_name
                                             ELSE sub_affiliate_dsp.sub_name END""",
                        (row.sub_affiliate, row.sub_affiliate_name, row.dsp_tag),
                    )
        except Exception:
            logger.warning("DSP sub-affiliate merge failed (non-fatal)", exc_info=True)

        token = _uuid.uuid4().hex
        _commission_cache[token] = result
        while len(_commission_cache) > 5:
            _commission_cache.pop(next(iter(_commission_cache)))

        try:
            with db() as conn:
                log_action(conn, "commission_compute", "sub_affiliate", None, {
                    "filename": raw_filename,
                    "previous_filename": prev_filename,
                    "rows": len(result.rows),
                    "total_payable": str(result.totals["total_payable"]),
                    "next_negative_carry": str(result.totals["next_negative_carry"]),
                })
                conn.execute(
                    """INSERT INTO commission_runs
                         (filename, previous_filename, rows_count, total_payable, next_negative_carry, created_by_user_id)
                       VALUES (?, ?, ?, ?, ?, ?)""",
                    (
                        raw_filename,
                        prev_filename,
                        len(result.rows),
                        str(result.totals["total_payable"]),
                        str(result.totals["next_negative_carry"]),
                        getattr(request.state, "user", {}).get("id") if getattr(request.state, "user", None) else None,
                    ),
                )
        except Exception:
            logger.warning("Commission run history logging failed (non-fatal)", exc_info=True)

        return templates.TemplateResponse(request, "commission_result.html", {
            "result": result,
            "download_token": token,
            "carry_saved": False,
            "carry_label": None,
            "carry_id": None,
        })
    finally:
        for p in (raw_path, previous_path):
            if p:
                with contextlib.suppress(OSError):
                    p.unlink(missing_ok=True)
        with contextlib.suppress(OSError):
            shutil.rmtree(tmp_dir, ignore_errors=True)


@router.get("/commission/export/{token}")
def commission_export(token: str) -> Response:
    result = _commission_cache.get(token)
    if result is None:
        raise HTTPException(404, "Result expired or not found - re-run the commission computation")
    data = export_commission_result(result)
    fname = f"Sub_Affiliate_Commission_{token[:8]}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/commission/save-carry/{token}", response_class=HTMLResponse)
async def commission_save_carry(request: Request, token: str,
                                label: str = Form(...)) -> Response:
    """Save the commission result Excel as a named carry-forward file for the next cut-off."""
    result = _commission_cache.get(token)
    if result is None:
        raise HTTPException(404, "Result expired — re-run the commission computation before saving.")

    label = label.strip()
    if not label:
        raise HTTPException(422, "A cut-off label is required (e.g. '2026 Jan 1-15').")

    # Generate the Excel bytes
    data = export_commission_result(result)

    # Store to disk
    safe_label = "".join(c if c.isalnum() or c in " .-_" else "_" for c in label)
    stored_filename = f"{safe_label}_{token[:8]}.xlsx"
    dest = COMMISSION_CARRY_DIR / stored_filename
    dest.write_bytes(data)

    # Record in DB
    carry_id = None
    try:
        with db() as conn:
            cur = conn.execute(
                """INSERT INTO commission_carry_files
                     (label, original_filename, stored_filename, total_payable, next_negative_carry, rows_count)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (
                    label,
                    getattr(result, "source_filename", stored_filename),
                    stored_filename,
                    str(result.totals["total_payable"]),
                    str(result.totals["next_negative_carry"]),
                    len(result.rows),
                ),
            )
            carry_id = cur.lastrowid
    except Exception:
        logger.exception("Failed to save commission carry-forward file")

    return templates.TemplateResponse(
        request, "commission_result.html",
        {"result": result, "download_token": token,
         "carry_saved": True, "carry_label": label, "carry_id": carry_id},
    )


@router.get("/commission/carry-file/{carry_id}")
def commission_carry_file(carry_id: int) -> Response:
    """Download a previously saved carry-forward file."""
    try:
        with db() as conn:
            row = conn.execute(
                "SELECT label, stored_filename FROM commission_carry_files WHERE id = ?",
                (carry_id,)
            ).fetchone()
    except Exception:
        raise HTTPException(500, "Database error")
    if row is None:
        raise HTTPException(404, "Carry-forward file not found.")
    dest = COMMISSION_CARRY_DIR / row["stored_filename"]
    if not dest.exists():
        raise HTTPException(404, "File has been removed from storage.")
    fname = f"CommissionCarry_{row['label'].replace(' ', '_')}.xlsx"
    return StreamingResponse(
        io.BytesIO(dest.read_bytes()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/commission/carry-file/{carry_id}/delete")
def commission_carry_delete(carry_id: int) -> Response:
    """Delete a saved carry-forward file from storage and DB."""
    try:
        with db() as conn:
            row = conn.execute(
                "SELECT stored_filename FROM commission_carry_files WHERE id = ?",
                (carry_id,)
            ).fetchone()
            if row is None:
                raise HTTPException(404, "Carry-forward file not found.")
            # Remove the file from disk (ignore if already gone)
            dest = COMMISSION_CARRY_DIR / row["stored_filename"]
            with contextlib.suppress(OSError):
                dest.unlink(missing_ok=True)
            conn.execute(
                "DELETE FROM commission_carry_files WHERE id = ?", (carry_id,)
            )
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(500, "Could not delete carry-forward file.")
    return RedirectResponse("/commission", status_code=303)


# ---------- DSP Assignments ---------------------------------------------------

@router.get("/commission/dsp", response_class=HTMLResponse)
def dsp_list(request: Request, message: str = "", error: str = "") -> Response:
    with db() as conn:
        rows = conn.execute(
            "SELECT * FROM sub_affiliate_dsp ORDER BY dsp_tag, sub_id"
        ).fetchall()
    return templates.TemplateResponse(request, "commission_dsp.html", {
        "rows": rows, "message": message or None, "error": error or None,
    })


@router.post("/commission/dsp/save", response_class=HTMLResponse)
async def dsp_save(request: Request,
                   sub_id: str = Form(...),
                   sub_name: str = Form(""),
                   dsp_tag: str = Form(...)) -> Response:
    sub_id  = sub_id.strip()
    dsp_tag = dsp_tag.strip()
    sub_name = sub_name.strip()
    if not sub_id or not dsp_tag:
        return RedirectResponse("/commission/dsp?error=Sub+Affiliate+ID+and+DSP+Tag+are+required",
                                status_code=303)
    with db() as conn:
        conn.execute(
            """INSERT INTO sub_affiliate_dsp (sub_id, sub_name, dsp_tag, updated_at)
               VALUES (?, ?, ?, datetime('now'))
               ON CONFLICT(sub_id) DO UPDATE SET
                 sub_name=excluded.sub_name,
                 dsp_tag=excluded.dsp_tag,
                 updated_at=datetime('now')""",
            (sub_id, sub_name, dsp_tag),
        )
    return RedirectResponse(f"/commission/dsp?message={sub_id}+saved", status_code=303)


@router.post("/commission/dsp/{row_id}/delete")
def dsp_delete(row_id: int) -> Response:
    with db() as conn:
        conn.execute("DELETE FROM sub_affiliate_dsp WHERE id = ?", (row_id,))
    return RedirectResponse("/commission/dsp?message=Deleted", status_code=303)


@router.post("/commission/dsp/bulk-import", response_class=HTMLResponse)
async def dsp_bulk_import(request: Request,
                          csv_text: str = Form(""),
                          csv_file: UploadFile = File(None)) -> Response:
    """Bulk-import DSP assignments.

    Accepts either pasted CSV text or an uploaded file.
    Expected format (one per line, comma-separated):
        sub_affiliate_id, dsp_tag
    or three columns:
        sub_affiliate_id, sub_affiliate_name, dsp_tag
    """

    lines: list[str] = []
    if csv_file and csv_file.filename:
        raw = await csv_file.read()
        lines = raw.decode("utf-8-sig", errors="replace").splitlines()
    elif csv_text.strip():
        lines = csv_text.strip().splitlines()

    if not lines:
        return RedirectResponse("/commission/dsp?error=No+data+provided", status_code=303)

    saved = 0
    skipped = 0
    with db() as conn:
        for line in lines:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) >= 3:
                sub_id, sub_name, dsp_tag = parts[0], parts[1], parts[2]
            elif len(parts) == 2:
                sub_id, sub_name, dsp_tag = parts[0], "", parts[1]
            else:
                skipped += 1
                continue
            if not sub_id or not dsp_tag:
                skipped += 1
                continue
            # Skip obvious header rows
            if sub_id.lower() in {"sub affiliate", "sub affiliate id", "id", "store", "affiliate"}:
                continue
            conn.execute(
                """INSERT INTO sub_affiliate_dsp (sub_id, sub_name, dsp_tag, updated_at)
                   VALUES (?, ?, ?, datetime('now'))
                   ON CONFLICT(sub_id) DO UPDATE SET
                     sub_name=excluded.sub_name,
                     dsp_tag=excluded.dsp_tag,
                     updated_at=datetime('now')""",
                (sub_id, sub_name, dsp_tag),
            )
            saved += 1

    msg = f"{saved}+entries+imported"
    if skipped:
        msg += f"+({skipped}+skipped)"
    return RedirectResponse(f"/commission/dsp?message={msg}", status_code=303)


# ---------- Detailed GGR Reconciliation --------------------------------------

@router.get("/reconcile/ggr/detailed", response_class=HTMLResponse)
def reconcile_detailed_form(request: Request) -> Response:
    return templates.TemplateResponse(request, "reconcile_detailed_upload.html", {"error": None})


@router.post("/reconcile/ggr/detailed", response_class=HTMLResponse)
async def reconcile_detailed_post(request: Request,
                                  raw_ggr_file: UploadFile = File(None),
                                  per_system_file: UploadFile = File(None),
                                  combined_file: UploadFile = File(None),
                                  account_code: str = Form("4111")) -> Response:
    allowed = (".xls", ".xlsx", ".xlsm", ".csv")
    tmp_dir = Path(tempfile.mkdtemp())
    saved_paths = []
    try:
        try:
            if raw_ggr_file and raw_ggr_file.filename and per_system_file and per_system_file.filename:
                for upload in (raw_ggr_file, per_system_file):
                    if not upload.filename.lower().endswith(allowed):
                        raise HTTPException(status_code=422, detail="Detailed GGR files must be .xls, .xlsx, .xlsm, or .csv")
                raw_path = tmp_dir / f"raw_{Path(raw_ggr_file.filename).name}"
                system_path = tmp_dir / f"system_{Path(per_system_file.filename).name}"
                with raw_path.open("wb") as out:
                    shutil.copyfileobj(raw_ggr_file.file, out)
                with system_path.open("wb") as out:
                    shutil.copyfileobj(per_system_file.file, out)
                saved_paths.extend([raw_path, system_path])
                cf = parse_two_file_ggr(raw_path, system_path)
                source_name = f"Raw: {raw_ggr_file.filename}; System: {per_system_file.filename}"
            elif combined_file and combined_file.filename:
                if not combined_file.filename.lower().endswith(".xls"):
                    raise HTTPException(status_code=422, detail="Legacy combined file must be .xls with PER SYSTEM and PER GGR sheets")
                combined_path = tmp_dir / Path(combined_file.filename).name
                with combined_path.open("wb") as out:
                    shutil.copyfileobj(combined_file.file, out)
                saved_paths.append(combined_path)
                cf = parse_combined_xls(combined_path)
                source_name = combined_file.filename
            else:
                return templates.TemplateResponse(request, "reconcile_detailed_upload.html", {
                    "error": "Upload both Raw GGR and Per System files, or upload one legacy combined .xls file."
                })
        except CombinedParseError as exc:
            raise HTTPException(status_code=422, detail=f"Detailed GGR reconciliation parse error: {exc}") from exc

        result = reconcile_detailed(cf, account_filter=account_code.strip() or None)

        # Cache for Excel export
        token = _uuid.uuid4().hex
        _detailed_cache[token] = result
        # Trim cache to last 5
        while len(_detailed_cache) > 5:
            _detailed_cache.pop(next(iter(_detailed_cache)))

        # Audit log
        with db() as conn:
            log_action(conn, "reconcile_detailed", "ggr", None, {
                "filename": source_name,
                "total": result.summary.total_records,
                "matched": result.summary.matched_records,
                "discrepancies": result.summary.discrepancy_count,
            })

        return templates.TemplateResponse(request, "reconcile_detailed_result.html",
                                           {"result": result, "download_token": token})
    finally:
        for p in saved_paths:
            with contextlib.suppress(OSError):
                p.unlink(missing_ok=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.get("/reconcile/ggr/detailed/export/{token}")
def reconcile_detailed_export(token: str) -> Response:
    result = _detailed_cache.get(token)
    if result is None:
        raise HTTPException(404, "Result expired or not found - re-run the reconciliation")

    from openpyxl import Workbook as _WB2
    from openpyxl.styles import Font as _FN2
    from openpyxl.styles import PatternFill as _PF2

    from app.services.excel_styles import (
        BOLD_FONT,
        DATA_FONT,
        HDR_FILL,
        NAVY_LIGHT,
        add_corp_header,
        auto_col_width,
        finalize_workbook,
        style_data_rows,
        write_column_headers,
    )

    period = result.summary.period_label or ""
    wb = _WB2()

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    add_corp_header(ws, "GGR Reconciliation Summary", period, 4)
    # Metrics block
    ws.cell(5, 1, "Metric").font = BOLD_FONT
    ws.cell(5, 2, "Value").font = BOLD_FONT
    ws.cell(5, 1).fill = HDR_FILL
    ws.cell(5, 2).fill = HDR_FILL
    ws.cell(5, 1).font = _FN2(bold=True, color="FFFFFF", size=9)
    ws.cell(5, 2).font = _FN2(bold=True, color="FFFFFF", size=9)
    for i, (label, val) in enumerate([
        ("Total Records Checked", result.summary.total_records),
        ("Matched Records",       result.summary.matched_records),
        ("Total Discrepancies",   result.summary.discrepancy_count),
    ], 6):
        ws.cell(i, 1, label).font = DATA_FONT
        ws.cell(i, 2, val).font = DATA_FONT
        if i % 2 == 0:
            ws.cell(i, 1).fill = _PF2("solid", fgColor="EBF3FB")
            ws.cell(i, 2).fill = _PF2("solid", fgColor="EBF3FB")

    r = 10
    ws.cell(r, 1, "DISCREPANCY BREAKDOWN BY WEEK").font = BOLD_FONT
    r += 1
    for hdr, col in [("Week", 1), ("Missing in System", 2), ("Wrong Side", 3), ("Amount Discrepancy", 4), ("In System Only", 5), ("Total", 6)]:
        ws.cell(r, col, hdr).font = _FN2(bold=True, color="FFFFFF")
        ws.cell(r, col).fill = _PF2("solid", fgColor=NAVY_LIGHT)
    r += 1
    for w in result.summary.by_week_total:
        miss    = result.summary.by_week_missing.get(w, 0)
        wrong   = result.summary.by_week_wrong.get(w, 0)
        amt     = result.summary.by_week_amount.get(w, 0)
        sysonly = result.summary.by_week_system_only.get(w, 0)
        ws.cell(r, 1, w)
        ws.cell(r, 2, miss)
        ws.cell(r, 3, wrong)
        ws.cell(r, 4, amt)
        ws.cell(r, 5, sysonly)
        ws.cell(r, 6, miss + wrong + amt + sysonly)
        r += 1

    if result.summary.findings:
        r += 1
        ws.cell(r, 1, "KEY FINDINGS").font = BOLD_FONT
        for f in result.summary.findings:
            r += 1
            ws.cell(r, 1, f).font = DATA_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    # ── Discrepancies + Full Reconciliation sheets ────────────────────────────
    det_headers = ["Week", "Outlet Code", "Outlet / Description", "Account Type",
                   "GGR Amount", "Expected Side", "System Debit", "System Credit",
                   "Actual Side", "Discrepancy Type", "Difference", "Nearest System Amt"]
    num_det = len(det_headers)
    for sheet_name, sheet_rows in (("Discrepancies", result.discrepancies),
                                    ("Full Reconciliation", result.rows)):
        ws = wb.create_sheet(sheet_name)
        hdr_r = add_corp_header(ws, sheet_name, period, num_det)
        write_column_headers(ws, hdr_r, det_headers)
        data_start = hdr_r + 1
        for ri, row in enumerate(sheet_rows, data_start):
            ws.cell(ri, 1, row.week_label)
            ws.cell(ri, 2, row.outlet_code)
            ws.cell(ri, 3, row.outlet_name)
            ws.cell(ri, 4, row.account_type)
            if row.ggr_amount:
                ws.cell(ri, 5, float(row.ggr_amount))
            ws.cell(ri, 6, row.expected_side)
            if row.system_debit is not None:
                ws.cell(ri, 7, float(row.system_debit))
            if row.system_credit is not None:
                ws.cell(ri, 8, float(row.system_credit))
            ws.cell(ri, 9, row.actual_side)
            ws.cell(ri, 10, row.discrepancy_type)
            if row.amount_difference is not None:
                ws.cell(ri, 11, float(row.amount_difference))
            if row.nearest_system_amount is not None:
                ws.cell(ri, 12, float(row.nearest_system_amount))
        last = ws.max_row
        style_data_rows(ws, data_start, last, num_det, money_cols={5, 7, 8, 11, 12})
        auto_col_width(ws)

    fname = f"GGR_Detailed_Reconciliation_{token[:8]}.xlsx"
    return StreamingResponse(
        io.BytesIO(finalize_workbook(wb)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Per Bank vs Per System Amount Reconciliation --------------------

@router.get("/reconcile/bank-system", response_class=HTMLResponse)
def bank_system_form(request: Request) -> Response:
    return templates.TemplateResponse(request, "bank_system_upload.html", {"error": None})


@router.post("/reconcile/bank-system", response_class=HTMLResponse)
async def bank_system_upload(
    request: Request,
    system_file: UploadFile = File(...),
    bank_file: UploadFile = File(...),
    tolerance: str = Form("0.05"),
    bank_balance: str = Form("0"),
    book_balance: str = Form("0"),
    company_name: str = Form(""),
    bank_account: str = Form(""),
    statement_date: str = Form(""),
) -> Response:
    import pathlib
    import shutil
    import tempfile
    from decimal import Decimal, InvalidOperation

    from app.parsers.bank_statement_generic import (
        GenericStatementParseError,
        parse_generic_statement,
    )

    def _to_dec(v, default="0"):
        try:
            return Decimal(str(v).strip().replace(",", "") or default)
        except InvalidOperation:
            return Decimal(default)

    tol          = _to_dec(tolerance, "0.05")
    bank_bal     = _to_dec(bank_balance)
    book_bal     = _to_dec(book_balance)

    tmp = pathlib.Path(tempfile.mkdtemp())
    try:
        sys_path  = tmp / pathlib.Path(system_file.filename).name
        bank_path = tmp / pathlib.Path(bank_file.filename).name
        sys_path.write_bytes(await system_file.read())
        bank_path.write_bytes(await bank_file.read())

        try:
            system_txns = parse_generic_statement(sys_path)
            bank_txns   = parse_generic_statement(bank_path)
        except GenericStatementParseError as exc:
            return templates.TemplateResponse(request, "bank_system_upload.html", {"error": str(exc)})

        result = reconcile_bank_system(
            system_txns, bank_txns,
            system_filename=system_file.filename,
            bank_filename=bank_file.filename,
            tolerance=tol,
            bank_balance=bank_bal,
            book_balance=book_bal,
            company_name=company_name.strip(),
            bank_account=bank_account.strip(),
            statement_date=statement_date.strip(),
        )

        token = _uuid.uuid4().hex
        _bank_system_cache[token] = result
        while len(_bank_system_cache) > 10:
            _bank_system_cache.pop(next(iter(_bank_system_cache)))

        return templates.TemplateResponse(request, "bank_system_result.html", {
            "result": result,
            "download_token": token,
        })
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@router.get("/reconcile/bank-system/export/{token}")
def bank_system_export(token: str) -> Response:
    result = _bank_system_cache.get(token)
    if result is None:
        raise HTTPException(404, "Result expired — re-run the reconciliation")

    from openpyxl import Workbook as _WB3

    from app.services.excel_styles import (
        add_corp_header,
        auto_col_width,
        finalize_workbook,
        style_data_rows,
        write_column_headers,
    )

    wb = _WB3()
    wb.remove(wb.active)

    headers = ["Status", "System Date", "System Ref", "System Description",
               "System Debit", "System Credit",
               "Bank Date", "Bank Ref", "Bank Description",
               "Bank Debit", "Bank Credit", "Difference"]
    ncols = len(headers)

    period = f"{result.system_filename} vs {result.bank_filename}"
    for sheet_name, rows in [
        ("System to Bank", result.sys_to_bank),
        ("Bank to System", result.bank_to_sys),
        ("All Rows",       result.rows),
    ]:
        ws = wb.create_sheet(sheet_name)
        hdr_r = add_corp_header(ws, sheet_name, period, ncols)
        write_column_headers(ws, hdr_r, headers)
        dr = hdr_r + 1
        for r in rows:
            ws.cell(dr, 1,  r.status)
            ws.cell(dr, 2,  r.sys_date or "")
            ws.cell(dr, 3,  r.sys_ref or "")
            ws.cell(dr, 4,  r.sys_description)
            if r.sys_debit is not None:
                ws.cell(dr, 5, float(r.sys_debit))
            if r.sys_credit is not None:
                ws.cell(dr, 6, float(r.sys_credit))
            ws.cell(dr, 7,  r.bank_date or "")
            ws.cell(dr, 8,  r.bank_ref or "")
            ws.cell(dr, 9,  r.bank_description)
            if r.bank_debit is not None:
                ws.cell(dr, 10, float(r.bank_debit))
            if r.bank_credit is not None:
                ws.cell(dr, 11, float(r.bank_credit))
            if r.difference is not None:
                ws.cell(dr, 12, float(r.difference))
            dr += 1
        style_data_rows(ws, hdr_r + 1, ws.max_row, ncols, money_cols={5, 6, 10, 11, 12})
        auto_col_width(ws)

    fname = f"BankSystem_Reconciliation_{token[:8]}.xlsx"
    return StreamingResponse(
        io.BytesIO(finalize_workbook(wb)),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Bank Reconciliation Workspace ------------------------------------

@router.get("/reconciliation-workspace", response_class=HTMLResponse)
def reconciliation_workspace_page(request: Request) -> Response:
    with db() as conn:
        runs = conn.execute(
            """SELECT r.*,
                      (SELECT COUNT(*) FROM bank_reconciliation_items i WHERE i.run_id = r.id) AS item_count,
                      (SELECT COUNT(*) FROM bank_reconciliation_items i WHERE i.run_id = r.id AND i.status = 'matched') AS matched_count
               FROM bank_reconciliation_runs r ORDER BY r.id DESC LIMIT 30"""
        ).fetchall()
    return templates.TemplateResponse(request, "bank_reconcile_upload.html", {"runs": runs})


@router.get("/bank-reconcile", response_class=HTMLResponse)
def bank_reconcile_legacy_page(request: Request) -> Response:
    return reconciliation_workspace_page(request)


@router.post("/reconciliation-workspace", response_class=HTMLResponse)
async def reconciliation_workspace_upload(request: Request,
                                          bank_file: UploadFile = File(...),
                                          system_file: UploadFile = File(...),
                                          reconciliation_type: str = Form("general_file_reconciliation")) -> Response:
    allowed = (".xlsx", ".xlsm", ".xls", ".csv")
    if reconciliation_type not in {"general_file_reconciliation", "bank_vs_system", "system_vs_raw", "bank_vs_raw", "left_vs_right"}:
        raise HTTPException(status_code=422, detail="Invalid reconciliation type.")
    if not bank_file.filename.lower().endswith(allowed) or not system_file.filename.lower().endswith(allowed):
        raise HTTPException(status_code=422, detail="Both reconciliation files must be .xlsx, .xlsm, .xls, or .csv.")
    tmpdir = Path(tempfile.mkdtemp())
    bank_path = tmpdir / Path(bank_file.filename).name
    system_path = tmpdir / Path(system_file.filename).name
    try:
        with bank_path.open("wb") as out:
            shutil.copyfileobj(bank_file.file, out)
        with system_path.open("wb") as out:
            shutil.copyfileobj(system_file.file, out)
        try:
            run_id = reconcile_files(bank_path, system_path, reconciliation_type=reconciliation_type)
        except (GenericStatementParseError, ReconciliationError) as exc:
            raise HTTPException(status_code=422, detail=f"Reconciliation error: {exc}") from exc
    finally:
        for _p in (bank_path, system_path):
            with contextlib.suppress(OSError):
                _p.unlink(missing_ok=True)
        with contextlib.suppress(OSError):
            tmpdir.rmdir()
    return RedirectResponse(f"/reconciliation-workspace/{run_id}", status_code=303)


@router.post("/bank-reconcile", response_class=HTMLResponse)
async def bank_reconcile_legacy_upload(request: Request,
                                       bank_file: UploadFile = File(...),
                                       system_file: UploadFile = File(...),
                                       reconciliation_type: str = Form("general_file_reconciliation")) -> Response:
    return await reconciliation_workspace_upload(request, bank_file, system_file, reconciliation_type)


@router.get("/reconciliation-workspace/{run_id}", response_class=HTMLResponse)
def reconciliation_workspace_result(request: Request, run_id: int, status: str = "") -> Response:
    with db() as conn:
        run = conn.execute("SELECT * FROM bank_reconciliation_runs WHERE id = ?", (run_id,)).fetchone()
        if run is None:
            raise HTTPException(404, "Reconciliation run not found")
        # status filter: only a fixed literal SQL fragment is appended; the
        # status value itself always travels through a ? placeholder in params.
        status_sql = " AND status = ?" if status else ""
        params = (run_id, status) if status else (run_id,)
        items = conn.execute(
            "SELECT * FROM bank_reconciliation_items WHERE run_id = ?" + status_sql + " ORDER BY status, id",
            params,
        ).fetchall()
        counts = conn.execute(
            "SELECT status, COUNT(*) AS n FROM bank_reconciliation_items WHERE run_id = ? GROUP BY status", (run_id,)
        ).fetchall()
    return templates.TemplateResponse(request, "bank_reconcile_result.html",
                                       {"run": run, "items": items, "counts": counts, "status": status})


# ---------- Open Items -------------------------------------------------------

@router.get("/open-items", response_class=HTMLResponse)
def open_items_page(request: Request) -> Response:
    user = _current_user(request)
    is_admin = user["role"] == "admin"
    company_id = int(user["company_id"] or 1)
    with db() as conn:
        if is_admin:
            runs = conn.execute(
                """SELECT r.*,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'open') AS open_count,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'partial') AS partial_count,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'closed') AS closed_count
                   FROM open_item_runs r ORDER BY r.id DESC LIMIT 30"""
            ).fetchall()
        else:
            runs = conn.execute(
                """SELECT r.*,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'open') AS open_count,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'partial') AS partial_count,
                          (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'closed') AS closed_count
                   FROM open_item_runs r WHERE r.company_id = ? ORDER BY r.id DESC LIMIT 30""",
                (company_id,)
            ).fetchall()
    return templates.TemplateResponse(request, "open_items_upload.html", {"runs": runs})


@router.post("/open-items", response_class=HTMLResponse)
async def open_items_upload(request: Request,
                            ledger_file: UploadFile = File(...),
                            account_filter: str = Form(...),
                            open_side: str = Form("auto"),
                            tolerance: str = Form("0.05")) -> Response:
    user = _current_user(request)
    company_id = int(user["company_id"] or 1)
    allowed = (".xlsx", ".xlsm", ".xls", ".csv")
    if not ledger_file.filename.lower().endswith(allowed):
        raise HTTPException(status_code=422, detail="Open items file must be .xlsx, .xlsm, .xls, or .csv.")
    tmpdir = Path(tempfile.mkdtemp())
    path = tmpdir / Path(ledger_file.filename).name
    try:
        with path.open("wb") as out:
            shutil.copyfileobj(ledger_file.file, out)
        try:
            run_id = create_open_items_run(
                path,
                account_filter=account_filter.strip(),
                open_side=open_side,
                tolerance=Decimal(tolerance or "0.05"),
                company_id=company_id,
            )
        except (OpenItemsError, InvalidOperation) as exc:
            with db() as _conn:
                runs = _conn.execute(
                    """SELECT r.*,
                              (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'open') AS open_count,
                              (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'partial') AS partial_count,
                              (SELECT COUNT(*) FROM open_item_details d WHERE d.run_id = r.id AND d.status = 'closed') AS closed_count
                       FROM open_item_runs r WHERE r.company_id = ? ORDER BY r.id DESC LIMIT 30""",
                    (company_id,)
                ).fetchall()
            return templates.TemplateResponse(request, "open_items_upload.html", {
                "runs": runs, "error": str(exc),
            })
    finally:
        with contextlib.suppress(OSError):
            path.unlink(missing_ok=True)
        with contextlib.suppress(OSError):
            tmpdir.rmdir()
    return RedirectResponse(f"/open-items/{run_id}", status_code=303)


@router.get("/open-items/{run_id}", response_class=HTMLResponse)
def open_items_result(request: Request, run_id: int, status: str = "") -> Response:
    user = _current_user(request)
    with db() as conn:
        run = conn.execute("SELECT * FROM open_item_runs WHERE id = ?", (run_id,)).fetchone()
        if run is None:
            raise HTTPException(status_code=404, detail="Open items run not found")
        if user["role"] != "admin" and int(run["company_id"] or 1) != int(user["company_id"] or 1):
            raise HTTPException(status_code=404, detail="Open items run not found")
        # Build WHERE from a fixed base; only literal SQL fragments are
        # appended to where_sql; status travels through a ? placeholder.
        where_sql = "WHERE e.run_id = ?"
        params: list = [run_id]
        if status:
            where_sql += " AND e.status = ?"
            params.append(status)

        details = conn.execute(
            "SELECT e.status, e.open_date, e.account_code, e.account_title,"
            "       e.reference, e.description,"
            "       e.original_amount, e.closed_amount, e.open_balance, e.aging_days"
            "  FROM open_item_details e"
            "  " + where_sql +
            "  ORDER BY e.open_date, e.id",
            params,
        ).fetchall()

        closures = conn.execute("""
            SELECT open_date, open_reference, close_date, close_reference,
                   amount, close_description
              FROM open_item_closures
             WHERE run_id = ?
             ORDER BY close_date
        """, (run_id,)).fetchall()

        counts = conn.execute("""
            SELECT status, COUNT(*) AS n
              FROM open_item_details
             WHERE run_id = ?
             GROUP BY status
             ORDER BY status
        """, (run_id,)).fetchall()

    return templates.TemplateResponse(request, "open_items_result.html", {
        "run": run,
        "details": details,
        "closures": closures,
        "counts": counts,
        "status": status,
    })


@router.get("/open-items/{run_id}/export")
def open_items_export(request: Request, run_id: int) -> Response:
    user = _current_user(request)
    with db() as conn:
        run = conn.execute("SELECT company_id FROM open_item_runs WHERE id = ?", (run_id,)).fetchone()
        if run is None:
            raise HTTPException(status_code=404, detail="Open items run not found")
        if user["role"] != "admin" and int(run["company_id"] or 1) != int(user["company_id"] or 1):
            raise HTTPException(status_code=404, detail="Open items run not found")
    data = export_open_items_xlsx(run_id)
    buf = io.BytesIO(data)
    fname = f"open_items_{run_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------- Operations Manager Reconciliation --------------------------------

@router.get("/ops-reconciliation", response_class=HTMLResponse)
def ops_reconciliation(request: Request, company_id: int | None = None, months: int = 3) -> Response:
    """Show ops-approval requests vs. bank debits, plus duplicate detection."""
    user = _current_user(request)
    if not (_is_operations_manager(user) or _is_accounting_user(user)):
        raise HTTPException(status_code=403, detail="Operations managers only")

    with db() as conn:
        companies = _operations_access_companies(conn, user)
        if not companies:
            # Fallback: accounting user sees their own company
            if user["company_id"]:
                companies = conn.execute(
                    "SELECT * FROM companies WHERE id = ? AND is_active = 1",
                    (int(user["company_id"]),)
                ).fetchall()
            if not companies:
                companies = conn.execute(
                    "SELECT * FROM companies WHERE is_active = 1 ORDER BY id LIMIT 1"
                ).fetchall()

        company_ids = [int(c["id"]) for c in companies]
        if company_id and int(company_id) in company_ids:
            selected_cid = int(company_id)
        else:
            selected_cid = company_ids[0] if company_ids else 0

        # All reimbursements > 50 000 for this company (any non-cancelled status)
        cutoff = (_date.today() - _timedelta(days=months * 30)).isoformat()
        reqs = conn.execute("""
            SELECT pr.id, pr.created_at, pr.payee_name, pr.amount, pr.status,
                   u.full_name  AS requester_name,
                   u.username   AS requester_username,
                   d.name       AS department_name
              FROM payment_requests pr
              LEFT JOIN users       u ON u.id = pr.requester_user_id
              LEFT JOIN departments d ON d.id = pr.department_id
             WHERE pr.company_id = ?
               AND pr.request_type = 'reimbursement'
               AND pr.status != 'cancelled'
               AND CAST(REPLACE(pr.amount, ',', '') AS REAL) > 50000
               AND pr.created_at >= ?
             ORDER BY pr.created_at DESC
        """, (selected_cid, cutoff)).fetchall()

        # Bank debits for this company — window is 2× the month filter
        bank_cutoff = (_date.today() - _timedelta(days=months * 60)).isoformat()
        bank_debits = conn.execute("""
            SELECT bt.id, bt.transaction_date, bt.description,
                   bt.counterparty_name, bt.reference_number, bt.check_number,
                   CAST(REPLACE(bt.debit_amount, ',', '') AS REAL) AS debit_amount
              FROM bank_transactions bt
              JOIN uploaded_files uf ON uf.id = bt.uploaded_file_id
             WHERE uf.company_id = ?
               AND bt.transaction_date >= ?
               AND CAST(REPLACE(bt.debit_amount, ',', '') AS REAL) > 50000
             ORDER BY bt.transaction_date DESC
        """, (selected_cid, bank_cutoff)).fetchall()

    # ── Match requests to bank debits (amount exact, ±45 days) ───────────────
    matched_bank_ids: set[int] = set()
    ops_requests = []

    for r in reqs:
        try:
            req_amount = float(str(r["amount"]).replace(",", ""))
        except (ValueError, TypeError):
            req_amount = 0.0

        req_date = r["created_at"][:10] if r["created_at"] else ""

        bank_matches = []
        for tx in bank_debits:
            tx_amount = tx["debit_amount"] or 0.0
            if abs(tx_amount - req_amount) > 0.01:
                continue
            # Date window ±45 days
            try:
                rd = _date.fromisoformat(req_date)
                td = _date.fromisoformat(tx["transaction_date"][:10])
                if abs((td - rd).days) <= 45:
                    bank_matches.append(tx)
                    matched_bank_ids.add(tx["id"])
            except (ValueError, TypeError):
                pass

        ops_requests.append({
            "id": r["id"],
            "created_at": r["created_at"],
            "requester_name": r["requester_name"],
            "requester_username": r["requester_username"],
            "department_name": r["department_name"],
            "payee_name": r["payee_name"],
            "amount": req_amount,
            "status": r["status"],
            "is_matched": bool(bank_matches),
            "bank_matches": bank_matches,
            "duplicate_flag": False,  # set below
        })

    # ── Duplicate detection (same payee + amount, within 30 days) ────────────
    dup_key_map: dict[tuple, list] = defaultdict(list)
    for req in ops_requests:
        key = (req["payee_name"], round(req["amount"], 2))
        dup_key_map[key].append(req)

    duplicate_groups = []
    for (payee, amt), group in dup_key_map.items():
        if len(group) < 2:
            continue
        # Check if any two requests are within 30 days
        dates = []
        for req in group:
            try:
                dates.append(_date.fromisoformat(req["created_at"][:10]))
            except (ValueError, TypeError):
                dates.append(None)

        is_dup_group = False
        for i in range(len(dates)):
            for j in range(i + 1, len(dates)):
                if dates[i] and dates[j] and abs((dates[i] - dates[j]).days) <= 30:
                    is_dup_group = True
                    break
            if is_dup_group:
                break

        if is_dup_group:
            for req in group:
                req["duplicate_flag"] = True
            duplicate_groups.append({
                "payee": payee,
                "amount": amt,
                "requests": group,
            })

    # ── KPI counts ────────────────────────────────────────────────────────────
    matched_count = sum(1 for r in ops_requests if r["is_matched"])
    unmatched_count = sum(
        1 for r in ops_requests
        if not r["is_matched"] and r["status"] in ("approved", "paid")
    )
    pending_count = sum(
        1 for r in ops_requests
        if r["status"] not in ("approved", "paid", "cancelled")
    )
    duplicate_count = sum(1 for r in ops_requests if r["duplicate_flag"])

    # ── Bank debits with no matching request ──────────────────────────────────
    unmatched_bank = [tx for tx in bank_debits if tx["id"] not in matched_bank_ids]

    return templates.TemplateResponse(request, "ops_reconciliation.html", {
        "ops_requests": ops_requests,
        "matched_count": matched_count,
        "unmatched_count": unmatched_count,
        "pending_count": pending_count,
        "duplicate_count": duplicate_count,
        "unmatched_bank": unmatched_bank,
        "duplicate_groups": duplicate_groups,
        "companies": companies,
        "selected_company_id": selected_cid,
        "months": months,
    })


# ---------- Private helpers --------------------------------------------------

def _commission_history():
    """Return commission run history rows, empty list on error."""
    try:
        with db() as conn:
            return conn.execute(
                "SELECT * FROM commission_runs ORDER BY id DESC LIMIT 50"
            ).fetchall()
    except Exception:
        logger.exception("Failed to fetch commission run history")
        return []


def _commission_saved_carries():
    """Return saved carry-forward files list, newest first."""
    try:
        with db() as conn:
            return conn.execute(
                "SELECT * FROM commission_carry_files ORDER BY id DESC LIMIT 30"
            ).fetchall()
    except Exception:
        logger.exception("Failed to fetch saved carry-forward files")
        return []
